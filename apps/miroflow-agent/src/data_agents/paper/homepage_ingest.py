from __future__ import annotations

import hashlib
import json
import logging
import re
import time
import warnings
from html import escape
from io import BytesIO
from collections.abc import Mapping
from dataclasses import dataclass, replace
from decimal import Decimal
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Callable
from urllib.parse import unquote, urldefrag, urljoin, urlparse
from uuid import UUID

import httpx
from bs4 import BeautifulSoup

from ..normalization import build_stable_id
from ..professor.canonical_writer import (
    _upsert_professor_paper_link,
    upsert_source_page_for_url,
)
from ..professor.homepage_publications import (
    HomepagePublication,
    HomepagePublicationExtractionResult,
    _COMMON_ROMANIZED_CHINESE_SURNAMES,
    _is_suspicious_rule_publication,
    _looks_like_author_list,
    extract_publications_from_html,
    extract_publications_with_diagnostics_from_html,
)
from ..professor.homepage_source_filter import is_homepage_publication_ingest_url
from ..professor.name_selection import is_unsafe_professor_paper_evidence_identity
from ..professor.profile import (
    _find_sztu_fragment_profile_node,
    _scope_html_to_fragment_profile,
)
from ..storage.postgres.paper_full_text import (
    paper_full_text_exists,
    upsert_paper_full_text,
)
from ..storage.postgres.homepage_recursion_ledger import (
    upsert_homepage_recursion_ledger_entry,
)
from ..storage.postgres.pipeline_run import close_pipeline_run, open_pipeline_run
from ..storage.postgres.title_resolution_cache import PostgresTitleResolutionCache
from .canonical_writer import upsert_paper
from .full_text_fetcher import fetch_and_extract_full_text
from .homepage_http import fetch_homepage_html
from .quality_promotion import NEEDS_ENRICHMENT
from .title_quality import is_plausible_paper_title
from .title_resolver import ResolvedPaper, resolve_paper_by_title

logger = logging.getLogger(__name__)

_DEFAULT_PUBLICATION_EXTRACTOR = extract_publications_from_html
_DRY_RUN_SENTINEL_RUN_ID = UUID("00000000-0000-0000-0000-000000000000")
_AUTHOR_NAME_MATCH_SCORE = Decimal("1.0")
_LINK_MATCH_REASON = "homepage_title_resolution"
_LINK_MATCH_REASON_PAGE_ONLY = "prof_page_declaration"
_PROF_PAGE_ONLY_SOURCE = "prof_page_only"
_TIER2_PAGE_ROLES = frozenset({"official_profile", "official_publication_page"})
_TIER3_PAGE_ROLES = frozenset({"personal_homepage", "lab_homepage"})
_FRAGMENT_PROFILE_PAGE_ROLES = frozenset(
    {"official_profile", "official_publication_page"}
)
_OWNED_HOMEPAGE_PAGE_ROLES = (
    "official_publication_page",
    "personal_homepage",
    "lab_homepage",
)
_DEFAULT_PROF_PAGE_PDF_FETCH_CAP = 20
_BULK_EXTERNAL_RESOLUTION_MAX_PUBLICATIONS = 80
_BULK_EXTERNAL_RESOLUTION_MAX_PER_PROFESSOR = 12
_MAX_SECOND_HOP_PUBLICATION_PAGES = 3
_SZU_BIGDATA_PUBLICATION_SECOND_HOP_MAX_PAGES = 8
_SZU_BIGDATA_ATTACHMENT_TIMEOUT = httpx.Timeout(
    connect=5.0,
    read=20.0,
    write=5.0,
    pool=5.0,
)
_SZU_BIGDATA_ATTACHMENT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
}
_BULK_TITLE_RESOLUTION_TIMEOUT = httpx.Timeout(
    connect=2.0,
    read=3.0,
    write=2.0,
    pool=2.0,
)
_AUTHOR_INITIAL_HINT_RE = re.compile(r"\b[A-Z]\.")
_AUTHOR_LIST_DELIMITER_RE = re.compile(r"\s*(?:[,，;；]|\band\b|&)\s*", re.I)
_CJK_RE = re.compile(r"[\u3400-\u9fff]")
_JCR_QUARTILE_LABEL_RE = re.compile(r"^(?:SCI\s*)?JCR\s*Q[1-4]$", re.I)
_JCR_METRIC_LABEL_RE = re.compile(
    r"^(?:SCI\s*)?JCR\s*:?\s*Q[1-4](?:\s*/?\s*IF\s*:?\s*[\d.]+)?$",
    re.I,
)
_MONTH_DAY_LABEL_RE = re.compile(
    r"^(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
    r"Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|Nov(?:ember)?|"
    r"Dec(?:ember)?)\.?\s+\d{1,2}$",
    re.I,
)
_DOI_URL_OR_VALUE_RE = re.compile(
    r"(?:https?://(?:dx\.)?doi\.org/|doi\s*[:：]\s*)?"
    r"(?P<doi>10\.\d{4,9}/[^\s\]）)>\"'，,;；。]+)",
    re.I,
)
_QUALIFICATION_TITLE_RE = re.compile(
    r"(注册会计师|资格考试|全科通过|可豁免|\bACCA\b)",
    re.I,
)
_PROCEEDINGS_LABEL_ONLY_TITLE_RE = re.compile(
    r"^(?:in\s+)?proceedings\s+of\s+"
    r"(?:(?:the|a)\s+)?(?:\d{1,2}(?:st|nd|rd|th)\s+)?"
    r".*\b(?:conference|congress|symposium|workshop|aaai|cvpr|eccv|iccv|"
    r"icml|iclr|ijcai|acl|emnlp|kdd|usenix|ieee|acm|springer)\b.*"
    r"(?:\b(?:19|20)\d{2}\b|\([A-Z0-9][A-Z0-9'&/ .-]{1,40}\))\.?$",
    re.I,
)
_IN_VENUE_LABEL_ONLY_TITLE_RE = re.compile(
    r"^in\s+(?:"
    r"(?:ieee|acm|elsevier|springer|nature|science|cell)\s+"
    r")?"
    r"(?:transactions?|journal|letters|conference|congress|symposium|workshop|"
    r"proceedings?|neurocomputing|chemistry|materials|communications?|"
    r"signal\s+processing|bioinformatics|automatica|robotics|pattern\s+recognition)"
    r"\b.*"
    r"(?:\([A-Z0-9][A-Z0-9'&/ .-]{1,40}\))?\.?$",
    re.I,
)
_VENUE_COUNT_METRIC_ONLY_TITLE_RE = re.compile(
    r"^(?:[A-Z][A-Z0-9&./-]{1,12}\s*[*×x]\s*\d+\s*[,;，；]\s*)+"
    r"[A-Z][A-Z0-9&./-]{1,12}\s*[*×x]\s*\d+\.?$",
    re.I,
)
_CONNECTIVE_AUTHOR_FRAGMENT_TITLE_RE = re.compile(
    r"^(?:and)\s+[A-Z][A-Za-z'’.-]+"
    r"(?:\s+[A-Z][A-Za-z'’.-]+)?(?:\s+[A-Za-z]\.?){0,3}[*#†‡]*$"
    r"|^(?:etc)\.?\s+[A-Z][A-Za-z'’.-]+"
    r"(?:\s+[A-Za-z]\.?){0,3}(?:\s+[A-Z][A-Za-z'’.-]+)?[*#†‡]*$",
    re.I,
)
_LEADING_CONTRIBUTION_MARKER_TITLE_RE = re.compile(r"^\s*[*#†‡]{1,4}\s+")
_LEADING_ETC_TITLE_PREFIX_RE = re.compile(r"^etc\.?\s+[A-Z].{20,}$", re.I)
_TRAILING_REFERENCE_LINK_LABELS_RE = re.compile(
    r"(?:\s*\[(?:paper|code|pdf|doi|arxiv|project|page|slides|bibtex|link|"
    r"video|dataset)\]\s*)+$",
    re.I,
)
_TRAILING_REFERENCE_VENUE_MARKER_RE = re.compile(
    r"\s+(?:"
    r"(?:in\s+)?(?:proceedings\s+of\s+)?(?:the\s+)?"
    r"(?:European\s+Conference\s+on\s+Computer\s+Vision|"
    r"IEEE/?CVF\s+Conference\s+on\s+Computer\s+Vision\s+and\s+Pattern\s+Recognition|"
    r"Conference\s+on\s+Computer\s+Vision\s+and\s+Pattern\s+Recognition|"
    r"International\s+Conference\s+on\s+Machine\s+Learning|"
    r"International\s+Conference\s+on\s+Learning\s+Representations|"
    r"Advances\s+in\s+Neural\s+Information\s+Processing\s+Systems|"
    r"AAAI\s+Conference\s+on\s+Artificial\s+Intelligence|"
    r"International\s+Joint\s+Conference\s+on\s+Artificial\s+Intelligence|"
    r"Association\s+for\s+Computational\s+Linguistics|"
    r"Empirical\s+Methods\s+in\s+Natural\s+Language\s+Processing|"
    r"NeurIPS|ICML|ICLR|CVPR|ECCV|ICCV|AAAI|ACL|EMNLP|KDD|IJCAI|WWW))"
    r"\b.*?\b(?:19|20)\d{2}\b\.?$",
    re.I,
)
_SPACED_NAME_FRAGMENT_TITLE_RE = re.compile(
    r"^[A-Z][A-Za-z'’.-]{1,40}\s+(?:[A-Za-z]\s+){2,5}[A-Za-z][*#†‡]*$"
)
_PUBLICATION_MARKER_LEGEND_TITLE_RE = re.compile(
    r"^in\s+publications?\s+marked\s+with\b.*\bauthors?\s+are\s+ordered\b",
    re.I,
)
_DATE_VOLUME_METADATA_TITLE_RE = re.compile(
    r"^(?:19|20)\d{2}\s+"
    r"(?:jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)"
    r"[a-z]*\.?\s+\d{1,2}\s*;"
    r"\s*\d+[A-Za-z]?(?:\s+Suppl\s+\d+)?"
    r"\s*:\s*[A-Za-z0-9.-]+"
    r"(?:\.?\s+doi\s*:\s*10\.\d{4,9}/\S+)?\.?$",
    re.I,
)
_BIBLIOGRAPHIC_FRAGMENT_TITLE_RE = re.compile(
    r"^(?:pp?\.?\s*)?\d+\s*[-–]\s*\d+$"
    r"|^(?:年|[,，])\s*[,，]?\s*ISBN\b.*\bPage\b"
    r"|^.*\bISBN\b.*\bPage\b.*$"
    r"|^\d{1,5}\s*[,，]\s*\d+\s*[-–]\s*\d+.*$"
    r"|^\d{1,5}\s*[,，]\s*\d{1,6}\s*\.?\s*\[?\s*doi\s*\]?$"
    r"|^\d{1,5}\s*[,，]\s*e[0-9A-Za-z]+(?:\s+[0-9A-Za-z]+)?$"
    r"|^[A-Za-z]{1,12}\.?\s*[,，]\s*(?:19|20)\d{2}\s*[,，]\s*\d{1,5}\s*[,，]\s*\d+\s*[-–]\s*\d+$"
    r"|^[A-Za-z]{1,12}\.?\s+(?:19|20)\d{2}\s*[,，]\s*\d{1,5}\s*[,，]\s*\d+\s*[-–]\s*\d+$",
    re.I,
)
_ELLIPSIS_AUTHOR_FRAGMENT_RE = re.compile(r"^(?:…|\.\.\.).*(?:[,，].*){2,}$")
_LOWERCASE_CONTINUATION_FRAGMENT_RE = re.compile(
    r"^[a-z]\s+[a-z][A-Za-z-]+(?:\s+[A-Za-z][A-Za-z-]+){2,}$"
)
_AUTHOR_YEAR_FRAGMENT_RE = re.compile(
    r"^[A-Z][A-Za-z'’-]{1,40}[*#†‡]?\s*[\(（](?:19|20)\d{2}[\)）]$"
)
_CONCATENATED_AUTHOR_FRAGMENT_RE = re.compile(r"^[A-Z][a-z]{1,24}[A-Z][a-z]{1,24}$")
_SEMICOLON_SURNAME_INITIAL_AUTHOR_FRAGMENT_RE = re.compile(
    r"^[A-Z][A-Za-z?'’-]{1,40}\s*[,，]\s*(?:[A-Z]\.?\s*){1,5}"
    r"[;；]\s*[A-Z][A-Za-z?'’-]{1,40}$"
)
_NO_SPACE_AND_AUTHOR_FRAGMENT_RE = re.compile(
    r"^and[A-Z][A-Za-z'’-]+(?:\s+[A-Z]\.?)?(?:\s+[A-Z][A-Za-z'’-]+)?[*#†‡]*$"
)
_PDF_FETCH_CAP_ERRORS = frozenset(
    {
        "pdf_too_large",
        "timeout",
        "pdf_content_type_disallowed",
        "redirect_cap_exceeded",
    }
)
_NON_PUBLICATION_LABEL_TITLES = frozenset(
    {
        "book chapters",
        "degree source",
        "in chinese",
        "invited talks",
        "manufacturing",
        "social networks",
        "transportation and disaster management",
        "healthcare and service systems",
    }
)
_KNOWN_VENUE_ONLY_TITLES = frozenset(
    {
        "angew. chem",
        "acs applied energy materials",
        "acs applied materials & interfaces",
        "acs energy letters",
        "advanced energy materials",
        "advanced functional materials",
        "advanced materials",
        "angew chem int edit",
        "applied health economics and health policy",
        "cell reports physical science",
        "chemical communications",
        "chemical engineering journal",
        "energy & environmental science",
        "energy storage materials",
        "journal of the american chemical society",
        "nano letters",
        "nature communications",
        "periodica polytechnica architecture",
        "personal and ubiquitous computing",
        "synfacts highlights",
        "the journal of physical chemistry letters",
        "自然 · 通讯",
    }
)
_SECOND_HOP_PUBLICATION_LINK_KEYWORDS = (
    "publication",
    "publications",
    "paper",
    "papers",
    "research-output",
    "research outputs",
    "selected publications",
    "selected papers",
    "论文",
    "发表论文",
    "学术成果",
    "科研成果",
)
_COUNT_ONLY_PUBLICATION_CLAIM_RE = re.compile(
    r"(?P<claim>"
    r"[^。；;\n]{0,40}"
    r"(?:发表|收录|刊发|出版|累计|共|已|在国内外期刊)"
    r"[^。；;\n]{0,40}"
    r"(?:论文|SCI|SSCI|EI)"
    r"[^。；;\n]{0,20}"
    r"\d+\s*(?:余|多|来)?\s*篇"
    r"[^。；;\n]{0,40}"
    r")",
    re.I,
)
_COUNT_ONLY_PUBLICATION_TEXT_SELECTORS = (
    ".v_news_content",
    ".detailtext11",
    ".newsdetail11",
    "article",
    "main",
)
_NON_HTML_SECOND_HOP_EXTENSIONS = (
    ".pdf",
    ".doc",
    ".docx",
    ".ppt",
    ".pptx",
    ".xls",
    ".xlsx",
    ".zip",
    ".rar",
)


@dataclass(frozen=True)
class _HomepageLink:
    href: str
    text: str
    title: str | None


class _HomepagePublicationLinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.links: list[_HomepageLink] = []
        self._active_href: str | None = None
        self._active_title: str | None = None
        self._active_text: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "a":
            return
        attr_map = {key.lower(): value for key, value in attrs if value is not None}
        self._active_href = attr_map.get("href")
        self._active_title = attr_map.get("title")
        self._active_text = []

    def handle_data(self, data: str) -> None:
        if self._active_href is not None:
            self._active_text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() != "a" or self._active_href is None:
            return
        self.links.append(
            _HomepageLink(
                href=self._active_href,
                text=" ".join(part.strip() for part in self._active_text).strip(),
                title=self._active_title,
            )
        )
        self._active_href = None
        self._active_title = None
        self._active_text = []


class _SkipCurrentProfessor(Exception):
    """Internal sentinel for handled per-professor early exits."""


def _synthesize_page_only_resolution(
    publication,
    *,
    canonical_name: str,
) -> ResolvedPaper:
    """Build a synthetic ResolvedPaper from prof-page data only.

    Used when external title resolution (Crossref / OpenAlex / S2) fails
    to find the publication — typically the preprint case (paper recently
    accepted but not yet indexed in external DBs). Per Paper Review
    §3.1 P4 and OpenSpec change `prof-paper-patent-from-page-flow` spec
    Requirement "Preprint listed on professor page", the system MUST
    still create a paper canonical record with page-only data and let
    enrichment fill in DOI / abstract / etc. on the next cron run.
    """
    authors_text = (publication.authors_text or "").strip()
    if authors_text:
        authors = tuple(_split_page_authors(authors_text)) or (
            f"{canonical_name} et al.",
        )
    else:
        authors = (f"{canonical_name} et al.",)
    return ResolvedPaper(
        title=publication.clean_title,
        doi=_homepage_publication_doi(publication),
        openalex_id=None,
        arxiv_id=None,
        abstract=None,
        pdf_url=getattr(publication, "pdf_url", None),
        authors=authors,
        year=publication.year,
        venue=publication.venue_text,
        match_confidence=1.0,
        match_source=_PROF_PAGE_ONLY_SOURCE,
    )


def _homepage_publication_doi(publication) -> str | None:
    for value in (
        getattr(publication, "source_anchor", None),
        getattr(publication, "raw_title", None),
    ):
        if not isinstance(value, str):
            continue
        match = _DOI_URL_OR_VALUE_RE.search(value)
        if match:
            return match.group("doi").rstrip(" .。")
    return None


def _split_page_authors(authors_text: str) -> list[str]:
    """Best-effort author split from prof-page free-text. Conservative;
    keeps original text if no clear delimiter detected."""
    candidates = [
        item.strip()
        for item in authors_text.replace(";", ",").replace("、", ",").split(",")
    ]
    return [c for c in candidates if c]


def _should_skip_external_title_resolution(
    publication,
    *,
    publication_count: int,
) -> bool:
    return False


def _attach_professor_page_pdf_url(
    resolved: ResolvedPaper,
    publication,
) -> ResolvedPaper:
    pdf_url = getattr(publication, "pdf_url", None)
    if not pdf_url:
        return resolved
    return replace(resolved, pdf_url=pdf_url)


def _homepage_evidence_source_type(prof: dict[str, Any]) -> str | None:
    page_role = str(prof.get("homepage_page_role") or "").strip()
    if page_role in _TIER2_PAGE_ROLES:
        return "prof_homepage_tier2"
    if page_role in _TIER3_PAGE_ROLES:
        return "prof_homepage_tier3"
    return None


def _publication_evidence_page_id(
    conn,
    *,
    prof: dict[str, Any],
    publication: HomepagePublication,
) -> Any | None:
    source_url = _normalize_source_url(getattr(publication, "source_url", None))
    homepage_url = _normalize_source_url(prof.get("homepage_url"))
    homepage_page_id = prof.get("homepage_page_id")
    if source_url and homepage_url and source_url == homepage_url and homepage_page_id:
        return homepage_page_id
    if not source_url:
        return homepage_page_id

    professor_id = str(prof["professor_id"])
    row = conn.execute(
        """
        SELECT page_id
        FROM source_page
        WHERE url = %s
          AND owner_scope_kind = 'professor'
          AND owner_scope_ref = %s
          AND page_role IN (%s, %s, %s, %s)
        LIMIT 1
        """,
        (
            source_url,
            professor_id,
            "official_profile",
            "official_publication_page",
            "personal_homepage",
            "lab_homepage",
        ),
    ).fetchone()
    if row is None:
        return None
    return _row_value(row, "page_id")


def _ensure_publication_evidence_page_id(
    conn,
    *,
    prof: dict[str, Any],
    publication: HomepagePublication,
    run_id: UUID,
) -> Any | None:
    page_id = _publication_evidence_page_id(conn, prof=prof, publication=publication)
    if page_id is not None:
        return page_id

    source_url = _normalize_source_url(getattr(publication, "source_url", None))
    if not source_url:
        return None
    homepage_url = _normalize_source_url(prof.get("homepage_url"))
    if homepage_url and source_url == homepage_url:
        return None
    if (
        homepage_url
        and not _is_same_personal_site_root(homepage_url, source_url)
        and not _is_szu_bigdata_publication_info_link(homepage_url, source_url)
    ):
        return prof.get("homepage_page_id")

    page_role = _publication_source_page_role(prof, source_url=source_url)
    if page_role is None:
        return None

    return upsert_source_page_for_url(
        conn,
        url=source_url,
        page_role=page_role,
        owner_scope_kind="professor",
        owner_scope_ref=str(prof["professor_id"]),
        is_official_source=page_role in _TIER2_PAGE_ROLES,
        run_id=run_id,
    )


def _publication_source_page_role(
    prof: dict[str, Any],
    *,
    source_url: str,
) -> str | None:
    homepage_role = str(prof.get("homepage_page_role") or "").strip()
    homepage_url = _normalize_source_url(prof.get("homepage_url"))
    if homepage_url and source_url == homepage_url:
        return homepage_role if homepage_role in _TIER2_PAGE_ROLES | _TIER3_PAGE_ROLES else None
    if homepage_role in {"official_profile", "official_publication_page"}:
        return "official_publication_page"
    if homepage_role in {"personal_homepage", "lab_homepage"}:
        return homepage_role
    return None


def _record_homepage_recursion_page_outcomes(
    conn,
    *,
    run_id: UUID,
    prof: dict[str, Any],
    outcomes: tuple[_HomepagePublicationPageOutcome, ...],
) -> None:
    professor_id = str(prof["professor_id"])
    homepage_url = _normalize_source_url(prof.get("homepage_url"))
    parent_source_page_id = prof.get("homepage_page_id")
    for outcome in outcomes:
        source_url = _normalize_source_url(outcome.page_url) or outcome.page_url
        source_page_id = (
            parent_source_page_id
            if homepage_url and source_url == homepage_url
            else None
        )
        page_role = _publication_source_page_role(prof, source_url=source_url)
        if source_page_id is not None:
            page_role = str(prof.get("homepage_page_role") or page_role or "unknown")
        if page_role is None:
            page_role = "unknown"
        try:
            upsert_homepage_recursion_ledger_entry(
                conn,
                run_id=run_id,
                professor_id=professor_id,
                parent_source_page_id=parent_source_page_id,
                source_page_id=source_page_id,
                url=source_url,
                page_role=page_role,
                discovery_source="homepage_paper_ingest",
                recursion_depth=outcome.recursion_depth,
                status=outcome.status,
                skip_reason=outcome.skip_reason,
                fetch_error_type=outcome.fetch_error_type,
                fetch_error_message=outcome.fetch_error_message,
                publications_extracted=outcome.publications_extracted,
                sections_detected=outcome.sections_detected,
                heading_texts=outcome.heading_texts,
            )
        except Exception as exc:
            logger.warning(
                "Failed to record homepage recursion ledger for %s: %s",
                source_url,
                exc,
            )


def _normalize_source_url(value: object) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip().rstrip("/")
    return normalized or None


def _row_value(row: object, column: str, index: int = 0) -> Any:
    if isinstance(row, Mapping):
        return row[column]
    return row[index]  # type: ignore[index]


def _personal_site_root_path(homepage_url: str) -> str:
    parsed = urlparse(homepage_url)
    path = parsed.path or "/"
    if path.endswith("/"):
        return path
    last_segment = path.rsplit("/", 1)[-1]
    if "." in last_segment:
        root = path.rsplit("/", 1)[0]
        return f"{root}/" if root else "/"
    return f"{path.rstrip('/')}/"


def _is_same_personal_site_root(homepage_url: str, candidate_url: str) -> bool:
    homepage = urlparse(homepage_url)
    candidate = urlparse(candidate_url)
    if candidate.scheme not in {"http", "https"}:
        return False
    if (candidate.hostname or "").lower() != (homepage.hostname or "").lower():
        return False
    root_path = _personal_site_root_path(homepage_url)
    if root_path == "/":
        return True
    candidate_path = candidate.path or "/"
    return candidate_path == root_path.rstrip("/") or candidate_path.startswith(root_path)


def _looks_like_publication_page_link(link: _HomepageLink, absolute_url: str) -> bool:
    parsed = urlparse(absolute_url)
    path = parsed.path.lower().replace("_", "-")
    if path.endswith(_NON_HTML_SECOND_HOP_EXTENSIONS):
        return False
    path_parts = [part for part in path.split("/") if part]
    stem_parts = [part.rsplit(".", 1)[0] for part in path_parts]
    haystacks = (
        path,
        (link.text or "").lower().replace("_", "-"),
        (link.title or "").lower().replace("_", "-"),
    )
    if any(part in {"pub", "pubs"} for part in stem_parts):
        return True
    return any(
        keyword in haystack
        for haystack in haystacks
        for keyword in _SECOND_HOP_PUBLICATION_LINK_KEYWORDS
    )


def _discover_second_hop_publication_page_candidates(
    html: str,
    *,
    page_url: str,
    max_pages: int | None = None,
) -> list[_SecondHopPublicationPageCandidate]:
    if max_pages is None:
        max_pages = _second_hop_publication_page_limit(page_url)
    parser = _HomepagePublicationLinkParser()
    try:
        parser.feed(html)
    except Exception:
        return []

    discovered: list[_SecondHopPublicationPageCandidate] = []
    seen = {page_url.rstrip("/")}
    for link in parser.links:
        href = (link.href or "").strip()
        if not href or href.startswith(("#", "mailto:", "javascript:", "tel:")):
            continue
        absolute_url = urldefrag(urljoin(page_url, href))[0]
        normalized = absolute_url.rstrip("/")
        if normalized in seen:
            continue
        if not _looks_like_publication_page_link(link, absolute_url):
            continue
        seen.add(normalized)
        if not _is_same_personal_site_root(
            page_url,
            absolute_url,
        ) and not _is_szu_bigdata_publication_info_link(
            page_url,
            absolute_url,
        ):
            discovered.append(
                _SecondHopPublicationPageCandidate(
                    page_url=absolute_url,
                    status="skipped",
                    skip_reason="outside_personal_site_root",
                )
            )
            continue
        discovered.append(_SecondHopPublicationPageCandidate(page_url=absolute_url))
        active_count = sum(1 for item in discovered if item.status == "candidate")
        if active_count >= max_pages:
            break
    return discovered


def _discover_second_hop_publication_page_urls(
    html: str,
    *,
    page_url: str,
    max_pages: int | None = None,
) -> list[str]:
    return [
        candidate.page_url
        for candidate in _discover_second_hop_publication_page_candidates(
            html,
            page_url=page_url,
            max_pages=max_pages,
        )
        if candidate.status == "candidate"
    ]


def _second_hop_publication_page_limit(page_url: str) -> int:
    parsed = urlparse(page_url)
    if (
        (parsed.hostname or "").casefold() == "bigdata.szu.edu.cn"
        and (parsed.path or "").startswith("/kycg/")
    ):
        return _SZU_BIGDATA_PUBLICATION_SECOND_HOP_MAX_PAGES
    return _MAX_SECOND_HOP_PUBLICATION_PAGES


def _is_szu_bigdata_publication_info_link(page_url: str, candidate_url: str) -> bool:
    page = urlparse(page_url)
    candidate = urlparse(candidate_url)
    if (page.hostname or "").casefold() != "bigdata.szu.edu.cn":
        return False
    if (candidate.hostname or "").casefold() != "bigdata.szu.edu.cn":
        return False
    source_path = page.path or "/"
    candidate_path = candidate.path or "/"
    if not source_path.startswith("/kycg/"):
        return False
    return bool(re.fullmatch(r"/info/\d+/\d+\.htm", candidate_path))


def _is_szu_bigdata_publication_info_page(page_url: str) -> bool:
    parsed = urlparse(page_url)
    return (
        (parsed.hostname or "").casefold() == "bigdata.szu.edu.cn"
        and re.fullmatch(r"/info/\d+/\d+\.htm", parsed.path or "") is not None
    )


def _augment_szu_bigdata_publication_attachment_text(
    html: str,
    *,
    page_url: str,
) -> str:
    if not _is_szu_bigdata_publication_info_page(page_url):
        return html

    parser = _HomepagePublicationLinkParser()
    try:
        parser.feed(html)
    except Exception:
        return html

    attachment_sections: list[str] = []
    seen_urls: set[str] = set()
    for link in parser.links:
        href = (link.href or "").strip()
        label = (link.text or link.title or "").strip()
        if not href:
            continue
        attachment_url = urljoin(page_url, href)
        if attachment_url in seen_urls:
            continue
        lower_label = label.casefold()
        lower_path = urlparse(attachment_url).path.casefold()
        if not (
            lower_label.endswith((".xlsx", ".pdf"))
            or lower_path.endswith((".xlsx", ".pdf"))
            or (
                "download.jsp" in lower_path
                and lower_label.endswith((".xlsx", ".pdf"))
            )
        ):
            continue
        seen_urls.add(attachment_url)
        try:
            attachment_bytes = _fetch_szu_bigdata_attachment_bytes(
                attachment_url,
                referer_url=page_url,
            )
            section_html = _szu_bigdata_attachment_publication_section_html(
                attachment_bytes,
                filename=label,
            )
        except Exception as exc:  # noqa: BLE001 - attachment enrichment is best effort.
            logger.warning(
                "Failed to extract SZU BigData publication attachment %s from %s: %s",
                attachment_url,
                page_url,
                exc,
            )
            continue
        if section_html:
            attachment_sections.append(section_html)

    if not attachment_sections:
        return html
    return f"{html}\n<section><h2>附件论文</h2>{''.join(attachment_sections)}</section>"


def _fetch_szu_bigdata_attachment_bytes(url: str, *, referer_url: str) -> bytes:
    headers = {**_SZU_BIGDATA_ATTACHMENT_HEADERS, "Referer": referer_url}
    with httpx.Client(
        trust_env=False,
        follow_redirects=True,
        timeout=_SZU_BIGDATA_ATTACHMENT_TIMEOUT,
        headers=headers,
    ) as client:
        response = client.get(url)
        response.raise_for_status()
        return response.content


def _szu_bigdata_attachment_publication_section_html(
    content: bytes,
    *,
    filename: str,
) -> str | None:
    lower_filename = filename.casefold()
    if lower_filename.endswith(".xlsx"):
        return _szu_bigdata_xlsx_publication_table_html(content)
    if lower_filename.endswith(".pdf"):
        return _szu_bigdata_pdf_publication_text_html(content)
    if content.startswith(b"PK\x03\x04"):
        return _szu_bigdata_xlsx_publication_table_html(content)
    if content.startswith(b"%PDF"):
        return _szu_bigdata_pdf_publication_text_html(content)
    return None


def _szu_bigdata_xlsx_publication_table_html(content: bytes) -> str | None:
    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    tables: list[str] = []
    for sheet in workbook.worksheets:
        rows: list[str] = []
        for row in sheet.iter_rows(values_only=True):
            values = [
                _normalize_attachment_cell_value(value)
                for value in row
                if _normalize_attachment_cell_value(value)
            ]
            if not values:
                continue
            cells = "".join(f"<td>{escape(value)}</td>" for value in values)
            rows.append(f"<tr>{cells}</tr>")
        if rows:
            tables.append(f"<table>{''.join(rows)}</table>")
    return "".join(tables) or None


def _szu_bigdata_pdf_publication_text_html(content: bytes) -> str | None:
    from pdfminer.high_level import extract_text

    text = _normalize_attachment_cell_value(extract_text(BytesIO(content)))
    if not text:
        return None
    return f"<p>{escape(text)}</p>"


def _normalize_attachment_cell_value(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("_x000D_", " ")
    return re.sub(r"\s+", " ", text).strip()


def _extract_publications_from_homepage_source_pages(
    html: str,
    *,
    page_url: str,
    publication_extractor: Callable[..., list[HomepagePublication]],
    use_rule_diagnostics: bool,
) -> _HomepagePublicationExtractionReport:
    publications, zero_extraction_pages = _extract_publications_from_single_source_page(
        html,
        page_url=page_url,
        publication_extractor=publication_extractor,
        use_rule_diagnostics=use_rule_diagnostics,
    )
    page_outcomes = [
        _HomepagePublicationPageOutcome(
            page_url=page_url,
            status="zero_extraction" if zero_extraction_pages else "processed",
            recursion_depth=0,
            publications_extracted=len(publications),
            sections_detected=(
                zero_extraction_pages[0].sections_detected
                if zero_extraction_pages
                else 0
            ),
            heading_texts=(
                zero_extraction_pages[0].heading_texts
                if zero_extraction_pages
                else ()
            ),
        )
    ]
    seen_sources = {page_url.rstrip("/")}
    for candidate in _discover_second_hop_publication_page_candidates(
        html,
        page_url=page_url,
    ):
        if candidate.status == "skipped":
            page_outcomes.append(
                _HomepagePublicationPageOutcome(
                    page_url=candidate.page_url,
                    status="skipped",
                    recursion_depth=1,
                    skip_reason=candidate.skip_reason,
                )
            )
            continue
        source_url = candidate.page_url
        normalized = source_url.rstrip("/")
        if normalized in seen_sources:
            continue
        seen_sources.add(normalized)
        try:
            source_html = fetch_homepage_html(source_url)
        except (
            httpx.HTTPStatusError,
            httpx.TransportError,
        ) as exc:
            logger.warning(
                "Second-hop homepage publication page fetch failed for %s: %s",
                source_url,
                exc,
            )
            page_outcomes.append(
                _HomepagePublicationPageOutcome(
                    page_url=source_url,
                    status="fetch_failed",
                    recursion_depth=1,
                    fetch_error_type=type(exc).__name__,
                    fetch_error_message=str(exc),
                )
            )
            continue
        page_publications, page_zero_extraction = (
            _extract_publications_from_single_source_page(
                source_html,
                page_url=source_url,
                publication_extractor=publication_extractor,
                use_rule_diagnostics=use_rule_diagnostics,
            )
        )
        publications.extend(page_publications)
        zero_extraction_pages.extend(page_zero_extraction)
        page_outcomes.append(
            _HomepagePublicationPageOutcome(
                page_url=source_url,
                status="zero_extraction" if page_zero_extraction else "processed",
                recursion_depth=1,
                publications_extracted=len(page_publications),
                sections_detected=(
                    page_zero_extraction[0].sections_detected
                    if page_zero_extraction
                    else 0
                ),
                heading_texts=(
                    page_zero_extraction[0].heading_texts
                    if page_zero_extraction
                    else ()
                ),
            )
        )
    return _HomepagePublicationExtractionReport(
        publications=_dedupe_homepage_publications(publications),
        zero_extraction_pages=tuple(zero_extraction_pages),
        page_outcomes=tuple(page_outcomes),
    )


def _extract_publications_from_single_source_page(
    html: str,
    *,
    page_url: str,
    publication_extractor: Callable[..., list[HomepagePublication]],
    use_rule_diagnostics: bool,
) -> tuple[list[HomepagePublication], list[_ZeroExtractionPublicationPage]]:
    html = _augment_szu_bigdata_publication_attachment_text(html, page_url=page_url)
    if not use_rule_diagnostics:
        publications = publication_extractor(html, page_url=page_url)
        if publications:
            return publications, []
        result = extract_publications_with_diagnostics_from_html(
            html,
            page_url=page_url,
        )
        if result.publications:
            return result.publications, []
        return publications, _zero_extraction_pages_from_result(page_url, result)

    result = extract_publications_with_diagnostics_from_html(html, page_url=page_url)
    return result.publications, _zero_extraction_pages_from_result(page_url, result)


def _filter_publications_for_professor_evidence(
    prof: dict[str, Any],
    publications: list[HomepagePublication],
) -> list[HomepagePublication]:
    if not publications:
        return []
    author_variants = _professor_author_name_variants(prof)
    filtered: list[HomepagePublication] = []
    for publication in publications:
        if not _requires_author_match_for_page_evidence(publication):
            filtered.append(publication)
            continue
        if _publication_authors_match_professor(publication, author_variants):
            filtered.append(publication)
    return filtered


def _requires_author_match_for_page_evidence(
    publication: HomepagePublication,
) -> bool:
    source_url = _normalize_source_url(getattr(publication, "source_url", None))
    if not source_url:
        return False
    parsed = urlparse(source_url)
    if (parsed.hostname or "").casefold() != "bigdata.szu.edu.cn":
        return False
    path = parsed.path or "/"
    return path.startswith("/kycg/") or re.fullmatch(r"/info/1016/\d+\.htm", path) is not None


def _publication_authors_match_professor(
    publication: HomepagePublication,
    author_variants: tuple[str, ...],
) -> bool:
    authors_text = getattr(publication, "authors_text", None)
    if not authors_text or not author_variants:
        return False
    normalized_authors = _normalize_author_match_text(authors_text)
    if not normalized_authors:
        return False
    return any(
        _normalize_author_match_text(variant) in normalized_authors
        for variant in author_variants
        if _normalize_author_match_text(variant)
    )


def _professor_author_name_variants(prof: dict[str, Any]) -> tuple[str, ...]:
    raw_values: list[object] = [
        prof.get("canonical_name"),
        prof.get("canonical_name_en"),
    ]
    aliases = prof.get("aliases")
    if isinstance(aliases, (list, tuple, set)):
        raw_values.extend(aliases)
    elif isinstance(aliases, str):
        raw_values.append(aliases)

    variants: list[str] = []
    seen: set[str] = set()
    for value in raw_values:
        if not isinstance(value, str):
            continue
        normalized = re.sub(r"\s+", " ", value).strip()
        if not normalized:
            continue
        key = normalized.casefold()
        if key in seen:
            continue
        seen.add(key)
        variants.append(normalized)
    return tuple(variants)


def _normalize_author_match_text(value: str) -> str:
    return re.sub(r"[\W_]+", "", value, flags=re.UNICODE).casefold()


def _zero_extraction_pages_from_result(
    page_url: str,
    result: HomepagePublicationExtractionResult,
) -> list[_ZeroExtractionPublicationPage]:
    if result.sections_detected <= 0 or result.publications:
        return []
    return [
        _ZeroExtractionPublicationPage(
            page_url=page_url,
            sections_detected=result.sections_detected,
            heading_texts=tuple(
                _truncate_issue_detail_text(text)
                for text in result.heading_texts[:5]
                if text
            ),
        )
    ]


def _truncate_issue_detail_text(text: str, *, limit: int = 160) -> str:
    normalized = re.sub(r"\s+", " ", str(text)).strip()
    if len(normalized) <= limit:
        return normalized
    return normalized[: limit - 1].rstrip() + "…"


def _dedupe_homepage_publications(
    publications: list[HomepagePublication],
) -> list[HomepagePublication]:
    seen: set[tuple[str, int | None]] = set()
    deduped: list[HomepagePublication] = []
    for publication in publications:
        publication = _normalize_homepage_publication_for_ingest(publication)
        key = (_normalize_guard_title(publication.clean_title).casefold(), publication.year)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(publication)
    return deduped


def _filter_new_publications_for_professor(
    professor_id: str,
    publications: list[HomepagePublication],
    seen_publication_keys_by_professor: dict[str, set[tuple[str, int | None]]],
) -> list[HomepagePublication]:
    seen = seen_publication_keys_by_professor.setdefault(professor_id, set())
    filtered: list[HomepagePublication] = []
    for publication in publications:
        publication = _normalize_homepage_publication_for_ingest(publication)
        key = (_normalize_guard_title(publication.clean_title).casefold(), publication.year)
        if key in seen:
            continue
        seen.add(key)
        filtered.append(publication)
    return filtered


def _normalize_homepage_publication_for_ingest(
    publication: HomepagePublication,
) -> HomepagePublication:
    clean_title = _normalize_guard_title(publication.clean_title)
    clean_title = _LEADING_CONTRIBUTION_MARKER_TITLE_RE.sub("", clean_title).strip()
    clean_title = _normalize_reference_style_title(clean_title)
    if clean_title == publication.clean_title:
        return publication
    return replace(publication, clean_title=clean_title)


def _normalize_reference_style_title(clean_title: str) -> str:
    normalized = _TRAILING_REFERENCE_LINK_LABELS_RE.sub("", clean_title).strip()
    normalized = _strip_trailing_reference_venue_year(normalized)
    normalized = _strip_leading_reference_author_prefix(normalized)
    normalized = _strip_trailing_reference_venue_year(normalized)
    return _normalize_guard_title(normalized)


def _strip_trailing_reference_venue_year(clean_title: str) -> str:
    match = _TRAILING_REFERENCE_VENUE_MARKER_RE.search(clean_title)
    if match is None or match.start() < 24:
        return clean_title
    stripped = clean_title[: match.start()].strip(" ,，.;；")
    return stripped or clean_title


def _strip_leading_reference_author_prefix(clean_title: str) -> str:
    for match in re.finditer(r"\s+(?=[A-Z][A-Za-z0-9])", clean_title):
        split_at = match.end()
        prefix = clean_title[:split_at].strip(" ,，;；")
        suffix = clean_title[split_at:].strip()
        if len(suffix) < 24 or not _looks_like_titleish_context_text(suffix):
            continue
        if _looks_like_short_author_sequence(prefix) or _looks_like_pubmed_author_list(
            prefix
        ):
            return suffix
    return clean_title


def _profile_raw_text_fallback_content(prof: dict[str, Any]) -> str | None:
    raw_text = prof.get("profile_raw_text")
    if not isinstance(raw_text, str):
        return None
    normalized = raw_text.strip()
    return normalized or None


def _fragment_profile_name(homepage_url: object) -> str | None:
    if not isinstance(homepage_url, str):
        return None
    fragment = unquote(urlparse(homepage_url).fragment or "")
    if not fragment.casefold().startswith("prof-"):
        return None
    name = fragment[5:].strip()
    return name or None


def _is_fragment_official_profile(prof: dict[str, Any]) -> bool:
    if prof.get("homepage_page_role") not in _FRAGMENT_PROFILE_PAGE_ROLES:
        return False
    return _fragment_profile_name(prof.get("homepage_url")) is not None


def _stored_fragment_raw_text_matches_profile(
    prof: dict[str, Any],
    raw_text: str,
) -> bool:
    expected_name = str(
        prof.get("canonical_name") or _fragment_profile_name(prof.get("homepage_url")) or ""
    ).strip()
    if not expected_name:
        return False
    normalized = re.sub(r"\s+", " ", raw_text).strip()
    if not normalized:
        return False
    first_window = normalized[:240]
    first_name_at = first_window.find(expected_name)
    if first_name_at < 0 or first_name_at > 80:
        return False
    return True


def _scoped_fragment_profile_raw_text(prof: dict[str, Any]) -> str | None:
    if not _is_fragment_official_profile(prof):
        return None
    raw_text = _profile_raw_text_fallback_content(prof)
    if not raw_text:
        return None
    if not _stored_fragment_raw_text_matches_profile(prof, raw_text):
        return None
    return raw_text


def _scope_fragment_profile_html(prof: dict[str, Any], html: str) -> str:
    homepage_url = prof.get("homepage_url")
    if not isinstance(homepage_url, str) or not _is_fragment_official_profile(prof):
        return html
    scoped = _scope_html_to_fragment_profile(html, homepage_url)
    if scoped:
        return scoped
    try:
        node = _find_sztu_fragment_profile_node(
            BeautifulSoup(html, "html.parser"),
            homepage_url,
        )
    except Exception:
        return html
    return str(node) if node is not None else html


def _profile_raw_text_fetch_fallback_content(prof: dict[str, Any]) -> str | None:
    if _is_fragment_official_profile(prof):
        return _scoped_fragment_profile_raw_text(prof)
    return _profile_raw_text_fallback_content(prof)


def _known_unproductive_profile_raw_text(prof: dict[str, Any]) -> str | None:
    if prof.get("homepage_page_role") != "official_profile":
        return None
    if not _is_known_unproductive_official_profile_url(prof.get("homepage_url")):
        return None
    return _profile_raw_text_fallback_content(prof)


def _is_direct_professor_page_pdf(resolved: ResolvedPaper) -> bool:
    if not resolved.pdf_url:
        return False
    parsed = urlparse(resolved.pdf_url)
    hostname = (parsed.hostname or "").lower()
    return not (hostname.endswith("arxiv.org") and parsed.path.startswith("/pdf/"))


def _is_pdf_fetch_cap_error(fetch_error: str | None) -> bool:
    return fetch_error in _PDF_FETCH_CAP_ERRORS


def _is_malformed_publication_title(publication) -> bool:
    clean_title = str(getattr(publication, "clean_title", "") or "").strip()
    if not clean_title:
        return True
    if _publication_context_has_non_paper_record_marker(publication):
        return True
    if re.match(r"^in\s+(?:arxiv\s+)?preprint\b", clean_title, re.IGNORECASE):
        return True
    if re.search(r"\b(?:etc|et\s+al\.?)\b", clean_title, re.IGNORECASE):
        return True
    if _looks_like_common_romanized_person_name(clean_title):
        return True
    if _is_suspicious_rule_publication(
        publication
    ) and not _is_context_supported_short_publication_title(publication):
        return True
    if _is_non_publication_label_title(clean_title):
        return True
    if _is_short_venue_only_title(publication):
        return True
    if _is_context_supported_short_publication_title(publication):
        return False
    if not is_plausible_paper_title(clean_title):
        return True
    if _is_author_fragment_title(clean_title):
        return True
    if _is_author_list_title_with_titleish_context(publication):
        return True
    has_explicit_author_syntax = (
        any(mark in clean_title for mark in (",", "，", ";", "；", "*", "#", "†", "‡"))
        or _AUTHOR_INITIAL_HINT_RE.search(clean_title) is not None
    )
    if not has_explicit_author_syntax or not _looks_like_author_list(clean_title):
        return False
    return not bool(str(getattr(publication, "authors_text", "") or "").strip())


def _is_context_supported_short_publication_title(publication) -> bool:
    clean_title = str(getattr(publication, "clean_title", "") or "").strip()
    if not clean_title:
        return False
    if _publication_context_has_non_paper_record_marker(publication):
        return False
    if _is_non_publication_label_title(clean_title):
        return False
    if _is_author_fragment_title(clean_title):
        return False
    if _title_has_explicit_author_syntax(clean_title):
        return False
    if re.search(r"\b(?:etc|et\s+al\.?)\b", clean_title, re.IGNORECASE):
        return False
    if re.match(r"^in\s+(?:arxiv\s+)?preprint\b", clean_title, re.IGNORECASE):
        return False
    authors_text = str(getattr(publication, "authors_text", "") or "").strip()
    if authors_text and _normalize_guard_title(authors_text) == _normalize_guard_title(
        clean_title
    ):
        return False
    if _looks_like_common_romanized_person_name(clean_title):
        return False
    has_context = bool(
        str(getattr(publication, "venue_text", "") or "").strip()
        or getattr(publication, "year", None)
        or str(getattr(publication, "source_anchor", "") or "").strip()
        or str(getattr(publication, "pdf_url", "") or "").strip()
    )
    if is_plausible_paper_title(clean_title):
        return has_context
    has_strong_bibliographic_context = bool(
        str(getattr(publication, "authors_text", "") or "").strip()
        and (
            str(getattr(publication, "venue_text", "") or "").strip()
            or getattr(publication, "year", None)
        )
    )
    return has_strong_bibliographic_context


def _publication_context_has_non_paper_record_marker(publication) -> bool:
    fields = (
        getattr(publication, "raw_title", None),
        getattr(publication, "venue_text", None),
        getattr(publication, "source_anchor", None),
    )
    return any(
        isinstance(value, str)
        and re.search(r"\bpatents?\b|专利|CN\d{6,}|US\d{6,}", value, re.IGNORECASE)
        for value in fields
    )


def _title_has_explicit_author_syntax(title: str) -> bool:
    return (
        any(mark in title for mark in (",", "，", ";", "；", "*", "#", "†", "‡"))
        or _AUTHOR_INITIAL_HINT_RE.search(title) is not None
    )


def _looks_like_common_romanized_person_name(title: str) -> bool:
    tokens = [
        re.sub(r"[^A-Za-z'’-]", "", token).casefold()
        for token in _normalize_guard_title(title).split()
    ]
    tokens = [token for token in tokens if token]
    tokens = [
        token for token in tokens if token not in {"and", "etc", "et", "al"}
    ]
    if not 2 <= len(tokens) <= 4:
        return False
    return (
        tokens[0] in _COMMON_ROMANIZED_CHINESE_SURNAMES
        or tokens[-1] in _COMMON_ROMANIZED_CHINESE_SURNAMES
    )


def _is_non_publication_label_title(title: str) -> bool:
    normalized = _normalize_guard_title(title)
    return (
        normalized.casefold() in _NON_PUBLICATION_LABEL_TITLES
        or _JCR_QUARTILE_LABEL_RE.fullmatch(normalized) is not None
        or _JCR_METRIC_LABEL_RE.fullmatch(normalized) is not None
        or _MONTH_DAY_LABEL_RE.fullmatch(normalized) is not None
        or _QUALIFICATION_TITLE_RE.search(normalized) is not None
        or _PROCEEDINGS_LABEL_ONLY_TITLE_RE.fullmatch(normalized) is not None
        or _IN_VENUE_LABEL_ONLY_TITLE_RE.fullmatch(normalized) is not None
        or _VENUE_COUNT_METRIC_ONLY_TITLE_RE.fullmatch(normalized) is not None
        or _CONNECTIVE_AUTHOR_FRAGMENT_TITLE_RE.fullmatch(normalized) is not None
        or _SPACED_NAME_FRAGMENT_TITLE_RE.fullmatch(normalized) is not None
        or _LEADING_ETC_TITLE_PREFIX_RE.fullmatch(normalized) is not None
        or _PUBLICATION_MARKER_LEGEND_TITLE_RE.search(normalized) is not None
        or _DATE_VOLUME_METADATA_TITLE_RE.fullmatch(normalized) is not None
        or _BIBLIOGRAPHIC_FRAGMENT_TITLE_RE.fullmatch(normalized) is not None
        or _ELLIPSIS_AUTHOR_FRAGMENT_RE.fullmatch(normalized) is not None
        or _LOWERCASE_CONTINUATION_FRAGMENT_RE.fullmatch(normalized) is not None
        or _looks_like_cjk_author_list_title(normalized)
    )


def _is_author_fragment_title(title: str) -> bool:
    normalized = _normalize_guard_title(title)
    if _AUTHOR_YEAR_FRAGMENT_RE.fullmatch(normalized):
        return True
    if _CONCATENATED_AUTHOR_FRAGMENT_RE.fullmatch(normalized):
        return True
    if _SEMICOLON_SURNAME_INITIAL_AUTHOR_FRAGMENT_RE.fullmatch(normalized):
        return True
    if _NO_SPACE_AND_AUTHOR_FRAGMENT_RE.fullmatch(normalized):
        return True
    if _looks_like_pubmed_author_segment(normalized):
        return True
    return _looks_like_pubmed_author_list(normalized)


def _looks_like_pubmed_author_list(title: str) -> bool:
    normalized = re.sub(r"\s+", " ", title).strip()
    if "," not in normalized and "，" not in normalized:
        return False
    parts = [
        part.strip()
        for part in re.split(r"\s*[,，]\s*", normalized)
        if part.strip()
    ]
    if len(parts) < 2:
        return False
    return all(_looks_like_pubmed_author_segment(part) for part in parts)


def _looks_like_pubmed_author_segment(title: str) -> bool:
    normalized = re.sub(r"[*#†‡]", "", title)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    if not normalized or re.search(r"[,，;；:：()（）\d]", normalized):
        return False
    if len(normalized.split()) != 2:
        return False
    return (
        re.fullmatch(
            r"[A-Z][A-Za-z?'’-]{1,40}\s+(?:[A-Z]{1,5}|(?:[A-Z]\.?\s*){1,5})",
            normalized,
        )
        is not None
    )


def _is_short_venue_only_title(publication) -> bool:
    clean_title = str(getattr(publication, "clean_title", "") or "").strip()
    venue_text = str(getattr(publication, "venue_text", "") or "").strip()
    normalized_title = _normalize_guard_title(clean_title).casefold()
    if normalized_title in _KNOWN_VENUE_ONLY_TITLES:
        return True
    if not clean_title or not venue_text:
        return False
    normalized_venue = _normalize_guard_title(venue_text).casefold()
    if normalized_title != normalized_venue:
        return False
    if re.search(r"[:：?？]", clean_title):
        return False
    word_count = len(re.findall(r"[A-Za-z][A-Za-z'&-]*", clean_title))
    cjk_count = len(re.findall(r"[\u4e00-\u9fff]", clean_title))
    return 2 <= word_count <= 8 or 4 <= cjk_count <= 18


def _normalize_guard_title(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip(" .:：;；")


def _is_author_list_title_with_titleish_context(publication) -> bool:
    clean_title = str(getattr(publication, "clean_title", "") or "").strip()
    venue_text = str(getattr(publication, "venue_text", "") or "").strip()
    if not clean_title:
        return False
    if "学生" in clean_title:
        return bool(
            re.search(r"[,，;；]", clean_title)
            or _AUTHOR_INITIAL_HINT_RE.search(clean_title)
        )
    if not _looks_like_titleish_context_text(venue_text):
        return False
    return _looks_like_short_author_sequence(clean_title)


def _looks_like_titleish_context_text(text: str) -> bool:
    normalized = re.sub(r"\s+", " ", text).strip()
    if len(normalized) < 32:
        return False
    if re.search(r"[\u4e00-\u9fff]", normalized):
        return len(normalized) >= 12 and any(
            marker in normalized
            for marker in ("研究", "模型", "方法", "系统", "设计", "分析")
        )
    words = re.findall(r"[A-Za-z][A-Za-z'’-]*", normalized)
    return len(words) >= 5 or ":" in normalized


def _looks_like_short_author_sequence(title: str) -> bool:
    normalized = re.sub(r"[（）()【】\\[\\]{}]", " ", title)
    normalized = re.sub(r"[*#†‡]", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip(" ,，;；.")
    if len(normalized) > 100:
        return False
    if not _AUTHOR_LIST_DELIMITER_RE.search(normalized):
        return False
    parts = [
        part.strip()
        for part in _AUTHOR_LIST_DELIMITER_RE.split(normalized)
        if part.strip()
    ]
    if not 2 <= len(parts) <= 6:
        return False
    if not all(_looks_like_simple_person_name(part) for part in parts):
        return False
    if re.search(r"[,，;；]", normalized):
        return True
    return any(_has_common_romanized_chinese_surname(part) for part in parts)


def _looks_like_cjk_author_list_title(title: str) -> bool:
    if not re.search(r"[,，、;；&＆]", title):
        return False
    normalized = re.sub(r"[（）()【】\[\]{}]", " ", title)
    parts = [
        part.strip()
        for part in re.split(r"\s*(?:[,，、;；]|&|＆)\s*", normalized)
        if part.strip()
    ]
    if not 2 <= len(parts) <= 8:
        return False
    return all(re.fullmatch(r"[\u4e00-\u9fff]{2,4}", part) for part in parts)


def _looks_like_simple_person_name(text: str) -> bool:
    normalized = re.sub(r"\s+", " ", text).strip(" .")
    tokens = re.findall(r"[A-Za-z][A-Za-z'’-]*", normalized)
    if len(tokens) != 2:
        return False
    return all(1 < len(token) <= 30 for token in tokens)


def _has_common_romanized_chinese_surname(text: str) -> bool:
    tokens = [
        re.sub(r"[^A-Za-z'’-]", "", token).casefold()
        for token in re.findall(r"[A-Za-z][A-Za-z'’-]*", text)
    ]
    return any(token in _COMMON_ROMANIZED_CHINESE_SURNAMES for token in tokens)


@dataclass(frozen=True, slots=True)
class IngestReport:
    run_id: UUID
    profs_total: int
    profs_processed: int
    profs_skipped: int
    papers_linked_total: int
    full_text_fetched_total: int
    pipeline_issues_filed: int
    run_duration_seconds: float


@dataclass(frozen=True, slots=True)
class _ZeroExtractionPublicationPage:
    page_url: str
    sections_detected: int
    heading_texts: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class _SecondHopPublicationPageCandidate:
    page_url: str
    status: str = "candidate"
    skip_reason: str | None = None


@dataclass(frozen=True, slots=True)
class _HomepagePublicationPageOutcome:
    page_url: str
    status: str
    recursion_depth: int
    publications_extracted: int = 0
    sections_detected: int = 0
    heading_texts: tuple[str, ...] = ()
    skip_reason: str | None = None
    fetch_error_type: str | None = None
    fetch_error_message: str | None = None


@dataclass(frozen=True, slots=True)
class _HomepagePublicationExtractionReport:
    publications: list[HomepagePublication]
    zero_extraction_pages: tuple[_ZeroExtractionPublicationPage, ...]
    page_outcomes: tuple[_HomepagePublicationPageOutcome, ...] = ()


def _count_only_publication_claim(html: str) -> str | None:
    text = _scoped_html_to_issue_text(html) or _html_to_issue_text(html)
    if not text:
        return None
    match = _COUNT_ONLY_PUBLICATION_CLAIM_RE.search(text)
    if not match:
        return None
    return _truncate_issue_detail_text(match.group("claim"))


def _scoped_html_to_issue_text(html: str) -> str | None:
    try:
        soup = BeautifulSoup(html, "html.parser")
    except Exception:
        return None

    for selector in _COUNT_ONLY_PUBLICATION_TEXT_SELECTORS:
        texts = [
            node.get_text(" ", strip=True)
            for node in soup.select(selector)
            if node.get_text(" ", strip=True)
        ]
        if texts:
            return re.sub(r"\s+", " ", " ".join(texts)).strip()
    return None


def _html_to_issue_text(html: str) -> str:
    class _TextParser(HTMLParser):
        def __init__(self) -> None:
            super().__init__(convert_charrefs=True)
            self.parts: list[str] = []

        def handle_data(self, data: str) -> None:
            if data.strip():
                self.parts.append(data)

    parser = _TextParser()
    try:
        parser.feed(html)
    except Exception:
        text = html
    else:
        text = " ".join(parser.parts) or html
    return re.sub(r"\s+", " ", text).strip()


def _file_pipeline_issue(
    conn,
    *,
    run_id,
    issue_type,
    professor_id,
    message,
    details=None,
) -> None:
    description = f"[{issue_type}] {message}"
    evidence_snapshot = json.dumps(
        {
            "run_id": str(run_id),
            "issue_type": issue_type,
            "message": message,
            "details": details,
        },
        ensure_ascii=False,
    )
    existing = conn.execute(
        """
        SELECT issue_id
          FROM pipeline_issue
         WHERE professor_id = %s
           AND link_id IS NULL
           AND institution IS NULL
           AND stage = %s
           AND reported_by = %s
           AND description_hash = md5(%s)
           AND resolved = false
         LIMIT 1
        """,
        (
            professor_id,
            "paper_attribution",
            "homepage_paper_ingest",
            description,
        ),
    ).fetchone()
    if existing is not None:
        issue_id = existing["issue_id"] if isinstance(existing, dict) else existing[0]
        conn.execute(
            """
            UPDATE pipeline_issue
               SET evidence_snapshot = %s::jsonb,
                   severity = %s,
                   reported_at = now()
             WHERE issue_id = %s
            """,
            (evidence_snapshot, "medium", issue_id),
        )
        return

    conn.execute(
        """
        INSERT INTO pipeline_issue (
            professor_id,
            institution,
            stage,
            severity,
            description,
            evidence_snapshot,
            reported_by
        )
        VALUES (%s, %s, %s, %s, %s, %s::jsonb, %s)
        ON CONFLICT DO NOTHING
        """,
        (
            professor_id,
            None,
            "paper_attribution",
            "medium",
            description,
            evidence_snapshot,
            "homepage_paper_ingest",
        ),
    )


def run_homepage_paper_ingest(
    conn,
    *,
    institution=None,
    department: str | None = None,
    seed_id: str | int | None = None,
    limit=None,
    dry_run=False,
    resume_checkpoint_path: Path | None = None,
    prof_id: str | None = None,
    prof_page_pdf_fetch_cap: int | None = _DEFAULT_PROF_PAGE_PDF_FETCH_CAP,
    publication_extractor: Callable[..., list[HomepagePublication]] | None = None,
    include_owned_homepage_pages: bool = False,
    external_resolution_max_per_professor: int
    | None = None,
    enable_arxiv_title_search: bool = True,
    web_search=None,
) -> IngestReport:
    started_at = time.monotonic()
    run_id = _DRY_RUN_SENTINEL_RUN_ID
    profs_processed = 0
    profs_skipped = 0
    papers_linked_total = 0
    full_text_fetched_total = 0
    pipeline_issues_filed = 0
    profs_with_errors = 0
    prof_page_pdf_fetches_started = 0
    run_opened = False
    active_publication_extractor = (
        publication_extractor or extract_publications_from_html
    )
    use_rule_publication_diagnostics = (
        publication_extractor is None
        and active_publication_extractor is _DEFAULT_PUBLICATION_EXTRACTOR
    )
    publication_extraction_mode = (
        "custom" if publication_extractor is not None else "rule"
    )
    effective_external_resolution_max_per_professor = (
        max(0, external_resolution_max_per_professor)
        if external_resolution_max_per_professor is not None
        else None
    )

    try:
        if not dry_run:
            run_id = open_pipeline_run(
                conn,
                run_kind="backfill_real",
                run_scope={
                    "task": "homepage_paper_ingest",
                    "institution": institution,
                    "department": department,
                    "seed_id": str(seed_id) if seed_id is not None else None,
                    "limit": limit,
                    "prof_id": prof_id,
                    "resume_checkpoint_path": (
                        str(resume_checkpoint_path)
                        if resume_checkpoint_path is not None
                        else None
                    ),
                    "publication_extraction_mode": publication_extraction_mode,
                    "include_owned_homepage_pages": include_owned_homepage_pages,
                    "external_resolution_max_per_professor": (
                        effective_external_resolution_max_per_professor
                    ),
                    "enable_arxiv_title_search": enable_arxiv_title_search,
                    "web_search_enabled": web_search is not None,
                },
                triggered_by="homepage_paper_ingest",
            )
            run_opened = True
            _commit_if_available(conn)

        resume_set = _load_resume_set(resume_checkpoint_path)
        professors = _fetch_professors(
            conn,
            institution=institution,
            department=department,
            seed_id=seed_id,
            limit=limit,
            prof_id=prof_id,
            include_owned_homepage_pages=include_owned_homepage_pages,
        )
        external_resolution_attempts_by_professor: dict[str, int] = {}
        seen_publication_keys_by_professor: dict[str, set[tuple[str, int | None]]] = {}

        for prof in professors:
            professor_id = str(prof["professor_id"])
            resume_key = _resume_key_for_professor_row(prof)
            if resume_key in resume_set:
                profs_skipped += 1
                continue

            profs_processed += 1
            prof_pipeline_issues = 0
            prof_had_error = False
            prof_papers_linked = 0
            checkpoint_status = "succeeded"

            # psycopg3：嵌套 transaction() 自动用 SAVEPOINT；不需要 savepoint=True kwarg
            with conn.transaction():
                try:
                    scoped_raw_text = (
                        _scoped_fragment_profile_raw_text(prof)
                        or _known_unproductive_profile_raw_text(prof)
                    )
                    if scoped_raw_text:
                        html = scoped_raw_text
                    else:
                        try:
                            html = fetch_homepage_html(prof["homepage_url"])
                            html = _scope_fragment_profile_html(prof, html)
                        except (
                            httpx.HTTPStatusError,
                            httpx.TransportError,
                        ) as exc:
                            logger.warning(
                                "Homepage fetch failed for %s (%s): %s",
                                professor_id,
                                prof["homepage_url"],
                                exc,
                            )
                            raw_text_fallback = _profile_raw_text_fetch_fallback_content(
                                prof
                            )
                            if raw_text_fallback:
                                html = raw_text_fallback
                                prof_had_error = True
                                pipeline_issues_filed += 1
                                prof_pipeline_issues += 1
                                if not dry_run:
                                    _file_pipeline_issue(
                                        conn,
                                        run_id=run_id,
                                        issue_type="homepage_fetch_raw_text_fallback",
                                        professor_id=professor_id,
                                        message=(
                                            "Homepage fetch failed; using stored "
                                            "profile_raw_text for publication extraction"
                                        ),
                                        details={
                                            "homepage_url": prof["homepage_url"],
                                            "profile_raw_text_len": len(
                                                raw_text_fallback
                                            ),
                                            "fetch_error": str(exc),
                                        },
                                    )
                                logger.info(
                                    "Using profile_raw_text fallback for %s (%s)",
                                    professor_id,
                                    prof["homepage_url"],
                                )
                            else:
                                prof_had_error = True
                                checkpoint_status = "failed"
                                pipeline_issues_filed += 1
                                prof_pipeline_issues += 1
                                if not dry_run:
                                    _file_pipeline_issue(
                                        conn,
                                        run_id=run_id,
                                        issue_type="homepage_fetch_error",
                                        professor_id=professor_id,
                                        message=str(exc),
                                        details={"homepage_url": prof["homepage_url"]},
                                    )
                                raise _SkipCurrentProfessor()

                    extraction_report = _extract_publications_from_homepage_source_pages(
                        html,
                        page_url=prof["homepage_url"],
                        publication_extractor=active_publication_extractor,
                        use_rule_diagnostics=use_rule_publication_diagnostics,
                    )
                    if extraction_report.page_outcomes and not dry_run:
                        _record_homepage_recursion_page_outcomes(
                            conn,
                            run_id=run_id,
                            prof=prof,
                            outcomes=extraction_report.page_outcomes,
                        )
                    publications = extraction_report.publications
                    publications = _filter_publications_for_professor_evidence(
                        prof,
                        publications,
                    )
                    publications = _filter_new_publications_for_professor(
                        professor_id,
                        publications,
                        seen_publication_keys_by_professor,
                    )
                    if extraction_report.zero_extraction_pages:
                        pipeline_issues_filed += 1
                        prof_pipeline_issues += 1
                        prof_had_error = True
                        if not dry_run:
                            _file_pipeline_issue(
                                conn,
                                run_id=run_id,
                                issue_type="publication_section_zero_extraction",
                                professor_id=professor_id,
                                message=(
                                    "Detected publication section but extracted "
                                    "zero publication records"
                                ),
                                details={
                                    "homepage_url": prof["homepage_url"],
                                    "pages": [
                                        {
                                            "page_url": page.page_url,
                                            "sections_detected": (
                                                page.sections_detected
                                            ),
                                            "heading_texts": list(
                                                page.heading_texts
                                            ),
                                        }
                                        for page in (
                                            extraction_report.zero_extraction_pages
                                        )
                                    ],
                                },
                            )
                    elif not publications and (
                        count_only_claim := _count_only_publication_claim(html)
                    ):
                        pipeline_issues_filed += 1
                        prof_pipeline_issues += 1
                        prof_had_error = True
                        if not dry_run:
                            _file_pipeline_issue(
                                conn,
                                run_id=run_id,
                                issue_type="publication_source_sparse_count_only",
                                professor_id=professor_id,
                                message=(
                                    "Homepage mentions publication counts but "
                                    "contains no extractable paper titles"
                                ),
                                details={
                                    "homepage_url": prof["homepage_url"],
                                    "claim_snippet": count_only_claim,
                                },
                            )
                    evidence_source_type = _homepage_evidence_source_type(prof)
                    if publications and evidence_source_type is None:
                        pipeline_issues_filed += 1
                        prof_pipeline_issues += 1
                        prof_had_error = True
                        checkpoint_status = "failed"
                        if not dry_run:
                            _file_pipeline_issue(
                                conn,
                                run_id=run_id,
                                issue_type="missing_homepage_tier",
                                professor_id=professor_id,
                                message=(
                                    "Professor homepage page_role is missing or "
                                    "not mappable to paper evidence tier"
                                ),
                                details={
                                    "homepage_url": prof["homepage_url"],
                                    "homepage_page_role": prof.get(
                                        "homepage_page_role"
                                    ),
                                    "publications_count": len(publications),
                                },
                            )
                        continue

                    if 0 < len(publications) < 3:
                        pipeline_issues_filed += 1
                        prof_pipeline_issues += 1
                        prof_had_error = True
                        if not dry_run:
                            _file_pipeline_issue(
                                conn,
                                run_id=run_id,
                                issue_type="publications_under_threshold",
                                professor_id=professor_id,
                                message=(
                                    "Extracted fewer than 3 publications from homepage"
                                ),
                                details={
                                    "homepage_url": prof["homepage_url"],
                                    "publications_count": len(publications),
                                },
                            )

                    cache = None if dry_run else PostgresTitleResolutionCache(conn)
                    unresolved_count = 0
                    page_only_count = 0
                    with httpx.Client(
                        timeout=_BULK_TITLE_RESOLUTION_TIMEOUT,
                        trust_env=False,
                    ) as title_http_client:
                        for publication in publications:
                            if _is_malformed_publication_title(publication):
                                pipeline_issues_filed += 1
                                prof_pipeline_issues += 1
                                prof_had_error = True
                                if not dry_run:
                                    _file_pipeline_issue(
                                        conn,
                                        run_id=run_id,
                                        issue_type="malformed_publication_title",
                                        professor_id=professor_id,
                                        message=(
                                            "Homepage publication clean_title looks like "
                                            "an author list, so title resolution was "
                                            "skipped"
                                        ),
                                        details={
                                            "homepage_url": prof["homepage_url"],
                                            "raw_title": getattr(
                                                publication,
                                                "raw_title",
                                                None,
                                            ),
                                            "clean_title": publication.clean_title,
                                            "authors_text": publication.authors_text,
                                            "venue_text": publication.venue_text,
                                        },
                                    )
                                continue

                            skip_external_resolution = (
                                _should_skip_external_title_resolution(
                                    publication,
                                    publication_count=len(publications),
                                )
                            )
                            resolution_budget_exhausted = (
                                effective_external_resolution_max_per_professor
                                is not None
                                and external_resolution_attempts_by_professor.get(
                                    professor_id,
                                    0,
                                )
                                >= effective_external_resolution_max_per_professor
                            )
                            allow_realtime_resolution = (
                                not skip_external_resolution
                                and not resolution_budget_exhausted
                            )
                            if allow_realtime_resolution:
                                external_resolution_attempts_by_professor[
                                    professor_id
                                ] = (
                                    external_resolution_attempts_by_professor.get(
                                        professor_id,
                                        0,
                                    )
                                    + 1
                                )
                            resolved = resolve_paper_by_title(
                                publication.clean_title,
                                author_hint=prof["canonical_name"],
                                year_hint=publication.year,
                                enable_arxiv_title_search=enable_arxiv_title_search,
                                web_search=web_search,
                                http_client=title_http_client,
                                cache=cache,
                                cache_only=not allow_realtime_resolution,
                            )
                            is_page_only = resolved is None
                            if is_page_only:
                                # Preprint case (Paper Review §3.1 P4): no
                                # external DB hit — create record from
                                # page-only data; enrichment fills later.
                                if allow_realtime_resolution:
                                    unresolved_count += 1
                                page_only_count += 1
                                resolved = _synthesize_page_only_resolution(
                                    publication,
                                    canonical_name=prof["canonical_name"],
                                )
                            resolved = _attach_professor_page_pdf_url(
                                resolved,
                                publication,
                            )

                            derived_paper_id = _derive_paper_id(
                                publication.clean_title,
                                resolved_doi=resolved.doi,
                                resolved_arxiv_id=resolved.arxiv_id,
                            )
                            papers_linked_total += 1
                            prof_papers_linked += 1

                            actual_paper_id = derived_paper_id
                            existing_page_only_paper_id = None
                            existing_canonical_paper_id = None
                            if is_page_only and not dry_run:
                                existing_page_only_paper_id = (
                                    _find_existing_linked_paper_for_page_only(
                                        conn,
                                        professor_id=professor_id,
                                        clean_title=publication.clean_title,
                                        year=publication.year,
                                    )
                                )
                            if existing_page_only_paper_id is None and not dry_run:
                                existing_canonical_paper_id = (
                                    _find_existing_canonical_homepage_paper(
                                        conn,
                                        clean_title=publication.clean_title,
                                        year=resolved.year or publication.year,
                                        doi=resolved.doi,
                                        arxiv_id=resolved.arxiv_id,
                                        authors=resolved.authors,
                                    )
                                )
                            if not dry_run:
                                if existing_page_only_paper_id is not None:
                                    actual_paper_id = existing_page_only_paper_id
                                elif existing_canonical_paper_id is not None:
                                    actual_paper_id = existing_canonical_paper_id
                                else:
                                    paper_report = upsert_paper(
                                        conn,
                                        title_clean=publication.clean_title,
                                        title_raw=resolved.title,
                                        doi=resolved.doi,
                                        arxiv_id=resolved.arxiv_id,
                                        openalex_id=resolved.openalex_id,
                                        semantic_scholar_id=None,
                                        year=resolved.year,
                                        venue=resolved.venue,
                                        abstract_clean=resolved.abstract,
                                        authors_display=_authors_display(
                                            resolved.authors
                                        ),
                                        citation_count=None,
                                        canonical_source=resolved.match_source,
                                        run_id=run_id,
                                        title_resolution_source=resolved.match_source,
                                        quality_status=NEEDS_ENRICHMENT,
                                    )
                                    actual_paper_id = getattr(
                                        paper_report,
                                        "paper_id",
                                        derived_paper_id,
                                    )
                                _upsert_professor_paper_link(
                                    conn,
                                    professor_id=professor_id,
                                    paper_id=actual_paper_id,
                                    link_status="verified",
                                    evidence_source_type=evidence_source_type,
                                    evidence_page_id=_ensure_publication_evidence_page_id(
                                        conn,
                                        prof=prof,
                                        publication=publication,
                                        run_id=run_id,
                                    ),
                                    evidence_api_source=None,
                                    match_reason=(
                                        _LINK_MATCH_REASON_PAGE_ONLY
                                        if is_page_only
                                        else _LINK_MATCH_REASON
                                    ),
                                    author_name_match_score=_AUTHOR_NAME_MATCH_SCORE,
                                    topic_consistency_score=None,
                                    institution_consistency_score=None,
                                    is_officially_listed=True,
                                    run_id=run_id,
                                )

                            if paper_full_text_exists(conn, actual_paper_id):
                                continue

                            if is_page_only and skip_external_resolution:
                                continue

                            is_prof_page_pdf = _is_direct_professor_page_pdf(resolved)
                            if (
                                is_prof_page_pdf
                                and prof_page_pdf_fetch_cap is not None
                                and prof_page_pdf_fetches_started
                                >= prof_page_pdf_fetch_cap
                            ):
                                pipeline_issues_filed += 1
                                prof_pipeline_issues += 1
                                prof_had_error = True
                                if not dry_run:
                                    _file_pipeline_issue(
                                        conn,
                                        run_id=run_id,
                                        issue_type="pdf_fetch_cap_exceeded",
                                        professor_id=professor_id,
                                        message=(
                                            "Professor-page PDF fetch cap exceeded for "
                                            f"{resolved.pdf_url}"
                                        ),
                                        details={
                                            "paper_id": actual_paper_id,
                                            "paper_title": resolved.title,
                                            "pdf_url": resolved.pdf_url,
                                            "prof_page_pdf_fetch_cap": (
                                                prof_page_pdf_fetch_cap
                                            ),
                                        },
                                    )
                                continue
                            if is_prof_page_pdf:
                                prof_page_pdf_fetches_started += 1

                            extract = fetch_and_extract_full_text(
                                resolved,
                                paper_id=actual_paper_id,
                            )
                            if extract.fetch_error is None:
                                full_text_fetched_total += 1
                            elif (
                                is_prof_page_pdf
                                and _is_pdf_fetch_cap_error(extract.fetch_error)
                            ):
                                pipeline_issues_filed += 1
                                prof_pipeline_issues += 1
                                prof_had_error = True
                                if not dry_run:
                                    _file_pipeline_issue(
                                        conn,
                                        run_id=run_id,
                                        issue_type="pdf_fetch_cap_violation",
                                        professor_id=professor_id,
                                        message=(
                                            "Professor-page PDF fetch violated configured "
                                            f"fetch policy for {resolved.pdf_url}"
                                        ),
                                        details={
                                            "paper_id": actual_paper_id,
                                            "paper_title": resolved.title,
                                            "pdf_url": resolved.pdf_url,
                                            "fetch_error": extract.fetch_error,
                                        },
                                    )
                            if not dry_run:
                                upsert_paper_full_text(
                                    conn,
                                    paper_id=actual_paper_id,
                                    extract=extract,
                                    run_id=run_id,
                                )

                    if publications and unresolved_count == len(publications):
                        pipeline_issues_filed += 1
                        prof_pipeline_issues += 1
                        prof_had_error = True
                        if not dry_run:
                            _file_pipeline_issue(
                                conn,
                                run_id=run_id,
                                issue_type="all_titles_unresolvable",
                                professor_id=professor_id,
                                message="All homepage publication titles were unresolvable",
                                details={"publications_count": len(publications)},
                            )
                except _SkipCurrentProfessor:
                    pass
                except Exception as exc:  # noqa: BLE001
                    prof_had_error = True
                    checkpoint_status = "failed"
                    pipeline_issues_filed += 1
                    prof_pipeline_issues += 1
                    logger.exception(
                        "Professor processing crashed for %s: %s",
                        professor_id,
                        exc,
                    )
                    if not dry_run:
                        _file_pipeline_issue(
                            conn,
                            run_id=run_id,
                            issue_type="prof_processing_crashed",
                            professor_id=professor_id,
                            message=str(exc),
                            details={"homepage_url": prof["homepage_url"]},
                        )

            if not dry_run:
                _commit_if_available(conn)
            if prof_had_error:
                profs_with_errors += 1
            _append_checkpoint_line(
                resume_checkpoint_path,
                prof_id=professor_id,
                homepage_url=prof.get("homepage_url"),
                status=checkpoint_status,
                papers_linked=prof_papers_linked,
                pipeline_issues=prof_pipeline_issues,
                dry_run=dry_run,
            )
            if not dry_run and resume_checkpoint_path is not None:
                resume_set.add(resume_key)

    except KeyboardInterrupt:
        if run_opened:
            close_pipeline_run(
                conn,
                run_id,
                status="failed",
            )
            _commit_if_available(conn)
        raise
    except Exception as exc:
        if run_opened:
            close_pipeline_run(
                conn,
                run_id,
                status="failed",
                error_summary={"msg": str(exc)},
            )
            _commit_if_available(conn)
        raise
    else:
        if run_opened:
            close_pipeline_run(
                conn,
                run_id,
                status="succeeded",
                items_processed=profs_processed,
                items_failed=profs_with_errors,
            )
            _commit_if_available(conn)

    return IngestReport(
        run_id=run_id,
        profs_total=len(professors),
        profs_processed=profs_processed,
        profs_skipped=profs_skipped,
        papers_linked_total=papers_linked_total,
        full_text_fetched_total=full_text_fetched_total,
        pipeline_issues_filed=pipeline_issues_filed,
        run_duration_seconds=time.monotonic() - started_at,
    )


def _authors_display(authors: tuple[str, ...]) -> str | None:
    if not authors:
        return None
    return ", ".join(author for author in authors if author)


def _commit_if_available(conn) -> None:
    commit = getattr(conn, "commit", None)
    if callable(commit):
        commit()


def _append_checkpoint_line(
    checkpoint_path: Path | None,
    *,
    prof_id: str,
    homepage_url: object,
    status: str,
    papers_linked: int,
    pipeline_issues: int,
    dry_run: bool,
) -> None:
    if dry_run or checkpoint_path is None:
        return

    checkpoint_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "prof_id": prof_id,
        "homepage_url": homepage_url if isinstance(homepage_url, str) else None,
        "resume_key": _resume_key(prof_id, homepage_url),
        "status": status,
        "papers_linked": papers_linked,
        "pipeline_issues": pipeline_issues,
    }
    with checkpoint_path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(payload, ensure_ascii=False))
        handle.write("\n")


def _derive_paper_id(
    clean_title: str,
    *,
    resolved_doi: str | None,
    resolved_arxiv_id: str | None,
) -> str:
    if resolved_doi:
        return build_stable_id("paper", f"doi:{resolved_doi}")
    if resolved_arxiv_id:
        return build_stable_id("paper", f"arxiv:{resolved_arxiv_id}")
    title_sha1 = hashlib.sha1(clean_title.encode("utf-8")).hexdigest()
    return build_stable_id("paper", f"title:{title_sha1}")


def _find_existing_linked_paper_for_page_only(
    conn,
    *,
    professor_id: str,
    clean_title: str,
    year: int | None,
) -> str | None:
    title_key = _page_only_reuse_title_key(clean_title)
    if not title_key:
        return None
    row = conn.execute(
        """
        SELECT COALESCE(pma.canonical_paper_id, p.paper_id) AS paper_id
          FROM professor_paper_link ppl
          JOIN paper p ON p.paper_id = ppl.paper_id
          LEFT JOIN paper_merge_alias pma ON pma.old_paper_id = p.paper_id
          JOIN paper resolved_p
            ON resolved_p.paper_id = COALESCE(pma.canonical_paper_id, p.paper_id)
         WHERE ppl.professor_id = %s
           AND ppl.link_status = 'verified'
           AND COALESCE(ppl.is_officially_listed, false) = true
           AND regexp_replace(lower(p.title_clean), '\\s+', '', 'g') = %s
           AND p.year IS NOT DISTINCT FROM %s
           AND COALESCE(resolved_p.identity_status, 'unverified') NOT IN ('rejected', 'merged')
         ORDER BY CASE
                    WHEN resolved_p.doi IS NOT NULL
                      OR resolved_p.openalex_id IS NOT NULL
                      OR resolved_p.arxiv_id IS NOT NULL
                      OR resolved_p.semantic_scholar_id IS NOT NULL
                    THEN 0
                    WHEN resolved_p.canonical_source <> 'prof_page_only'
                    THEN 1
                    ELSE 2
                  END,
                  resolved_p.updated_at DESC NULLS LAST,
                  resolved_p.paper_id
         LIMIT 1
        """,
        (professor_id, title_key, year),
    ).fetchone()
    if row is None:
        return None
    if isinstance(row, Mapping):
        value = row.get("paper_id")
    else:
        value = row[0]
    return value if isinstance(value, str) and value else None


def _find_existing_canonical_homepage_paper(
    conn,
    *,
    clean_title: str,
    year: int | None,
    doi: str | None,
    arxiv_id: str | None,
    authors: tuple[str, ...],
) -> str | None:
    conditions: list[str] = []
    params: list[object] = []
    normalized_doi = str(doi).strip().lower() if doi else ""
    normalized_arxiv_id = str(arxiv_id).strip().lower() if arxiv_id else ""
    if normalized_doi:
        conditions.append("lower(p.doi) = %s")
        params.append(normalized_doi)
    if normalized_arxiv_id:
        conditions.append("lower(p.arxiv_id) = %s")
        params.append(normalized_arxiv_id)

    title_key = _page_only_reuse_title_key(clean_title)
    if title_key:
        # Content anchor: title+year only (no author-overlap gate). The author gate
        # caused co-authored papers on different professor pages to miss (each page
        # synthesizes a different "ProfName et al." author string). The exact
        # whitespace-stripped-lowercased title + year is a strong enough anchor.
        # See OpenSpec change ingest-dedup-anchor-before-insert.
        conditions.append(
            "(regexp_replace(lower(p.title_clean), '\\s+', '', 'g') = %s"
            " AND p.year IS NOT DISTINCT FROM %s)"
        )
        params.extend([title_key, year])

    if not conditions:
        return None

    row = conn.execute(
        f"""
        SELECT COALESCE(pma.canonical_paper_id, p.paper_id) AS paper_id
          FROM paper p
          LEFT JOIN paper_merge_alias pma ON pma.old_paper_id = p.paper_id
          JOIN paper resolved_p
            ON resolved_p.paper_id = COALESCE(pma.canonical_paper_id, p.paper_id)
         WHERE ({' OR '.join(conditions)})
           AND COALESCE(resolved_p.identity_status, 'unverified') NOT IN ('rejected', 'merged')
         ORDER BY CASE
                    WHEN resolved_p.doi IS NOT NULL
                      OR resolved_p.openalex_id IS NOT NULL
                      OR resolved_p.arxiv_id IS NOT NULL
                      OR resolved_p.semantic_scholar_id IS NOT NULL
                    THEN 0
                    WHEN resolved_p.canonical_source <> 'prof_page_only'
                    THEN 1
                    ELSE 2
                  END,
                  resolved_p.updated_at DESC NULLS LAST,
                  resolved_p.paper_id
         LIMIT 1
        """,
        tuple(params),
    ).fetchone()
    if row is None:
        return None
    if isinstance(row, Mapping):
        value = row.get("paper_id")
    else:
        value = row[0]
    return value if isinstance(value, str) and value else None


def _canonical_author_lookup_terms(authors: tuple[str, ...]) -> tuple[str, ...]:
    terms: list[str] = []
    for author in authors:
        value = str(author or "").strip().casefold()
        if not value:
            continue
        if _CJK_RE.search(value):
            compact = re.sub(r"\s+", "", value)
            if len(compact) >= 2:
                terms.append(compact)
            continue
        parts = [
            part
            for part in re.split(r"[^a-z0-9]+", value)
            if len(part) > 1 and not part.isdigit()
        ]
        if parts:
            terms.append(parts[-1])
    return tuple(dict.fromkeys(terms))


def _page_only_reuse_title_key(clean_title: str) -> str:
    return re.sub(r"\s+", "", _normalize_guard_title(clean_title)).casefold()


def _fetch_professors(
    conn,
    *,
    institution: str | None,
    department: str | None = None,
    seed_id: str | int | None = None,
    limit: int | None,
    prof_id: str | None,
    include_owned_homepage_pages: bool = False,
) -> list[dict[str, Any]]:
    # V003 schema: professor.institution / homepage_url 已迁出主表。
    # institution 走 professor_affiliation 多对多；homepage_url 走 source_page
    # via primary_official_profile_page_id FK。
    params: list[Any] = []
    ctes, seed_join = _seed_scope_ctes_and_join(seed_id, params)
    if seed_id is None:
        institution_expr = "primary_aff.institution"
        department_expr = "primary_aff.department"
    else:
        institution_expr = "COALESCE(seed_scope.institution, primary_aff.institution)"
        department_expr = "COALESCE(seed_scope.department, primary_aff.department)"

    if include_owned_homepage_pages:
        query = [
            "WITH selected_professors AS (",
            "SELECT p.professor_id,",
            "       p.canonical_name,",
            "       p.canonical_name_en,",
            "       p.aliases,",
            "       p.primary_official_profile_page_id,",
            "       p.profile_raw_text,",
            "       primary_aff.title AS affiliation_title,",
            f"       COALESCE({institution_expr}, '') AS institution",
            "  FROM professor p",
            seed_join,
            "  LEFT JOIN LATERAL (",
            "    SELECT pa.institution, pa.department, pa.title",
            "    FROM professor_affiliation pa",
            "    WHERE pa.professor_id = p.professor_id",
            "    ORDER BY pa.is_primary DESC,",
            "             pa.is_current DESC,",
            "             pa.start_year DESC NULLS LAST,",
            "             pa.created_at DESC NULLS LAST,",
            "             pa.affiliation_id DESC",
            "    LIMIT 1",
            "  ) primary_aff ON TRUE",
            " WHERE TRUE",
        ]
        if institution:
            query.append(f"AND {institution_expr} ILIKE %s")
            params.append(f"%{institution}%")
        if department:
            query.append(f"AND {department_expr} ILIKE %s")
            params.append(f"%{department}%")
        if prof_id:
            query.append("AND p.professor_id = %s")
            params.append(prof_id)
        if limit is not None:
            query.append("LIMIT %s")
            params.append(limit)
        query.extend(
            [
                ")",
                "SELECT p.professor_id::text AS professor_id,",
                "       p.canonical_name,",
                "       p.canonical_name_en,",
                "       p.aliases,",
                "       p.institution,",
                "       sp.page_id AS homepage_page_id,",
                "       sp.url AS homepage_url,",
                "       sp.page_role AS homepage_page_role,",
                "       p.profile_raw_text,",
                "       p.affiliation_title",
                "  FROM selected_professors p",
                "  LEFT JOIN source_page primary_sp",
                "    ON primary_sp.page_id = p.primary_official_profile_page_id",
                "  JOIN source_page sp ON (",
                "       sp.page_id = p.primary_official_profile_page_id",
                "       OR (",
                "            sp.owner_scope_kind = 'professor'",
                "            AND sp.owner_scope_ref = p.professor_id::text",
                "            AND sp.page_role IN (%s, %s, %s)",
                "            AND (",
                "                 sp.page_role <> 'personal_homepage'",
                "                 OR sp.url_host IS DISTINCT FROM primary_sp.url_host",
                "            )",
                "       )",
                "  )",
                " WHERE sp.url IS NOT NULL",
                " ORDER BY p.professor_id::text, sp.page_id",
            ]
        )
        params.extend(_OWNED_HOMEPAGE_PAGE_ROLES)
        if ctes:
            query[0] = "WITH " + ", ".join(
                ctes + [query[0].removeprefix("WITH ")]
            )
        rows = conn.execute(
            " ".join(part for part in query if part), tuple(params)
        ).fetchall()
        return _filter_homepage_ingest_professor_rows(_normalize_professor_rows(rows))

    query = [
        "SELECT p.professor_id::text AS professor_id,",
        "       p.canonical_name,",
        "       p.canonical_name_en,",
        "       p.aliases,",
        f"       COALESCE({institution_expr}, '') AS institution,",
        "       sp.page_id AS homepage_page_id,",
        "       sp.url AS homepage_url,",
        "       sp.page_role AS homepage_page_role,",
        "       p.profile_raw_text,",
        "       primary_aff.title AS affiliation_title",
        "  FROM professor p",
        seed_join,
        "  LEFT JOIN LATERAL (",
        "    SELECT pa.institution, pa.department, pa.title",
        "    FROM professor_affiliation pa",
        "    WHERE pa.professor_id = p.professor_id",
        "    ORDER BY pa.is_primary DESC,",
        "             pa.is_current DESC,",
        "             pa.start_year DESC NULLS LAST,",
        "             pa.created_at DESC NULLS LAST,",
        "             pa.affiliation_id DESC",
        "    LIMIT 1",
        "  ) primary_aff ON TRUE",
        "  LEFT JOIN source_page sp ON sp.page_id = p.primary_official_profile_page_id",
        " WHERE sp.url IS NOT NULL",
    ]
    if ctes:
        query.insert(0, "WITH " + ", ".join(ctes))
    if institution:
        query.append(f"AND {institution_expr} ILIKE %s")
        params.append(f"%{institution}%")
    if department:
        query.append(f"AND {department_expr} ILIKE %s")
        params.append(f"%{department}%")
    if prof_id:
        query.append("AND p.professor_id = %s")
        params.append(prof_id)
    if limit is not None:
        query.append("LIMIT %s")
        params.append(limit)

    rows = conn.execute(" ".join(query), tuple(params)).fetchall()
    return _filter_homepage_ingest_professor_rows(_normalize_professor_rows(rows))


def _seed_scope_ctes_and_join(
    seed_id: str | int | None,
    params: list[Any],
) -> tuple[list[str], str]:
    if seed_id is None:
        return [], ""

    params.append(str(seed_id))
    return [
        (
            "latest_seed_run AS ("
            " SELECT pr.run_id"
            " FROM pipeline_run pr"
            " WHERE pr.run_kind = 'roster_crawl'"
            "   AND pr.status = 'succeeded'"
            "   AND pr.run_scope->>'seed_id' = %s"
            " ORDER BY (COALESCE(pr.run_scope->>'trigger_mode', '') = 'full') DESC,"
            "          pr.started_at DESC NULLS LAST,"
            "          pr.created_at DESC NULLS LAST,"
            "          pr.run_id DESC"
            " LIMIT 1"
            ")"
        ),
        (
            "seed_professors AS ("
            " SELECT DISTINCT pa.professor_id, pa.institution, pa.department"
            " FROM professor_affiliation pa"
            " JOIN latest_seed_run lr ON lr.run_id = pa.run_id"
            ")"
        ),
    ], "  JOIN seed_professors seed_scope ON seed_scope.professor_id = p.professor_id"


def _filter_homepage_ingest_professor_rows(
    rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    filtered: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    professors_with_owned_pages = {
        str(row.get("professor_id") or "")
        for row in rows
        if row.get("homepage_page_role") in _OWNED_HOMEPAGE_PAGE_ROLES
        and is_homepage_publication_ingest_url(row.get("homepage_url"))
    }
    for row in rows:
        if is_unsafe_professor_paper_evidence_identity(
            row.get("canonical_name"),
            affiliation_title=row.get("affiliation_title"),
        ):
            continue
        homepage_url = row.get("homepage_url")
        if _should_skip_known_unproductive_profile_row(
            row,
            professors_with_owned_pages=professors_with_owned_pages,
        ):
            continue
        if not is_homepage_publication_ingest_url(
            homepage_url
        ) and not (
            _scoped_fragment_profile_raw_text(row)
            or _is_fragment_official_profile(row)
        ):
            continue
        normalized_url = _normalize_source_url(homepage_url)
        if not normalized_url:
            continue
        key = (str(row.get("professor_id") or ""), normalized_url.casefold())
        if key in seen:
            continue
        seen.add(key)
        filtered.append(row)
    return filtered


def _should_skip_known_unproductive_profile_row(
    row: dict[str, Any],
    *,
    professors_with_owned_pages: set[str],
) -> bool:
    professor_id = str(row.get("professor_id") or "")
    if professor_id not in professors_with_owned_pages:
        return False
    if row.get("homepage_page_role") != "official_profile":
        return False
    if _scoped_fragment_profile_raw_text(row):
        return False
    return _is_known_unproductive_official_profile_url(row.get("homepage_url"))


def _is_known_unproductive_official_profile_url(value: object) -> bool:
    if not isinstance(value, str):
        return False
    parsed = urlparse(value)
    hostname = (parsed.hostname or "").lower()
    path = parsed.path.rstrip("/").lower()
    return hostname == "csse.szu.edu.cn" and path == "/pages/user/index"


def _normalize_professor_rows(rows) -> list[dict[str, Any]]:
    normalized_rows: list[dict[str, Any]] = []
    for row in rows:
        if isinstance(row, dict):
            normalized_rows.append(row)
            continue
        normalized_rows.append(
            {
                "professor_id": row[0],
                "canonical_name": row[1],
                "canonical_name_en": row[2] if len(row) >= 10 else None,
                "aliases": row[3] if len(row) >= 10 else [],
                "institution": row[4] if len(row) >= 10 else row[2],
                "homepage_page_id": (
                    row[5] if len(row) >= 10 else (row[3] if len(row) > 5 else None)
                ),
                "homepage_url": (
                    row[6] if len(row) >= 10 else (row[4] if len(row) > 5 else row[3])
                ),
                "homepage_page_role": (
                    row[7] if len(row) >= 10 else (row[5] if len(row) > 5 else row[4])
                ),
                "profile_raw_text": (
                    row[8] if len(row) >= 10 else (row[6] if len(row) > 6 else None)
                ),
                "affiliation_title": (
                    row[9] if len(row) >= 10 else (row[7] if len(row) > 7 else None)
                ),
            }
        )
    return normalized_rows


def _load_resume_set(checkpoint_path: Path | None) -> set[str]:
    if checkpoint_path is None or not checkpoint_path.exists():
        return set()

    resume_keys: set[str] = set()
    with checkpoint_path.open(encoding="utf-8") as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line:
                continue
            try:
                payload = json.loads(line)
            except json.JSONDecodeError:
                logger.warning("Ignoring corrupted checkpoint line: %r", line)
                continue
            if not isinstance(payload, dict):
                continue
            if not _is_safe_resume_checkpoint(payload):
                continue
            value = payload.get("resume_key")
            if isinstance(value, str) and value:
                resume_keys.add(value)
                continue
            prof_id = payload.get("prof_id")
            homepage_url = payload.get("homepage_url")
            if isinstance(prof_id, str) and isinstance(homepage_url, str):
                resume_keys.add(_resume_key(prof_id, homepage_url))
    return resume_keys


def _is_safe_resume_checkpoint(payload: Mapping[str, Any]) -> bool:
    if payload.get("status") != "succeeded":
        return False
    papers_linked = _checkpoint_int(payload.get("papers_linked"))
    pipeline_issues = _checkpoint_int(payload.get("pipeline_issues"))
    return papers_linked > 0 or pipeline_issues > 0


def _checkpoint_int(value: object) -> int:
    if isinstance(value, bool):
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, str):
        try:
            return int(value)
        except ValueError:
            return 0
    return 0


def _resume_key_for_professor_row(prof: Mapping[str, Any]) -> str:
    return _resume_key(str(prof["professor_id"]), prof.get("homepage_url"))


def _resume_key(prof_id: str, homepage_url: object) -> str:
    normalized_url = _normalize_source_url(homepage_url)
    return f"{prof_id}|{normalized_url or ''}"
