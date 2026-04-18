from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from pathlib import Path
import re
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from psycopg import Connection
from psycopg.types.json import Jsonb

from ..normalization import normalize_company_name, normalize_person_name
from ..storage.postgres.connection import connect
from ._company_id import generate_company_id
from .team_parser import parse_team_raw


HEADER_ALIASES: dict[str, str] = {
    "序号": "sequence_no",
    "项目名称": "project_name",
    "行业领域": "industry",
    "子领域": "sub_industry",
    "业务": "business",
    "地区": "region",
    "投资轮次": "latest_funding_round",
    "投资时间": "latest_funding_time_raw",
    "投资金额": "latest_funding_amount_raw",
    "参考转化金额（万人民币）": "latest_funding_cny_wan",
    "比例": "latest_funding_ratio",
    "投资方": "latest_investors_raw",
    "FA信息": "latest_fa_info",
    "注册资金": "registered_capital",
    "高新企业": "is_high_tech",
    "简介": "description",
    "Logo链接": "logo_url",
    "星级": "star_rating",
    "状态": "status_raw",
    "备注": "remarks",
    "公司名称": "company_name_xlsx",
    "企业名称": "company_name_xlsx",
    "国别": "country_xlsx",
    "成立日期": "established_date",
    "网址": "website_xlsx",
    "法人代表": "legal_representative",
    "团队": "team_raw",
    "注册地址": "registered_address",
    "企业联系电话": "contact_phone",
    "联系邮箱": "contact_email",
    "成立年限": "years_established",
    "参保人数": "reported_insured_count",
    "股东数": "reported_shareholder_count",
    "投资数": "reported_investment_count",
    "专利数": "reported_patent_count",
    "商标数": "reported_trademark_count",
    "著作权": "reported_copyright_count",
    "招聘数": "reported_recruitment_count",
    "新闻数": "reported_news_count",
    "机构方数量": "reported_institution_count",
    "融资总次数": "reported_funding_round_count",
    "融资总额": "reported_total_funding_raw",
    "估值": "reported_valuation_raw",
}

_MISSING_MARKERS = {"-", "--", "—", "－", "N/A", "n/a", "NULL", "null"}
_TRUE_SET = {"是", "true", "1", "yes", "y"}
_FALSE_SET = {"否", "false", "0", "no", "n"}
_DATE_PATTERNS = ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d")
_NON_MERGED_CONTINUATION_KEYS = {
    "latest_funding_round",
    "latest_funding_time_raw",
    "latest_funding_amount_raw",
    "latest_funding_cny_wan",
    "latest_funding_ratio",
    "latest_investors_raw",
    "latest_fa_info",
}


@dataclass(frozen=True, slots=True)
class ImportReport:
    batch_id: UUID
    rows_read: int
    records_parsed: int
    records_new_company: int
    records_updated_company: int
    records_failed: int
    team_members_inserted: int
    funding_events_inserted: int
    lineage_rows: int


@dataclass(slots=True)
class _MergedCompanyRow:
    source_row_number: int
    source_row_numbers: list[int]
    values: dict[str, str | None]

    def raw_row_json(self) -> dict[str, Any]:
        payload = dict(self.values)
        payload["source_row_number"] = self.source_row_number
        payload["source_row_numbers"] = list(self.source_row_numbers)
        return payload


def import_company_xlsx_to_postgres(
    xlsx_path: Path,
    *,
    dsn: str,
    seed_id: str,
    triggered_by: str = "manual",
) -> ImportReport:
    """Import the canonical company xlsx into Postgres."""
    source = Path(xlsx_path)
    file_content_hash = sha256(source.read_bytes()).hexdigest()
    rows_read, merged_rows = _load_merged_rows(source)

    with connect(dsn) as conn:
        _ensure_seed_exists(conn, seed_id)
        if _batch_exists(conn, seed_id, file_content_hash):
            raise LookupError(
                f"import_batch already exists for seed_id={seed_id} "
                f"and file_content_hash={file_content_hash}"
            )

        pipeline_run_id = _insert_pipeline_run(
            conn,
            seed_id=seed_id,
            source=source,
            file_content_hash=file_content_hash,
            triggered_by=triggered_by,
        )
        batch_id = _insert_import_batch(
            conn,
            seed_id=seed_id,
            source=source,
            file_content_hash=file_content_hash,
            triggered_by=triggered_by,
        )

        records_new_company = 0
        records_updated_company = 0
        records_failed = 0
        team_members_inserted = 0
        funding_events_inserted = 0
        lineage_rows = 0
        records_parsed = len(merged_rows)

        for merged_row in merged_rows:
            raw_row_json = merged_row.raw_row_json()
            try:
                with conn.transaction():
                    company_id, is_new = _upsert_company(
                        conn,
                        batch_id=batch_id,
                        values=merged_row.values,
                    )
                    snapshot_id = _insert_company_snapshot(
                        conn,
                        company_id=company_id,
                        batch_id=batch_id,
                        source_row_number=merged_row.source_row_number,
                        values=merged_row.values,
                        raw_row_json=raw_row_json,
                    )
                    inserted_team_members = _insert_team_members(
                        conn,
                        company_id=company_id,
                        snapshot_id=snapshot_id,
                        team_raw=merged_row.values.get("team_raw"),
                    )
                    inserted_funding_events = _upsert_funding_event(
                        conn,
                        company_id=company_id,
                        company_name=merged_row.values.get("company_name_xlsx"),
                        values=merged_row.values,
                    )
                    _insert_source_row_lineage(
                        conn,
                        batch_id=batch_id,
                        source_row_number=merged_row.source_row_number,
                        target_entity_id=company_id,
                        resolution_status="created" if is_new else "matched",
                        resolution_reason=(
                            "created canonical company"
                            if is_new
                            else "matched existing canonical company"
                        ),
                        raw_row_json=raw_row_json,
                    )

                if is_new:
                    records_new_company += 1
                else:
                    records_updated_company += 1
                team_members_inserted += inserted_team_members
                funding_events_inserted += inserted_funding_events
                lineage_rows += 1
            except Exception as exc:
                records_failed += 1
                _insert_source_row_lineage(
                    conn,
                    batch_id=batch_id,
                    source_row_number=merged_row.source_row_number,
                    target_entity_id=None,
                    resolution_status="failed",
                    resolution_reason=str(exc),
                    raw_row_json=raw_row_json,
                )
                lineage_rows += 1

        run_status = "succeeded" if records_failed == 0 else "partial"
        _finalize_import_batch(
            conn,
            batch_id=batch_id,
            rows_read=rows_read,
            records_parsed=records_parsed,
            records_new=records_new_company,
            records_updated=records_updated_company,
            records_failed=records_failed,
            run_status=run_status,
        )
        _finalize_pipeline_run(
            conn,
            run_id=pipeline_run_id,
            items_processed=records_parsed,
            items_failed=records_failed,
            status=run_status,
        )

        return ImportReport(
            batch_id=batch_id,
            rows_read=rows_read,
            records_parsed=records_parsed,
            records_new_company=records_new_company,
            records_updated_company=records_updated_company,
            records_failed=records_failed,
            team_members_inserted=team_members_inserted,
            funding_events_inserted=funding_events_inserted,
            lineage_rows=lineage_rows,
        )


def _load_merged_rows(
    xlsx_path: Path,
    *,
    max_header_scan_rows: int = 32,
) -> tuple[int, list[_MergedCompanyRow]]:
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            raise ValueError(f"workbook has no active sheet: {xlsx_path}")

        header_row_index, column_mapping = _detect_header_row(
            sheet=sheet,
            max_header_scan_rows=max_header_scan_rows,
        )

        rows_read = 0
        merged_rows: list[_MergedCompanyRow] = []
        current_row: _MergedCompanyRow | None = None

        for row_index, row in enumerate(
            sheet.iter_rows(min_row=header_row_index + 1, values_only=True),
            start=header_row_index + 1,
        ):
            mapped_values = _extract_mapped_values(row, column_mapping)
            if not any(mapped_values.values()):
                continue

            rows_read += 1
            sequence_no = mapped_values.get("sequence_no")
            company_name = mapped_values.get("company_name_xlsx")

            if not sequence_no and not company_name:
                if current_row is None:
                    continue
                _merge_continuation_row(current_row, mapped_values, row_index)
                continue

            if not company_name and not _is_numeric_sequence(sequence_no):
                current_row = None
                continue

            current_row = _start_merged_row(mapped_values, row_index)
            merged_rows.append(current_row)

        return rows_read, merged_rows
    finally:
        workbook.close()


def _detect_header_row(sheet, max_header_scan_rows: int) -> tuple[int, dict[int, str]]:
    best_row_index = 0
    best_mapping: dict[int, str] = {}
    best_score = -1

    for row_index, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=max_header_scan_rows, values_only=True),
        start=1,
    ):
        mapping: dict[int, str] = {}
        for column_index, value in enumerate(row):
            header = _normalize_cell(value)
            if not header:
                continue
            canonical = HEADER_ALIASES.get(header)
            if canonical is None or canonical in mapping.values():
                continue
            mapping[column_index] = canonical

        score = len(mapping)
        if "company_name_xlsx" not in mapping.values():
            continue
        if score > best_score:
            best_row_index = row_index
            best_mapping = mapping
            best_score = score

    if not best_mapping:
        raise ValueError("unable to detect header row containing 公司名称/企业名称")
    return best_row_index, best_mapping


def _extract_mapped_values(
    row: tuple[object, ...],
    column_mapping: dict[int, str],
) -> dict[str, str | None]:
    mapped_values: dict[str, str | None] = {}
    for column_index, key in column_mapping.items():
        value = row[column_index] if column_index < len(row) else None
        mapped_values[key] = _normalize_cell(value)
    return mapped_values


def _start_merged_row(
    values: dict[str, str | None],
    row_index: int,
) -> _MergedCompanyRow:
    return _MergedCompanyRow(
        source_row_number=row_index,
        source_row_numbers=[row_index],
        values=dict(values),
    )


def _merge_continuation_row(
    merged_row: _MergedCompanyRow,
    values: dict[str, str | None],
    row_index: int,
) -> None:
    merged_row.source_row_numbers.append(row_index)
    for key, value in values.items():
        if key in _NON_MERGED_CONTINUATION_KEYS:
            continue
        if value and not merged_row.values.get(key):
            merged_row.values[key] = value


def _ensure_seed_exists(conn: Connection, seed_id: str) -> None:
    row = conn.execute(
        """
        SELECT seed_id
        FROM seed_registry
        WHERE seed_id = %s
        """,
        (seed_id,),
    ).fetchone()
    if row is None:
        raise LookupError(f"seed_id does not exist in seed_registry: {seed_id}")


def _batch_exists(conn: Connection, seed_id: str, file_content_hash: str) -> bool:
    row = conn.execute(
        """
        SELECT 1 AS exists_flag
        FROM import_batch
        WHERE seed_id = %s
          AND file_content_hash = %s
        LIMIT 1
        """,
        (seed_id, file_content_hash),
    ).fetchone()
    return row is not None


def _insert_pipeline_run(
    conn: Connection,
    *,
    seed_id: str,
    source: Path,
    file_content_hash: str,
    triggered_by: str,
) -> UUID:
    row = conn.execute(
        """
        INSERT INTO pipeline_run (
            run_kind,
            run_scope,
            seed_id,
            started_at,
            status,
            triggered_by
        )
        VALUES (%s, %s, %s, now(), %s, %s)
        RETURNING run_id
        """,
        (
            "import_xlsx",
            Jsonb(
                {
                    "xlsx_path": str(source.resolve()),
                    "file_content_hash": file_content_hash,
                }
            ),
            seed_id,
            "running",
            triggered_by,
        ),
    ).fetchone()
    assert row is not None
    return row["run_id"]


def _insert_import_batch(
    conn: Connection,
    *,
    seed_id: str,
    source: Path,
    file_content_hash: str,
    triggered_by: str,
) -> UUID:
    row = conn.execute(
        """
        INSERT INTO import_batch (
            seed_id,
            source_file,
            file_content_hash,
            started_at,
            run_status,
            triggered_by,
            error_summary
        )
        VALUES (%s, %s, %s, now(), %s, %s, %s)
        RETURNING batch_id
        """,
        (
            seed_id,
            str(source.resolve()),
            file_content_hash,
            "running",
            triggered_by,
            Jsonb({}),
        ),
    ).fetchone()
    assert row is not None
    return row["batch_id"]


def _upsert_company(
    conn: Connection,
    *,
    batch_id: UUID,
    values: dict[str, str | None],
) -> tuple[str, bool]:
    registered_name = values.get("company_name_xlsx")
    if not registered_name:
        raise ValueError("company_name_xlsx is required")

    canonical_name = normalize_company_name(registered_name)
    if not canonical_name:
        raise ValueError(f"unable to normalize company name: {registered_name}")

    company_id = generate_company_id(
        unified_credit_code=None,
        website=values.get("website_xlsx"),
        registered_name=registered_name,
    )
    is_new = _company_exists(conn, company_id) is False

    province, city, district = _derive_region_parts(
        values.get("region"),
        values.get("registered_address"),
    )
    is_shenzhen = any(
        "深圳" in part
        for part in (
            values.get("region") or "",
            values.get("registered_address") or "",
            city or "",
        )
    )

    conn.execute(
        """
        INSERT INTO company (
            company_id,
            canonical_name,
            registered_name,
            website,
            hq_province,
            hq_city,
            hq_district,
            is_shenzhen,
            country,
            identity_status,
            first_seen_batch_id,
            first_seen_at,
            last_refreshed_at
        )
        VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, now(), now()
        )
        ON CONFLICT (company_id) DO UPDATE
        SET last_refreshed_at = EXCLUDED.last_refreshed_at,
            website = COALESCE(company.website, EXCLUDED.website),
            registered_name = COALESCE(company.registered_name, EXCLUDED.registered_name),
            updated_at = now()
        """,
        (
            company_id,
            canonical_name,
            registered_name,
            values.get("website_xlsx"),
            province,
            city,
            district,
            is_shenzhen,
            values.get("country_xlsx") or "国内",
            "resolved",
            batch_id,
        ),
    )
    return company_id, is_new


def _company_exists(conn: Connection, company_id: str) -> bool:
    row = conn.execute(
        """
        SELECT 1 AS exists_flag
        FROM company
        WHERE company_id = %s
        """,
        (company_id,),
    ).fetchone()
    return row is not None


def _insert_company_snapshot(
    conn: Connection,
    *,
    company_id: str,
    batch_id: UUID,
    source_row_number: int,
    values: dict[str, str | None],
    raw_row_json: dict[str, Any],
) -> UUID:
    row = conn.execute(
        """
        INSERT INTO company_snapshot (
            company_id,
            import_batch_id,
            snapshot_kind,
            source_row_number,
            project_name,
            industry,
            sub_industry,
            business,
            region,
            description,
            logo_url,
            star_rating,
            status_raw,
            remarks,
            is_high_tech,
            company_name_xlsx,
            country_xlsx,
            established_date,
            years_established,
            website_xlsx,
            legal_representative,
            registered_address,
            registered_capital,
            contact_phone,
            contact_email,
            reported_insured_count,
            reported_shareholder_count,
            reported_investment_count,
            reported_patent_count,
            reported_trademark_count,
            reported_copyright_count,
            reported_recruitment_count,
            reported_news_count,
            reported_institution_count,
            reported_funding_round_count,
            reported_total_funding_raw,
            reported_valuation_raw,
            latest_funding_round,
            latest_funding_time_raw,
            latest_funding_time,
            latest_funding_amount_raw,
            latest_funding_cny_wan,
            latest_funding_ratio,
            latest_investors_raw,
            latest_fa_info,
            team_raw,
            raw_row_jsonb
        )
        VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
        )
        RETURNING snapshot_id
        """,
        (
            company_id,
            batch_id,
            "xlsx_import",
            source_row_number,
            values.get("project_name"),
            values.get("industry"),
            values.get("sub_industry"),
            values.get("business"),
            values.get("region"),
            values.get("description"),
            values.get("logo_url"),
            _parse_int(values.get("star_rating")),
            values.get("status_raw"),
            values.get("remarks"),
            _parse_bool(values.get("is_high_tech")),
            values.get("company_name_xlsx"),
            values.get("country_xlsx"),
            _parse_date(values.get("established_date")),
            _parse_int(values.get("years_established")),
            values.get("website_xlsx"),
            values.get("legal_representative"),
            values.get("registered_address"),
            values.get("registered_capital"),
            values.get("contact_phone"),
            values.get("contact_email"),
            _parse_int(values.get("reported_insured_count")),
            _parse_int(values.get("reported_shareholder_count")),
            _parse_int(values.get("reported_investment_count")),
            _parse_int(values.get("reported_patent_count")),
            _parse_int(values.get("reported_trademark_count")),
            _parse_int(values.get("reported_copyright_count")),
            _parse_int(values.get("reported_recruitment_count")),
            _parse_int(values.get("reported_news_count")),
            _parse_int(values.get("reported_institution_count")),
            _parse_int(values.get("reported_funding_round_count")),
            values.get("reported_total_funding_raw"),
            values.get("reported_valuation_raw"),
            values.get("latest_funding_round"),
            values.get("latest_funding_time_raw"),
            _parse_date(values.get("latest_funding_time_raw")),
            values.get("latest_funding_amount_raw"),
            _parse_decimal(values.get("latest_funding_cny_wan")),
            values.get("latest_funding_ratio"),
            values.get("latest_investors_raw"),
            values.get("latest_fa_info"),
            values.get("team_raw"),
            Jsonb(raw_row_json),
        ),
    ).fetchone()
    assert row is not None
    return row["snapshot_id"]


def _insert_team_members(
    conn: Connection,
    *,
    company_id: str,
    snapshot_id: UUID,
    team_raw: str | None,
) -> int:
    inserted = 0
    for member_order, member in enumerate(parse_team_raw(team_raw), start=1):
        conn.execute(
            """
            INSERT INTO company_team_member (
                company_id,
                snapshot_id,
                member_order,
                raw_name,
                raw_role,
                raw_intro,
                normalized_name
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """,
            (
                company_id,
                snapshot_id,
                member_order,
                member.raw_name,
                member.raw_role,
                member.raw_intro,
                normalize_person_name(member.raw_name),
            ),
        )
        inserted += 1
    return inserted


def _upsert_funding_event(
    conn: Connection,
    *,
    company_id: str,
    company_name: str | None,
    values: dict[str, str | None],
) -> int:
    funding_round = values.get("latest_funding_round")
    event_date = _parse_date(values.get("latest_funding_time_raw"))
    if not funding_round or event_date is None:
        return 0

    amount_cny_wan = _parse_decimal(values.get("latest_funding_cny_wan"))
    amount_cny_wan_text = (
        None if amount_cny_wan is None else format(amount_cny_wan, "f")
    )
    dedup_key = _funding_dedup_key(
        company_id=company_id,
        funding_round=funding_round,
        amount_cny_wan=amount_cny_wan_text,
        event_date=event_date,
    )
    is_new = _funding_event_exists(conn, company_id, dedup_key) is False

    event_subject_normalized = {
        "round": funding_round,
        "amount_raw": values.get("latest_funding_amount_raw"),
        "amount_cny_wan": amount_cny_wan_text,
        "investors_raw": values.get("latest_investors_raw"),
    }
    summary_bits = [company_name or company_id, funding_round]
    if values.get("latest_funding_amount_raw"):
        summary_bits.append(values["latest_funding_amount_raw"])
    event_summary = " ".join(summary_bits)

    conn.execute(
        """
        INSERT INTO company_signal_event (
            company_id,
            event_type,
            event_date,
            event_subject_normalized,
            event_summary,
            confidence,
            dedup_key
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (company_id, event_type, dedup_key) DO UPDATE
        SET event_date = EXCLUDED.event_date,
            event_subject_normalized = EXCLUDED.event_subject_normalized,
            event_summary = EXCLUDED.event_summary,
            confidence = EXCLUDED.confidence,
            updated_at = now()
        """,
        (
            company_id,
            "funding",
            event_date,
            Jsonb(event_subject_normalized),
            event_summary,
            Decimal("0.95"),
            dedup_key,
        ),
    )
    return 1 if is_new else 0


def _funding_event_exists(conn: Connection, company_id: str, dedup_key: str) -> bool:
    row = conn.execute(
        """
        SELECT 1 AS exists_flag
        FROM company_signal_event
        WHERE company_id = %s
          AND event_type = 'funding'
          AND dedup_key = %s
        """,
        (company_id, dedup_key),
    ).fetchone()
    return row is not None


def _funding_dedup_key(
    *,
    company_id: str,
    funding_round: str,
    amount_cny_wan: str | None,
    event_date: date,
) -> str:
    bucket_start = event_date.toordinal() // 14
    digest = sha256(
        f"{company_id}|funding|{funding_round}|{amount_cny_wan or ''}|{bucket_start}".encode(
            "utf-8"
        )
    ).hexdigest()
    return digest[:16]


def _insert_source_row_lineage(
    conn: Connection,
    *,
    batch_id: UUID,
    source_row_number: int,
    target_entity_id: str | None,
    resolution_status: str,
    resolution_reason: str | None,
    raw_row_json: dict[str, Any],
) -> None:
    conn.execute(
        """
        INSERT INTO source_row_lineage (
            batch_id,
            source_row_number,
            target_entity_type,
            target_entity_id,
            resolution_status,
            resolution_reason,
            raw_row_jsonb
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """,
        (
            batch_id,
            source_row_number,
            "company",
            target_entity_id,
            resolution_status,
            resolution_reason,
            Jsonb(raw_row_json),
        ),
    )


def _finalize_import_batch(
    conn: Connection,
    *,
    batch_id: UUID,
    rows_read: int,
    records_parsed: int,
    records_new: int,
    records_updated: int,
    records_failed: int,
    run_status: str,
) -> None:
    conn.execute(
        """
        UPDATE import_batch
        SET finished_at = now(),
            rows_read = %s,
            records_parsed = %s,
            records_new = %s,
            records_updated = %s,
            records_failed = %s,
            run_status = %s,
            error_summary = %s
        WHERE batch_id = %s
        """,
        (
            rows_read,
            records_parsed,
            records_new,
            records_updated,
            records_failed,
            run_status,
            Jsonb({"records_failed": records_failed}) if records_failed else Jsonb({}),
            batch_id,
        ),
    )


def _finalize_pipeline_run(
    conn: Connection,
    *,
    run_id: UUID,
    items_processed: int,
    items_failed: int,
    status: str,
) -> None:
    conn.execute(
        """
        UPDATE pipeline_run
        SET finished_at = now(),
            status = %s,
            items_processed = %s,
            items_failed = %s,
            error_summary = %s
        WHERE run_id = %s
        """,
        (
            status,
            items_processed,
            items_failed,
            Jsonb({"items_failed": items_failed}) if items_failed else Jsonb({}),
            run_id,
        ),
    )


def _derive_region_parts(
    region: str | None,
    registered_address: str | None,
) -> tuple[str | None, str | None, str | None]:
    if region:
        parts = [part.strip() for part in re.split(r"[-/]", region) if part.strip()]
        province = parts[0] if len(parts) >= 1 else None
        city = parts[1] if len(parts) >= 2 else None
        district = parts[2] if len(parts) >= 3 else None
        return province, city, district

    if registered_address and "深圳" in registered_address:
        return "广东省", "深圳市", None
    return None, None, None


def _normalize_cell(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text or text in _MISSING_MARKERS:
        return None
    return text


def _parse_bool(value: str | None) -> bool | None:
    if value is None:
        return None
    lowered = value.strip().lower()
    if lowered in _TRUE_SET:
        return True
    if lowered in _FALSE_SET:
        return False
    return None


def _parse_int(value: str | None) -> int | None:
    if not value:
        return None
    digits = re.sub(r"[^\d-]", "", value)
    if not digits or digits == "-":
        return None
    return int(digits)


def _parse_decimal(value: str | None) -> Decimal | None:
    if not value:
        return None
    normalized = value.replace(",", "").strip()
    try:
        return Decimal(normalized)
    except InvalidOperation:
        return None


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    for pattern in _DATE_PATTERNS:
        try:
            return datetime.strptime(value, pattern).date()
        except ValueError:
            continue
    return None


def _is_numeric_sequence(value: str | None) -> bool:
    if not value:
        return False
    return value.isdigit()
