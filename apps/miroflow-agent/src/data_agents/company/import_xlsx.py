from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
import re

from openpyxl import load_workbook

from ..normalization import normalize_company_name
from .models import (
    CompanyImportRecord,
    CompanyImportReport,
    CompanyImportResult,
    FinancingEvent,
)


HEADER_ALIASES: dict[str, str] = {
    "序号": "sequence_no",
    "项目名称": "project_name",
    "行业领域": "industry",
    "子领域": "sub_industry",
    "业务": "business",
    "地区": "region",
    "投资轮次": "financing_round",
    "投资时间": "financing_time",
    "投资金额": "financing_amount",
    "参考转化金额（万人民币）": "financing_amount_cny_wan",
    "比例": "financing_ratio",
    "投资方": "investor",
    "注册资金": "registered_capital",
    "高新企业": "is_high_tech",
    "简介": "description",
    "公司名称": "company_name",
    "企业名称": "company_name",
    "国别": "country",
    "成立日期": "established_date",
    "网址": "website",
    "法人代表": "legal_representative",
    "团队": "team_raw",
    "注册地址": "registered_address",
    "企业联系电话": "contact_phone",
    "联系邮箱": "contact_email",
    "专利数": "patent_count",
}

_MISSING_MARKERS = {"-", "--", "—", "－", "N/A", "n/a", "NULL", "null"}
_INVESTOR_SPLIT_RE = re.compile(r"[、,，;；/\n]+")
_TRUE_SET = {"是", "true", "1", "yes", "y"}
_FALSE_SET = {"否", "false", "0", "no", "n"}


@dataclass(slots=True)
class _MergedCompanyRow:
    source_row_numbers: list[int]
    values: dict[str, str | None]
    financing_events: list[FinancingEvent] = field(default_factory=list)
    investors: list[str] = field(default_factory=list)


def import_company_xlsx(
    xlsx_path: Path,
    *,
    sheet_name: str | None = None,
    max_header_scan_rows: int = 32,
) -> CompanyImportResult:
    source = Path(xlsx_path)
    workbook = load_workbook(source, data_only=True, read_only=True)
    try:
        active_sheet = workbook.active
        if active_sheet is None:
            raise ValueError(f"workbook has no active sheet: {source}")
        selected_sheet_name = sheet_name or active_sheet.title
        if selected_sheet_name not in workbook.sheetnames:
            available = ", ".join(workbook.sheetnames)
            raise ValueError(
                f"sheet '{selected_sheet_name}' not found in {source}. available sheets: {available}"
            )

        sheet = workbook[selected_sheet_name]
        header_row_index, column_mapping = _detect_header_row(
            sheet=sheet,
            max_header_scan_rows=max_header_scan_rows,
        )

        rows_read = 0
        rows_empty_skipped = 0
        rows_missing_company_name = 0
        continuation_rows_merged = 0
        orphan_continuation_rows = 0
        merged_rows: list[_MergedCompanyRow] = []
        current_row: _MergedCompanyRow | None = None

        for row_index, row in enumerate(
            sheet.iter_rows(min_row=header_row_index + 1, values_only=True),
            start=header_row_index + 1,
        ):
            mapped_values = _extract_mapped_values(row, column_mapping)
            if not any(mapped_values.values()):
                rows_empty_skipped += 1
                continue

            rows_read += 1
            sequence_no = mapped_values.get("sequence_no")
            company_name = mapped_values.get("company_name")

            if not sequence_no and not company_name:
                if current_row is None:
                    orphan_continuation_rows += 1
                    continue
                _merge_continuation_row(current_row, mapped_values, row_index)
                continuation_rows_merged += 1
                continue

            if not company_name:
                rows_missing_company_name += 1
                current_row = None
                continue

            current_row = _start_merged_row(mapped_values, row_index)
            merged_rows.append(current_row)

        parsed_records: list[CompanyImportRecord] = []
        for merged_row in merged_rows:
            company_name = merged_row.values.get("company_name")
            if not company_name:
                rows_missing_company_name += 1
                continue
            normalized_name = normalize_company_name(company_name)
            if not normalized_name:
                rows_missing_company_name += 1
                continue
            parsed_records.append(
                _build_company_record(
                    merged_row=merged_row,
                    normalized_name=normalized_name,
                )
            )

        deduped_records, duplicate_groups = _dedupe_records(parsed_records)
        report = CompanyImportReport(
            source_file=str(source),
            sheet_name=selected_sheet_name,
            header_row_index=header_row_index,
            rows_read=rows_read,
            rows_empty_skipped=rows_empty_skipped,
            rows_missing_company_name=rows_missing_company_name,
            continuation_rows_merged=continuation_rows_merged,
            orphan_continuation_rows=orphan_continuation_rows,
            company_rows_parsed=len(parsed_records),
            deduped_records=len(deduped_records),
            duplicate_groups=duplicate_groups,
            duplicate_records_discarded=len(parsed_records) - len(deduped_records),
        )
        return CompanyImportResult(records=deduped_records, report=report)
    finally:
        workbook.close()


def _detect_header_row(sheet, max_header_scan_rows: int) -> tuple[int, dict[int, str]]:
    best_row_index = 0
    best_mapping: dict[int, str] = {}
    best_score = -1

    for row_index, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=max_header_scan_rows, values_only=True),
        start=1,
    ):
        mapping: dict[int, str] = {}
        for column_index, value in enumerate(row):
            header = _normalize_cell(value)
            if not header:
                continue
            canonical = HEADER_ALIASES.get(header)
            if canonical is None:
                continue
            if canonical in mapping.values():
                continue
            mapping[column_index] = canonical

        score = len(mapping)
        if "company_name" not in mapping.values():
            continue
        if score > best_score:
            best_score = score
            best_row_index = row_index
            best_mapping = mapping

    if not best_mapping:
        raise ValueError("unable to detect header row containing 公司名称/企业名称")
    return best_row_index, best_mapping


def _extract_mapped_values(
    row: tuple[object, ...],
    column_mapping: dict[int, str],
) -> dict[str, str | None]:
    mapped_values: dict[str, str | None] = {}
    for column_index, key in column_mapping.items():
        value = row[column_index] if column_index < len(row) else None
        mapped_values[key] = _normalize_cell(value)
    return mapped_values


def _start_merged_row(values: dict[str, str | None], row_index: int) -> _MergedCompanyRow:
    merged = _MergedCompanyRow(
        source_row_numbers=[row_index],
        values=dict(values),
    )
    event = _build_financing_event(values)
    if event is not None:
        merged.financing_events.append(event)
    _append_investors(merged.investors, values.get("investor"))
    return merged


def _merge_continuation_row(
    merged_row: _MergedCompanyRow,
    values: dict[str, str | None],
    row_index: int,
) -> None:
    merged_row.source_row_numbers.append(row_index)

    for key, value in values.items():
        if value and not merged_row.values.get(key):
            merged_row.values[key] = value

    event = _build_financing_event(values)
    if event is not None:
        merged_row.financing_events.append(event)
    _append_investors(merged_row.investors, values.get("investor"))


def _build_financing_event(values: dict[str, str | None]) -> FinancingEvent | None:
    event = FinancingEvent(
        round=values.get("financing_round"),
        time=values.get("financing_time"),
        amount=values.get("financing_amount"),
        amount_cny_wan=values.get("financing_amount_cny_wan"),
        ratio=values.get("financing_ratio"),
        investor=values.get("investor"),
    )
    return None if event.is_empty() else event


def _build_company_record(
    merged_row: _MergedCompanyRow,
    normalized_name: str,
) -> CompanyImportRecord:
    values = merged_row.values
    return CompanyImportRecord(
        name=values["company_name"] or "",
        normalized_name=normalized_name,
        sequence_no=values.get("sequence_no"),
        project_name=values.get("project_name"),
        industry=values.get("industry"),
        sub_industry=values.get("sub_industry"),
        business=values.get("business"),
        region=values.get("region"),
        website=values.get("website"),
        legal_representative=values.get("legal_representative"),
        registered_capital=values.get("registered_capital"),
        description=values.get("description"),
        team_raw=values.get("team_raw"),
        registered_address=values.get("registered_address"),
        contact_phone=values.get("contact_phone"),
        contact_email=values.get("contact_email"),
        country=values.get("country"),
        established_date=values.get("established_date"),
        is_high_tech=_parse_bool(values.get("is_high_tech")),
        patent_count=_parse_int(values.get("patent_count")),
        financing_events=tuple(merged_row.financing_events),
        investors=tuple(merged_row.investors),
        source_row_numbers=tuple(merged_row.source_row_numbers),
    )


def _dedupe_records(
    records: list[CompanyImportRecord],
) -> tuple[list[CompanyImportRecord], int]:
    groups: dict[str, list[CompanyImportRecord]] = {}
    for record in records:
        groups.setdefault(record.normalized_name, []).append(record)

    duplicate_groups = sum(1 for grouped in groups.values() if len(grouped) > 1)
    winners = [_pick_most_complete(grouped) for grouped in groups.values()]
    winners.sort(
        key=lambda record: record.source_row_numbers[0]
        if record.source_row_numbers
        else 10**9
    )
    return winners, duplicate_groups


def _pick_most_complete(records: list[CompanyImportRecord]) -> CompanyImportRecord:
    return max(records, key=_record_rank)


def _record_rank(record: CompanyImportRecord) -> tuple[int, int, int, int, int]:
    first_row = record.source_row_numbers[0] if record.source_row_numbers else 10**9
    return (
        record.completeness_score(),
        len(record.source_row_numbers),
        len(record.financing_events),
        len(record.investors),
        -first_row,
    )


def _append_investors(existing: list[str], raw_investor: str | None) -> None:
    if not raw_investor:
        return
    for item in _split_multi_text(raw_investor):
        if item and item not in existing:
            existing.append(item)


def _split_multi_text(value: str) -> list[str]:
    parts = [_normalize_cell(part) for part in _INVESTOR_SPLIT_RE.split(value)]
    return [part for part in parts if part]


def _normalize_cell(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text in _MISSING_MARKERS:
        return None
    return text


def _parse_bool(value: str | None) -> bool | None:
    if value is None:
        return None
    lowered = value.strip().lower()
    if lowered in _TRUE_SET:
        return True
    if lowered in _FALSE_SET:
        return False
    return None


def _parse_int(value: str | None) -> int | None:
    if not value:
        return None
    digits = re.sub(r"[^\d]", "", value)
    if not digits:
        return None
    return int(digits)
