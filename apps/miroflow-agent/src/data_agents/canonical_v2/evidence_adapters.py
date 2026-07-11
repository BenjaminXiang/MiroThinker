"""Deterministic parsers for verified Canonical V2 landing inputs.

These adapters parse bytes supplied by an admitted offline workflow. They do not acquire data, open
original PostgreSQL/Milvus sources, or write canonical/publication state.
"""

from __future__ import annotations

import base64
import csv
from datetime import date, datetime, time
from io import BytesIO, StringIO
import json
from pathlib import Path
import sqlite3
import tempfile
from typing import Any

from openpyxl import load_workbook
from pydantic import JsonValue, ValidationError

from .contracts import ParseStatus, SourceError, SourceErrorKind
from .evidence_landing import (
    AdapterInput,
    ParsedRecordDraft,
    SourceAdapter,
    SourceAdapterError,
    UnverifiedSourceError,
)


_MISSING = object()


class _StrictJsonError(ValueError):
    """JSON bytes are syntactically accepted by stdlib but ambiguous as evidence."""


def _unique_json_object(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    value: dict[str, Any] = {}
    for key, item in pairs:
        if key in value:
            raise _StrictJsonError(f"JSON object contains duplicate key {key!r}")
        value[key] = item
    return value


def _reject_json_constant(value: str) -> None:
    raise _StrictJsonError(f"JSON contains non-standard numeric constant {value!r}")


def _load_json(content: bytes) -> Any:
    return json.loads(
        content,
        object_pairs_hook=_unique_json_object,
        parse_constant=_reject_json_constant,
    )


def _is_aware_iso_timestamp(value: Any) -> bool:
    if not isinstance(value, str) or not value.strip():
        return False
    candidate = f"{value[:-1]}+00:00" if value.endswith("Z") else value
    try:
        parsed = datetime.fromisoformat(candidate)
    except ValueError:
        return False
    return parsed.utcoffset() is not None


def _error(
    *,
    code: str,
    kind: SourceErrorKind,
    message: str,
    field_path: str | None = None,
    recoverable: bool = False,
) -> SourceError:
    return SourceError(
        error_code=code,
        error_kind=kind,
        message=message,
        field_path=field_path,
        recoverable=recoverable,
    )


def _corrupt(locator: str, message: str) -> ParsedRecordDraft:
    return ParsedRecordDraft(
        record_locator=locator,
        parse_status=ParseStatus.corrupt,
        payload={},
        errors=(
            _error(
                code="corrupt_content",
                kind=SourceErrorKind.corrupt_content,
                message=message,
            ),
        ),
    )


def _unsupported(locator: str, message: str) -> ParsedRecordDraft:
    return ParsedRecordDraft(
        record_locator=locator,
        parse_status=ParseStatus.unsupported,
        payload={},
        errors=(
            _error(
                code="unsupported_format",
                kind=SourceErrorKind.unsupported_format,
                message=message,
            ),
        ),
    )


def _schema_error(
    locator: str,
    message: str,
    *,
    payload: dict[str, JsonValue] | None = None,
) -> ParsedRecordDraft:
    return ParsedRecordDraft(
        record_locator=locator,
        parse_status=ParseStatus.quarantined,
        payload=payload or {},
        errors=(
            _error(
                code="schema_mismatch",
                kind=SourceErrorKind.schema_mismatch,
                message=message,
            ),
        ),
    )


def _field_path(parent: str | None, child: str) -> str:
    return f"{parent}.{child}" if parent else child


def _strip_unreadable(
    value: JsonValue,
    *,
    path: str | None = None,
) -> tuple[JsonValue | object, list[SourceError]]:
    if isinstance(value, dict):
        marker = value.get("$unreadable_external")
        if marker is not None:
            locator = str(marker)
            return (
                _MISSING,
                [
                    _error(
                        code="missing_external_content",
                        kind=SourceErrorKind.missing_external_content,
                        message=f"External content is unreadable at {locator}.",
                        field_path=path,
                        recoverable=True,
                    )
                ],
            )
        cleaned: dict[str, JsonValue] = {}
        errors: list[SourceError] = []
        for key, child in value.items():
            child_value, child_errors = _strip_unreadable(
                child,
                path=_field_path(path, key),
            )
            errors.extend(child_errors)
            if child_value is not _MISSING:
                cleaned[key] = child_value  # type: ignore[assignment]
        return cleaned, errors
    if isinstance(value, list):
        cleaned_list: list[JsonValue] = []
        errors = []
        for index, child in enumerate(value):
            child_value, child_errors = _strip_unreadable(
                child,
                path=_field_path(path, str(index)),
            )
            errors.extend(child_errors)
            if child_value is not _MISSING:
                cleaned_list.append(child_value)  # type: ignore[arg-type]
        return cleaned_list, errors
    return value, []


def _mapping_draft(
    value: Any,
    *,
    locator: str,
) -> ParsedRecordDraft:
    if not isinstance(value, dict):
        return _schema_error(locator, "A source record must be a JSON object.")
    cleaned, errors = _strip_unreadable(value)
    if not isinstance(cleaned, dict):
        return _schema_error(locator, "A source record must retain an object payload.")
    return ParsedRecordDraft(
        record_locator=locator,
        parse_status=ParseStatus.partial if errors else ParseStatus.parsed,
        payload=cleaned,
        errors=tuple(errors),
    )


def _json_lines(content: bytes) -> tuple[ParsedRecordDraft, ...]:
    records: list[ParsedRecordDraft] = []
    for line_number, raw_line in enumerate(content.splitlines(), start=1):
        locator = f"line:{line_number}"
        if not raw_line.strip():
            continue
        try:
            value = _load_json(raw_line)
        except (UnicodeDecodeError, json.JSONDecodeError, _StrictJsonError) as exc:
            records.append(_corrupt(locator, f"JSON line cannot be decoded: {exc}."))
            continue
        records.append(_mapping_draft(value, locator=locator))
    return tuple(records)


class _KindBoundAdapter:
    parser_name: str
    allowed_source_kinds: frozenset[str]

    def validate_source(self, value: AdapterInput) -> None:
        if value.source_kind not in self.allowed_source_kinds:
            raise SourceAdapterError(
                f"parser {self.parser_name} does not accept source kind {value.source_kind}"
            )


class HistoricalJsonlAdapter(_KindBoundAdapter):
    parser_name = "historical_jsonl"
    allowed_source_kinds = frozenset(
        {"historical_jsonl", "forensic_source", "verified_copy"}
    )

    def parse(self, value: AdapterInput) -> tuple[ParsedRecordDraft, ...]:
        return _json_lines(value.content)


class HistoricalJsonAdapter(_KindBoundAdapter):
    parser_name = "historical_json"
    allowed_source_kinds = frozenset({"historical_json"})

    def parse(self, value: AdapterInput) -> tuple[ParsedRecordDraft, ...]:
        try:
            decoded = _load_json(value.content)
        except (UnicodeDecodeError, json.JSONDecodeError, _StrictJsonError) as exc:
            return (_corrupt("$", f"JSON document cannot be decoded: {exc}."),)
        items = decoded if isinstance(decoded, list) else [decoded]
        return tuple(
            _mapping_draft(item, locator=f"item:{index}")
            for index, item in enumerate(items, start=1)
        )


class HistoricalCsvAdapter(_KindBoundAdapter):
    parser_name = "historical_csv"
    allowed_source_kinds = frozenset({"historical_csv"})

    def parse(self, value: AdapterInput) -> tuple[ParsedRecordDraft, ...]:
        try:
            text = value.content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            return (_corrupt("$", f"CSV bytes are not UTF-8: {exc}."),)
        reader = csv.DictReader(StringIO(text))
        if not reader.fieldnames:
            return (
                _unsupported("header", "CSV input requires a non-empty header row."),
            )
        normalized_headers = tuple(name.strip() for name in reader.fieldnames)
        if any(not name for name in normalized_headers):
            return (_schema_error("header", "CSV header names must be non-empty."),)
        if len(normalized_headers) != len(set(normalized_headers)):
            return (_schema_error("header", "CSV header names must be unique."),)
        records: list[ParsedRecordDraft] = []
        for row_number, row in enumerate(reader, start=2):
            extra_cells = row.get(None)
            payload = {
                str(key): cell
                for key, cell in row.items()
                if key is not None and cell not in {None, ""}
            }
            errors = (
                (
                    _error(
                        code="unheaded_csv_values",
                        kind=SourceErrorKind.schema_mismatch,
                        message=f"CSV row contains {len(extra_cells)} value(s) without headers.",
                        field_path="$extra_columns",
                        recoverable=True,
                    ),
                )
                if extra_cells
                else ()
            )
            records.append(
                ParsedRecordDraft(
                    record_locator=f"row:{row_number}",
                    parse_status=ParseStatus.partial if errors else ParseStatus.parsed,
                    payload=payload,
                    errors=errors,
                )
            )
        return tuple(records)


def _cell_value(
    value: Any, *, field_path: str
) -> tuple[JsonValue | object, SourceError | None]:
    if value is None or value == "":
        return _MISSING, None
    if isinstance(value, (str, int, float, bool)):
        return value, None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat(), None
    if isinstance(value, bytes):
        return (
            {"$base64": base64.b64encode(value).decode("ascii")},
            None,
        )
    return (
        _MISSING,
        _error(
            code="unsupported_cell_value",
            kind=SourceErrorKind.unsupported_format,
            message=f"Cell value type {type(value).__name__} is unsupported.",
            field_path=field_path,
        ),
    )


class HistoricalXlsxAdapter(_KindBoundAdapter):
    parser_name = "historical_xlsx"
    allowed_source_kinds = frozenset({"historical_xlsx"})

    def parse(self, value: AdapterInput) -> tuple[ParsedRecordDraft, ...]:
        try:
            workbook = load_workbook(
                BytesIO(value.content),
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except (
            Exception
        ) as exc:  # openpyxl raises several format-specific exception types
            return (_corrupt("$", f"XLSX workbook cannot be opened: {exc}."),)
        try:
            requested_sheet = value.parser.options.get("sheet")
            sheet_name = (
                str(requested_sheet) if requested_sheet else workbook.sheetnames[0]
            )
            if sheet_name not in workbook.sheetnames:
                return (
                    _schema_error(
                        "workbook",
                        f"Requested XLSX sheet {sheet_name} is absent.",
                    ),
                )
            sheet = workbook[sheet_name]
            rows = sheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            if not header_row:
                return (
                    _unsupported(f"sheet:{sheet_name}", "XLSX sheet has no header."),
                )
            headers = tuple(
                str(cell).strip() if cell is not None else "" for cell in header_row
            )
            if any(not header for header in headers):
                return (
                    _schema_error(
                        f"sheet:{sheet_name}:row:1",
                        "XLSX header cells must be non-empty.",
                    ),
                )
            if len(headers) != len(set(headers)):
                return (
                    _schema_error(
                        f"sheet:{sheet_name}:row:1",
                        "XLSX header names must be unique.",
                    ),
                )
            records: list[ParsedRecordDraft] = []
            for row_number, row in enumerate(rows, start=2):
                payload: dict[str, JsonValue] = {}
                errors: list[SourceError] = []
                for header, cell in zip(headers, row, strict=False):
                    normalized, error = _cell_value(cell, field_path=header)
                    if error is not None:
                        errors.append(error)
                    elif normalized is not _MISSING:
                        payload[header] = normalized  # type: ignore[assignment]
                records.append(
                    ParsedRecordDraft(
                        record_locator=f"sheet:{sheet_name}:row:{row_number}",
                        parse_status=ParseStatus.partial
                        if errors
                        else ParseStatus.parsed,
                        payload=payload,
                        errors=tuple(errors),
                    )
                )
            return tuple(records)
        finally:
            workbook.close()


class HistoricalSqliteAdapter(_KindBoundAdapter):
    parser_name = "historical_sqlite"
    allowed_source_kinds = frozenset({"historical_sqlite"})

    def parse(self, value: AdapterInput) -> tuple[ParsedRecordDraft, ...]:
        table_option = value.parser.options.get("table")
        if not isinstance(table_option, str) or not table_option.strip():
            raise SourceAdapterError(
                "historical_sqlite requires a non-empty table option"
            )
        table = table_option.strip()
        temporary_path: Path | None = None
        try:
            with tempfile.NamedTemporaryFile(
                prefix="canonical-v2-landing-",
                suffix=".sqlite",
                delete=False,
            ) as temporary:
                temporary.write(value.content)
                temporary_path = Path(temporary.name)
            connection = sqlite3.connect(
                f"file:{temporary_path.as_posix()}?mode=ro&immutable=1",
                uri=True,
            )
            connection.row_factory = sqlite3.Row
            try:
                exists = connection.execute(
                    "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
                    (table,),
                ).fetchone()
                if exists is None:
                    return (
                        _schema_error(
                            f"table:{table}",
                            f"SQLite table {table} is absent.",
                        ),
                    )
                quoted_table = table.replace('"', '""')
                rows = connection.execute(f'SELECT * FROM "{quoted_table}"').fetchall()
                records: list[ParsedRecordDraft] = []
                for row_number, row in enumerate(rows, start=1):
                    payload: dict[str, JsonValue] = {}
                    errors: list[SourceError] = []
                    for key in row.keys():
                        normalized, error = _cell_value(row[key], field_path=str(key))
                        if error is not None:
                            errors.append(error)
                        elif normalized is not _MISSING:
                            payload[str(key)] = normalized  # type: ignore[assignment]
                    records.append(
                        ParsedRecordDraft(
                            record_locator=f"table:{table}:row:{row_number}",
                            parse_status=(
                                ParseStatus.partial if errors else ParseStatus.parsed
                            ),
                            payload=payload,
                            errors=tuple(errors),
                        )
                    )
                return tuple(records)
            finally:
                connection.close()
        except sqlite3.DatabaseError as exc:
            return (_corrupt("$", f"SQLite database cannot be read: {exc}."),)
        finally:
            if temporary_path is not None:
                temporary_path.unlink(missing_ok=True)


class WalFpiSalvageAdapter(_KindBoundAdapter):
    parser_name = "wal_fpi_salvage"
    allowed_source_kinds = frozenset({"wal_fpi_salvage_records"})

    def parse(self, value: AdapterInput) -> tuple[ParsedRecordDraft, ...]:
        records: list[ParsedRecordDraft] = []
        for line_number, raw_line in enumerate(value.content.splitlines(), start=1):
            default_locator = f"line:{line_number}"
            if not raw_line.strip():
                continue
            try:
                envelope = _load_json(raw_line)
            except (UnicodeDecodeError, json.JSONDecodeError, _StrictJsonError) as exc:
                records.append(
                    _corrupt(
                        default_locator, f"Salvage envelope is invalid JSON: {exc}."
                    )
                )
                continue
            if not isinstance(envelope, dict):
                records.append(
                    _schema_error(
                        default_locator, "Salvage envelope must be an object."
                    )
                )
                continue
            locator_value = envelope.get("record_locator", default_locator)
            locator = str(locator_value) if locator_value else default_locator
            readable_fields = envelope.get("readable_fields")
            raw_errors = envelope.get("field_errors", [])
            if not isinstance(readable_fields, dict) or not isinstance(
                raw_errors, list
            ):
                records.append(
                    _schema_error(
                        locator,
                        "Salvage envelope requires readable_fields object and field_errors list.",
                    )
                )
                continue
            try:
                errors = tuple(SourceError.model_validate(item) for item in raw_errors)
            except ValidationError as exc:
                records.append(
                    _schema_error(locator, f"Salvage field error is invalid: {exc}.")
                )
                continue
            records.append(
                ParsedRecordDraft(
                    record_locator=locator,
                    parse_status=ParseStatus.partial if errors else ParseStatus.parsed,
                    payload=readable_fields,
                    errors=errors,
                )
            )
        return tuple(records)


class MilvusCopyRecordsAdapter(_KindBoundAdapter):
    parser_name = "milvus_copy_records"
    allowed_source_kinds = frozenset({"milvus_verified_copy_records"})

    def validate_source(self, value: AdapterInput) -> None:
        if value.source_kind != "milvus_verified_copy_records":
            raise UnverifiedSourceError(
                "Milvus records must come from a verified copy; original Milvus is forbidden"
            )

    def parse(self, value: AdapterInput) -> tuple[ParsedRecordDraft, ...]:
        records = list(_json_lines(value.content))
        validated: list[ParsedRecordDraft] = []
        for line_number, record in enumerate(records, start=1):
            if record.parse_status is not ParseStatus.parsed:
                validated.append(record)
                continue
            collection = record.payload.get("collection")
            primary_key = record.payload.get("primary_key")
            valid_collection = isinstance(collection, str) and bool(collection.strip())
            valid_primary_key = (
                isinstance(primary_key, str) and bool(primary_key.strip())
            ) or (isinstance(primary_key, int) and not isinstance(primary_key, bool))
            if not valid_collection or not valid_primary_key:
                validated.append(
                    _schema_error(
                        record.record_locator,
                        "Milvus copy record requires collection and primary_key.",
                        payload=record.payload,
                    )
                )
                continue
            validated.append(
                record.model_copy(
                    update={
                        "record_locator": (
                            f"collection:{collection}:primary_key:{primary_key}:line:{line_number}"
                        )
                    }
                )
            )
        return tuple(validated)


class CollectedResponseAdapter(_KindBoundAdapter):
    parser_name = "collected_response"
    allowed_source_kinds = frozenset({"newly_collected_response"})
    required_fields = frozenset(
        {"source_url", "retrieved_at", "status_code", "content_type", "body"}
    )

    def parse(self, value: AdapterInput) -> tuple[ParsedRecordDraft, ...]:
        try:
            envelope = _load_json(value.content)
        except (UnicodeDecodeError, json.JSONDecodeError, _StrictJsonError) as exc:
            return (_corrupt("$", f"Collected response is invalid JSON: {exc}."),)
        if not isinstance(envelope, dict):
            return (_schema_error("$", "Collected response must be an object."),)
        source_url = envelope.get("source_url")
        locator = (
            source_url if isinstance(source_url, str) and source_url.strip() else "$"
        )
        missing = sorted(self.required_fields - set(envelope))
        errors = [
            _error(
                code="schema_mismatch",
                kind=SourceErrorKind.schema_mismatch,
                message=f"Collected response field {field} is missing.",
                field_path=field,
            )
            for field in missing
        ]
        field_values = {
            "source_url": source_url,
            "retrieved_at": envelope.get("retrieved_at"),
            "status_code": envelope.get("status_code"),
            "content_type": envelope.get("content_type"),
        }
        validators = {
            "source_url": lambda item: isinstance(item, str) and bool(item.strip()),
            "retrieved_at": _is_aware_iso_timestamp,
            "status_code": lambda item: (
                isinstance(item, int)
                and not isinstance(item, bool)
                and 100 <= item <= 599
            ),
            "content_type": lambda item: isinstance(item, str) and bool(item.strip()),
        }
        for field, validator in validators.items():
            if field in envelope and not validator(field_values[field]):
                errors.append(
                    _error(
                        code="schema_mismatch",
                        kind=SourceErrorKind.schema_mismatch,
                        message=f"Collected response field {field} has an invalid value.",
                        field_path=field,
                    )
                )
        if errors:
            return (
                ParsedRecordDraft(
                    record_locator=locator,
                    parse_status=ParseStatus.partial,
                    payload=envelope,
                    errors=tuple(errors),
                ),
            )
        return (
            ParsedRecordDraft(
                record_locator=locator,
                parse_status=ParseStatus.parsed,
                payload=envelope,
            ),
        )


def default_source_adapters() -> tuple[SourceAdapter, ...]:
    return (
        HistoricalJsonlAdapter(),
        HistoricalJsonAdapter(),
        HistoricalCsvAdapter(),
        HistoricalXlsxAdapter(),
        HistoricalSqliteAdapter(),
        WalFpiSalvageAdapter(),
        MilvusCopyRecordsAdapter(),
        CollectedResponseAdapter(),
    )


__all__ = [
    "CollectedResponseAdapter",
    "HistoricalCsvAdapter",
    "HistoricalJsonAdapter",
    "HistoricalJsonlAdapter",
    "HistoricalSqliteAdapter",
    "HistoricalXlsxAdapter",
    "MilvusCopyRecordsAdapter",
    "WalFpiSalvageAdapter",
    "default_source_adapters",
]
