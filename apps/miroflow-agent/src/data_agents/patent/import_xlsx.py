from __future__ import annotations

from collections import Counter
from datetime import date, datetime
from pathlib import Path
import re
from typing import Any
import warnings

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .models import PatentImportRecord, PatentImportReport, PatentImportResult


_WS_RE = re.compile(r"\s+")
_HEADER_WS_RE = re.compile(r"[\s\u3000]+")
_TOKEN_SPLIT_RE = re.compile(r"[;；\n]+")
_DATE_RE = re.compile(r"(\d{4})[./-](\d{1,2})[./-](\d{1,2})")

_COLUMN_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "sequence_number": ("序号",),
    "title": ("标题 (中文)", "标题（中文）", "标题(中文)", "标题"),
    "title_en": ("标题 (英文)", "标题（英文）", "标题(英文)"),
    "abstract": ("摘要 (中文)", "摘要（中文）", "摘要(中文)"),
    "abstract_en": ("摘要 (英文)", "摘要（英文）", "摘要(英文)"),
    "applicants": ("申请人",),
    "patent_number": ("公开（公告）号", "公开(公告)号", "公开号", "专利号"),
    "publication_date": ("公开（公告）日", "公开(公告)日"),
    "filing_date": ("申请日",),
    "patent_type": ("专利类型",),
    "technology_effect_sentence": ("技术功效句",),
    "technology_effect_phrases": ("技术功效短语",),
    "expected_expiry_date": ("预估到期日",),
}
_REQUIRED_COLUMNS = ("title", "applicants")


_HEADER_LOOKUP: dict[str, str] = {}
for canonical_key, aliases in _COLUMN_HEADER_ALIASES.items():
    for alias in aliases:
        _HEADER_LOOKUP[_HEADER_WS_RE.sub("", alias)] = canonical_key


def import_patent_xlsx(path: Path) -> PatentImportResult:
    workbook_path = Path(path)
    read_only_max_row, read_only_max_column = _probe_read_only_dimensions(workbook_path)

    workbook = _load_workbook_without_style_warning(
        workbook_path,
        read_only=False,
        data_only=True,
    )
    try:
        worksheet = workbook.active
        if worksheet is None:
            raise ValueError(f"workbook has no active worksheet: {workbook_path}")
        header_row_index, column_indexes = _locate_header_row(worksheet)

        records: list[PatentImportRecord] = []
        skipped_counter: Counter[str] = Counter()
        rows_read = 0

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=header_row_index + 1, values_only=True),
            start=header_row_index + 1,
        ):
            if _is_empty_row(row):
                continue
            rows_read += 1
            record = _parse_record(row, row_index=row_index, column_indexes=column_indexes)
            if isinstance(record, PatentImportRecord):
                records.append(record)
                continue
            skipped_counter[record] += 1

        report = PatentImportReport(
            workbook_path=str(workbook_path),
            worksheet_title=worksheet.title,
            header_row_index=header_row_index,
            rows_read=rows_read,
            records_parsed=len(records),
            skipped_rows=rows_read - len(records),
            skip_reasons=dict(skipped_counter),
            read_only_max_row=read_only_max_row,
            read_only_max_column=read_only_max_column,
        )
        return PatentImportResult(records=records, report=report)
    finally:
        workbook.close()


def _probe_read_only_dimensions(path: Path) -> tuple[int, int]:
    workbook = _load_workbook_without_style_warning(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            raise ValueError(f"workbook has no active worksheet: {path}")
        return sheet.max_row, sheet.max_column
    finally:
        workbook.close()


def _load_workbook_without_style_warning(path: Path, **kwargs: Any) -> Workbook:
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style, apply openpyxl's default",
            category=UserWarning,
            module="openpyxl.styles.stylesheet",
        )
        return load_workbook(path, **kwargs)


def _locate_header_row(sheet: Worksheet) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        column_indexes = _resolve_header_columns(row)
        if not column_indexes:
            continue
        if all(required in column_indexes for required in _REQUIRED_COLUMNS):
            return row_index, column_indexes
    raise ValueError("unable to locate patent header row")


def _resolve_header_columns(row: tuple[object, ...]) -> dict[str, int]:
    column_indexes: dict[str, int] = {}
    for column_index, value in enumerate(row):
        label = _normalize_header_label(value)
        if label is None:
            continue
        canonical_key = _HEADER_LOOKUP.get(label)
        if canonical_key is None:
            continue
        column_indexes.setdefault(canonical_key, column_index)
    return column_indexes


def _normalize_header_label(value: object) -> str | None:
    text = _normalize_text(value)
    if text is None:
        return None
    return _HEADER_WS_RE.sub("", text)


def _is_empty_row(row: tuple[object, ...]) -> bool:
    return all(_normalize_text(cell) is None for cell in row)


def _parse_record(
    row: tuple[object, ...],
    *,
    row_index: int,
    column_indexes: dict[str, int],
) -> PatentImportRecord | str:
    title = _normalize_text(_get_cell(row, column_indexes, "title"))
    if title is None:
        return "missing_title"

    applicants = _split_tokens(_get_cell(row, column_indexes, "applicants"))
    if not applicants:
        return "missing_applicants"

    return PatentImportRecord(
        source_row=row_index,
        sequence_number=_normalize_sequence_number(
            _get_cell(row, column_indexes, "sequence_number")
        ),
        title=title,
        title_en=_normalize_text(_get_cell(row, column_indexes, "title_en")),
        abstract=_normalize_text(_get_cell(row, column_indexes, "abstract")),
        abstract_en=_normalize_text(_get_cell(row, column_indexes, "abstract_en")),
        applicants=applicants,
        patent_number=_normalize_patent_number(
            _get_cell(row, column_indexes, "patent_number")
        ),
        publication_date=_parse_date(_get_cell(row, column_indexes, "publication_date")),
        filing_date=_parse_date(_get_cell(row, column_indexes, "filing_date")),
        patent_type=_normalize_patent_type(_get_cell(row, column_indexes, "patent_type")),
        technology_effect_sentence=_normalize_text(
            _get_cell(row, column_indexes, "technology_effect_sentence")
        ),
        technology_effect_phrases=_split_tokens(
            _get_cell(row, column_indexes, "technology_effect_phrases")
        ),
        expected_expiry_date=_parse_date(
            _get_cell(row, column_indexes, "expected_expiry_date")
        ),
    )


def _get_cell(
    row: tuple[object, ...],
    column_indexes: dict[str, int],
    key: str,
) -> object | None:
    column_index = column_indexes.get(key)
    if column_index is None:
        return None
    if column_index >= len(row):
        return None
    return row[column_index]


def _normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return _WS_RE.sub(" ", text)


def _normalize_sequence_number(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = _normalize_text(value)
    if text is None:
        return None
    if text.endswith(".0"):
        head = text[:-2]
        if head.isdigit():
            return head
    return text


def _normalize_patent_number(value: object) -> str | None:
    text = _normalize_text(value)
    if text is None:
        return None
    return text.replace(" ", "").upper()


def _normalize_patent_type(value: object) -> str | None:
    text = _normalize_text(value)
    if text is None:
        return None
    if "实用新型" in text:
        return "实用新型"
    if "外观" in text:
        return "外观设计"
    if "发明" in text:
        return "发明"
    return text


def _split_tokens(value: object) -> tuple[str, ...]:
    text = _normalize_text(value)
    if text is None:
        return tuple()
    tokens = [
        _WS_RE.sub(" ", token).strip()
        for token in _TOKEN_SPLIT_RE.split(text)
        if token and token.strip()
    ]
    deduplicated: list[str] = []
    seen: set[str] = set()
    for token in tokens:
        if token in seen:
            continue
        seen.add(token)
        deduplicated.append(token)
    return tuple(deduplicated)


def _parse_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            parsed = from_excel(value)
        except Exception:
            parsed = None
        if isinstance(parsed, datetime):
            return parsed.date()
        if isinstance(parsed, date):
            return parsed

    text = _normalize_text(value)
    if text is None:
        return None
    match = _DATE_RE.search(text)
    if match is None:
        return None
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None
