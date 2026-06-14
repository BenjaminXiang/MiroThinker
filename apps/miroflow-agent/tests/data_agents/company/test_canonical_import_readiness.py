from __future__ import annotations

from pathlib import Path
from uuid import UUID

from openpyxl import Workbook

from src.data_agents.company._company_id import generate_company_id
from src.data_agents.company.canonical_import import (
    build_company_import_preflight,
    _evaluate_xlsx_baseline_readiness,
    _load_merged_rows,
    _upsert_company,
)


class _Result:
    def __init__(self, rows: list[dict]) -> None:
        self._rows = rows

    def fetchone(self):
        return self._rows[0] if self._rows else None


class _Conn:
    def __init__(self) -> None:
        self.calls: list[tuple[str, tuple | None]] = []

    def execute(self, query: str, params: tuple | None = None) -> _Result:
        sql = " ".join(query.split())
        self.calls.append((sql, params))
        if sql.startswith("SELECT 1 AS exists_flag"):
            return _Result([])
        return _Result([])


def test_xlsx_baseline_readiness_promotes_resolved_meaningful_row() -> None:
    readiness = _evaluate_xlsx_baseline_readiness(
        {
            "company_name_xlsx": "深圳示例科技有限公司",
            "industry": "医疗AI",
            "description": "公司提供面向医院的AI辅助诊断平台。",
        }
    )

    assert readiness.quality_status == "ready"
    assert readiness.blockers == ()


def test_xlsx_baseline_readiness_keeps_sparse_row_reviewable() -> None:
    readiness = _evaluate_xlsx_baseline_readiness(
        {"company_name_xlsx": "深圳示例科技有限公司"}
    )

    assert readiness.quality_status == "needs_review"
    assert "missing_meaningful_baseline_field" in readiness.blockers


def test_xlsx_baseline_readiness_records_unresolved_identity_blocker() -> None:
    readiness = _evaluate_xlsx_baseline_readiness(
        {
            "company_name_xlsx": "深圳示例科技有限公司",
            "industry": "医疗AI",
            "description": "公司提供面向医院的AI辅助诊断平台。",
        },
        identity_status="needs_review",
    )

    assert readiness.quality_status == "needs_review"
    assert "unresolved_identity" in readiness.blockers


def test_upsert_company_persists_ready_quality_status_from_xlsx_baseline() -> None:
    conn = _Conn()

    company_id, is_new = _upsert_company(
        conn,
        batch_id=UUID("aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa"),
        values={
            "company_name_xlsx": "深圳示例科技有限公司",
            "industry": "医疗AI",
            "description": "公司提供面向医院的AI辅助诊断平台。",
            "region": "深圳",
        },
        run_id=UUID("bbbbbbbb-bbbb-bbbb-bbbb-bbbbbbbbbbbb"),
    )

    insert_sql, insert_params = next(
        call for call in conn.calls if call[0].startswith("INSERT INTO company")
    )
    assert company_id
    assert is_new is True
    assert "quality_status" in insert_sql
    assert insert_params is not None
    assert "ready" in insert_params


def test_enterprise_master_headers_are_mapped_for_canonical_import(tmp_path: Path) -> None:
    workbook_path = tmp_path / "enterprise_master.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "公司名称",
            "项目名称",
            "行业",
            "省份地区",
            "业务",
            "成立时间",
            "法定代表人",
            "团队",
            "企业类型",
            "注册地址",
            "网址",
            "邮箱",
            "企业联系电话",
            "产品简介",
            "产品特点",
            "应用场景",
        ]
    )
    sheet.append(
        [
            "深圳示例机器人有限公司",
            "示例机器人",
            "机器人",
            "广东省深圳市",
            "工业巡检机器人研发商",
            "2020-01-02",
            "张三",
            "张三，创始人",
            "有限责任公司",
            "深圳市南山区",
            "https://example.com",
            "team@example.com",
            "13800000000",
            "示例机器人平台用于工厂巡检。",
            "具备机器视觉和自主导航能力。",
            "工业园区巡检、设备监测",
        ]
    )
    workbook.save(workbook_path)

    rows_read, rows = _load_merged_rows(workbook_path)

    assert rows_read == 1
    values = rows[0].values
    assert values["industry"] == "机器人"
    assert values["region"] == "广东省深圳市"
    assert values["established_date"] == "2020-01-02"
    assert values["legal_representative"] == "张三"
    assert values["contact_email"] == "team@example.com"
    assert values["product_intro"] == "示例机器人平台用于工厂巡检。"
    assert values["product_features"] == "具备机器视觉和自主导航能力。"
    assert values["application_scenarios_raw"] == "工业园区巡检、设备监测"


def test_shared_platform_domains_do_not_become_company_identity_anchor() -> None:
    first = generate_company_id(
        unified_credit_code=None,
        website="https://weibo.com/company-a",
        registered_name="深圳甲科技有限公司",
    )
    second = generate_company_id(
        unified_credit_code=None,
        website="https://weibo.com/company-b",
        registered_name="深圳乙科技有限公司",
    )

    assert first != second
    assert first == generate_company_id(
        unified_credit_code=None,
        website=None,
        registered_name="深圳甲科技有限公司",
    )


def test_company_import_preflight_reports_duplicate_identity_conflicts(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "conflict.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["公司名称", "项目名称", "网址", "产品简介", "应用场景"])
    sheet.append(
        [
            "深圳甲科技有限公司",
            "甲科技",
            "https://same.example.com",
            "甲科技平台提供工业巡检能力。",
            "工业巡检",
        ]
    )
    sheet.append(
        [
            "深圳乙科技有限公司",
            "乙科技",
            "https://same.example.com",
            "乙科技平台提供设备监测能力。",
            "设备监测",
        ]
    )
    workbook.save(workbook_path)

    report = build_company_import_preflight(workbook_path)

    assert report["rows_read"] == 2
    assert report["records_parsed"] == 2
    assert report["identity_conflict_count"] == 2
    assert report["duplicate_generated_company_id_groups"] == 1
    assert report["field_coverage"]["product_intro"] == 2
    assert report["field_coverage"]["application_scenarios_raw"] == 2


def test_company_import_preflight_reports_existing_company_diffs(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "overlap.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["公司名称", "网址", "产品简介"])
    sheet.append(["深圳甲科技有限公司", "https://new.example.com", "甲科技平台。"])
    workbook.save(workbook_path)

    generated_id = generate_company_id(
        unified_credit_code=None,
        website="https://new.example.com",
        registered_name="深圳甲科技有限公司",
    )
    report = build_company_import_preflight(
        workbook_path,
        existing_companies={
            generated_id: {
                "registered_name": "广州乙科技有限公司",
                "website": "https://old.example.com",
            }
        },
    )

    assert report["matched_existing_count"] == 1
    assert report["overlap_diff_count"] == 1
    assert report["overlap_diff_samples"] == [
        {
            "generated_company_id": generated_id,
            "source_row_number": 2,
            "company_name": "深圳甲科技有限公司",
                "diffs": {
                    "registered_name": {
                        "existing": "广州乙科技有限公司",
                        "incoming": "深圳甲科技有限公司",
                    },
                "website": {
                    "existing": "https://old.example.com",
                    "incoming": "https://new.example.com",
                },
            },
        }
    ]
