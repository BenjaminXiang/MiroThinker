from __future__ import annotations

from datetime import date
from pathlib import Path
import warnings

from openpyxl import Workbook, load_workbook

from src.data_agents.patent.import_xlsx import import_patent_xlsx


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[5]


def _real_patent_xlsx_path() -> Path:
    return _repo_root() / "docs" / "2025-12-05 专利.xlsx"


def test_import_patent_xlsx_reads_real_rows_despite_broken_read_only_dimensions():
    input_path = _real_patent_xlsx_path()
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style, apply openpyxl's default",
            category=UserWarning,
            module="openpyxl.styles.stylesheet",
        )
        read_only_workbook = load_workbook(input_path, read_only=True, data_only=True)
    read_only_sheet = read_only_workbook.active
    assert (read_only_sheet.max_row, read_only_sheet.max_column) == (1, 1)
    read_only_workbook.close()

    result = import_patent_xlsx(input_path)

    assert result.report.header_row_index == 1
    assert result.report.rows_read == 1930
    assert result.report.records_parsed == 1930
    assert result.report.skipped_rows == 0
    assert len(result.records) == 1930


def test_import_patent_xlsx_parses_and_normalizes_first_real_record():
    result = import_patent_xlsx(_real_patent_xlsx_path())
    first = result.records[0]

    assert first.source_row == 2
    assert first.sequence_number == "10001"
    assert first.title == "无人机的机臂锁紧组件、无人机机臂和无人机"
    assert first.title_en == "UAV arm locking assembly, UAV arm and UAV"
    assert first.applicants == ("深圳市飞米机器人科技有限公司",)
    assert first.patent_number == "CN223200311U"
    assert first.publication_date == date(2025, 8, 8)
    assert first.filing_date == date(2024, 8, 8)
    assert first.patent_type == "实用新型"
    assert first.technology_effect_phrases == ("第一固定件牢固锁定",)


def test_import_patent_xlsx_splits_multi_applicant_and_effect_phrase_fields():
    result = import_patent_xlsx(_real_patent_xlsx_path())
    second = result.records[1]

    assert second.sequence_number == "10002"
    assert second.applicants == (
        "深圳市普渡科技有限公司",
        "成都市普渡机器人有限公司",
    )
    assert second.technology_effect_phrases[:3] == (
        "不需要进行翻转换面",
        "提升使用体验感",
        "提升作业效率",
    )


def test_import_patent_xlsx_reports_skipped_rows_for_missing_required_fields(
    tmp_path: Path,
):
    workbook_path = tmp_path / "patent_import.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(
        [
            "序号",
            "标题 (中文)",
            "标题 (英文)",
            "摘要 (中文)",
            "摘要 (英文)",
            "申请人",
            "公开（公告）号",
            "公开（公告）日",
            "申请日",
            "专利类型",
            "技术功效句",
            "技术功效短语",
            "预估到期日",
        ]
    )
    sheet.append(
        [
            1,
            "有效专利标题",
            "Valid patent title",
            "摘要",
            "abstract",
            "测试公司A; 测试公司B",
            "cn12345a",
            "2025-01-01",
            "2024-01-01",
            "发明申请",
            "技术功效",
            "功效A; 功效B",
            "2044-01-01",
        ]
    )
    sheet.append(
        [
            2,
            "",
            "Missing title",
            "摘要",
            "abstract",
            "测试公司",
            "CN00002A",
            "2025-01-02",
            "2024-01-02",
            "发明申请",
            "",
            "",
            "2044-01-02",
        ]
    )
    sheet.append(
        [
            3,
            "缺失申请人的专利",
            "Missing applicants",
            "摘要",
            "abstract",
            "",
            "CN00003A",
            "2025-01-03",
            "2024-01-03",
            "实用新型",
            "",
            "",
            "2034-01-03",
        ]
    )
    workbook.save(workbook_path)

    result = import_patent_xlsx(workbook_path)

    assert len(result.records) == 1
    assert result.records[0].title == "有效专利标题"
    assert result.records[0].applicants == ("测试公司A", "测试公司B")
    assert result.records[0].patent_number == "CN12345A"
    assert result.records[0].publication_date == date(2025, 1, 1)
    assert result.records[0].filing_date == date(2024, 1, 1)
    assert result.records[0].patent_type == "发明"
    assert result.report.rows_read == 3
    assert result.report.records_parsed == 1
    assert result.report.skipped_rows == 2
    assert result.report.skip_reasons["missing_title"] == 1
    assert result.report.skip_reasons["missing_applicants"] == 1
