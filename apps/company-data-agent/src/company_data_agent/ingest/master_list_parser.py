"""Parser for Shenzhen company master-list inputs in CSV and Excel formats."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field


HEADER_ALIASES: dict[str, str] = {
    "企业名称": "name",
    "公司名称": "name",
    "name": "name",
    "company_name": "name",
    "统一社会信用代码": "credit_code",
    "信用代码": "credit_code",
    "credit_code": "credit_code",
    "unified_social_credit_code": "credit_code",
    "注册地址": "registered_address",
    "地址": "registered_address",
    "registered_address": "registered_address",
    "行业分类": "industry",
    "行业": "industry",
    "industry": "industry",
}

REQUIRED_CANONICAL_HEADERS = {"name", "credit_code"}


class ParsedMasterListRow(BaseModel):
    """Normalized row produced by the master-list parser."""

    model_config = ConfigDict(extra="forbid")

    row_number: int
    source_path: str
    raw_columns: dict[str, str | None]
    name: str
    credit_code: str
    registered_address: str | None = None
    industry: str | None = None
    extra_columns: dict[str, str | None] = Field(default_factory=dict)


class MasterListParseError(BaseModel):
    """Structured parse error for a single source row."""

    model_config = ConfigDict(extra="forbid")

    row_number: int
    source_path: str
    message: str
    raw_columns: dict[str, str | None]


class MasterListParseResult(BaseModel):
    """Rows and row-level failures emitted by the parser."""

    model_config = ConfigDict(extra="forbid")

    rows: list[ParsedMasterListRow]
    errors: list[MasterListParseError]
    source_path: str


class MasterListParser:
    """Parse CSV or Excel company master-list files into normalized rows."""

    def parse(self, source_path: str | Path) -> MasterListParseResult:
        path = Path(source_path)
        suffix = path.suffix.lower()
        if suffix == ".csv":
            raw_rows = self._read_csv(path)
        elif suffix in {".xlsx", ".xlsm"}:
            raw_rows = self._read_xlsx(path)
        else:
            raise ValueError(f"unsupported master list format: {path.suffix}")

        return self._normalize_rows(path, raw_rows)

    def _read_csv(self, path: Path) -> list[tuple[int, dict[str, str | None]]]:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if reader.fieldnames is None:
                raise ValueError("csv input must include a header row")
            return [
                (index, self._normalize_raw_columns(row))
                for index, row in enumerate(reader, start=2)
            ]

    def _read_xlsx(self, path: Path) -> list[tuple[int, dict[str, str | None]]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_cells is None:
            raise ValueError("excel input must include a header row")
        headers = [self._stringify_header(value) for value in header_cells]
        if not any(headers):
            raise ValueError("excel header row must not be empty")

        rows: list[tuple[int, dict[str, str | None]]] = []
        for index, values in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            row: dict[str, str | None] = {}
            for header, value in zip(headers, values, strict=False):
                if header is None:
                    continue
                row[header] = self._stringify_cell(value)
            rows.append((index, self._normalize_raw_columns(row)))
        workbook.close()
        return rows

    def _normalize_rows(
        self,
        path: Path,
        raw_rows: list[tuple[int, dict[str, str | None]]],
    ) -> MasterListParseResult:
        rows: list[ParsedMasterListRow] = []
        errors: list[MasterListParseError] = []

        for row_number, raw_columns in raw_rows:
            if self._is_blank_row(raw_columns):
                continue

            canonical_columns = self._map_to_canonical_columns(raw_columns)
            missing = [
                column
                for column in sorted(REQUIRED_CANONICAL_HEADERS)
                if not canonical_columns.get(column)
            ]
            if missing:
                errors.append(
                    MasterListParseError(
                        row_number=row_number,
                        source_path=str(path),
                        message=f"missing required columns: {', '.join(missing)}",
                        raw_columns=raw_columns,
                    )
                )
                continue

            rows.append(
                ParsedMasterListRow(
                    row_number=row_number,
                    source_path=str(path),
                    raw_columns=raw_columns,
                    name=canonical_columns["name"],
                    credit_code=canonical_columns["credit_code"],
                    registered_address=canonical_columns.get("registered_address"),
                    industry=canonical_columns.get("industry"),
                    extra_columns=self._extract_extra_columns(raw_columns),
                )
            )

        return MasterListParseResult(rows=rows, errors=errors, source_path=str(path))

    def _map_to_canonical_columns(
        self,
        raw_columns: dict[str, str | None],
    ) -> dict[str, str | None]:
        canonical: dict[str, str | None] = {}
        for raw_header, value in raw_columns.items():
            alias_key = raw_header.strip().lower()
            canonical_name = HEADER_ALIASES.get(raw_header) or HEADER_ALIASES.get(alias_key)
            if canonical_name is None:
                continue
            if canonical_name not in canonical or not canonical[canonical_name]:
                canonical[canonical_name] = value
        return canonical

    def _extract_extra_columns(
        self,
        raw_columns: dict[str, str | None],
    ) -> dict[str, str | None]:
        extra: dict[str, str | None] = {}
        for raw_header, value in raw_columns.items():
            alias_key = raw_header.strip().lower()
            canonical_name = HEADER_ALIASES.get(raw_header) or HEADER_ALIASES.get(alias_key)
            if canonical_name is None:
                extra[raw_header] = value
        return extra

    def _normalize_raw_columns(
        self,
        row: dict[str, Any],
    ) -> dict[str, str | None]:
        normalized: dict[str, str | None] = {}
        for key, value in row.items():
            if key is None:
                continue
            header = str(key).strip()
            if not header:
                continue
            normalized[header] = self._stringify_cell(value)
        return normalized

    def _is_blank_row(self, row: dict[str, str | None]) -> bool:
        return not any(value for value in row.values())

    def _stringify_header(self, value: Any) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    def _stringify_cell(self, value: Any) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None
