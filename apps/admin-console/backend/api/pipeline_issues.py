"""P2.2 — pipeline_issue feed API.

Surfaces the `pipeline_issue` table for the dashboard. Supports filtering
by stage / severity / resolved / reported_by (the cleanup-round label)
and pagination. Also exposes a per-guard summary used by the dashboard
"数据质量动态" strip: last run time + 7-day counts per reported_by.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Literal

from fastapi import APIRouter, Depends, HTTPException, Query
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from backend.deps import get_pg_conn

router = APIRouter(prefix="/api")


class ChatFeedbackIssueContext(BaseModel):
    session_id: str | None = None
    query: str | None = None
    query_type: str | None = None
    answer_text: str | None = None
    answer_style: str | None = None
    feedback_type: str | None = None
    note: str | None = None
    citations: list[dict[str, Any]] = Field(default_factory=list)
    citation_map: dict[str, Any] = Field(default_factory=dict)
    structured_payload: dict[str, Any] = Field(default_factory=dict)


class PipelineIssueRow(BaseModel):
    issue_id: str
    professor_id: str | None
    link_id: str | None
    institution: str | None
    stage: str
    severity: str
    description: str
    evidence_snapshot: dict[str, Any] | None
    reported_by: str | None
    reported_at: datetime
    resolved: bool
    resolved_at: datetime | None
    resolution_notes: str | None
    resolution_round: str | None
    domain: str | None = None
    issue_type: str | None = None
    task_id: str | None = None
    source_rows: list[int] = Field(default_factory=list)
    recommended_action: str | None = None
    chat_feedback: ChatFeedbackIssueContext | None = None


class PipelineIssueListResponse(BaseModel):
    total: int
    page: int
    page_size: int
    items: list[PipelineIssueRow]


class PipelineIssueUpdateRequest(BaseModel):
    resolved: bool
    resolution_notes: str | None = None
    resolution_round: str | None = None


class SourceRowCell(BaseModel):
    column_index: int
    column_letter: str
    header: str | None
    value: str | None


class SourceRowPreview(BaseModel):
    row_number: int
    cells: list[SourceRowCell]


class PipelineIssueSourceRowsResponse(BaseModel):
    issue_id: str
    task_id: str | None
    domain: str | None
    upload_path: str | None
    sheet_name: str | None
    header_row_number: int | None
    rows: list[SourceRowPreview]
    warning: str | None = None


class GuardRunSummary(BaseModel):
    reported_by: str
    last_run_at: datetime
    rows_last_run: int
    rows_last_7_days: int
    severity_breakdown: dict[str, int]


def _pipeline_issue_from_row(row: Any) -> PipelineIssueRow:
    evidence = _json_object(row["evidence_snapshot"])
    return PipelineIssueRow(
        issue_id=row["issue_id"],
        professor_id=row["professor_id"],
        link_id=row["link_id"],
        institution=row["institution"],
        stage=row["stage"],
        severity=row["severity"],
        description=row["description"],
        evidence_snapshot=row["evidence_snapshot"],
        reported_by=row["reported_by"],
        reported_at=row["reported_at"],
        resolved=row["resolved"],
        resolved_at=row["resolved_at"],
        resolution_notes=row["resolution_notes"],
        resolution_round=row["resolution_round"],
        domain=_optional_text(evidence.get("domain")),
        issue_type=_optional_text(evidence.get("issue_type")),
        task_id=_optional_text(evidence.get("task_id")),
        source_rows=_source_row_numbers(evidence.get("source_rows")),
        recommended_action=_optional_text(evidence.get("recommended_action")),
        chat_feedback=_chat_feedback_context(evidence),
    )


@router.get("/pipeline-issues", response_model=PipelineIssueListResponse)
def list_pipeline_issues(
    stage: str | None = Query(default=None),
    severity: Literal["low", "medium", "high"] | None = Query(default=None),
    resolved: bool | None = Query(default=None),
    reported_by: str | None = Query(default=None),
    professor_id: str | None = Query(default=None),
    task_id: str | None = Query(default=None),
    domain: str | None = Query(default=None),
    issue_type: str | None = Query(default=None),
    q: str | None = Query(default=None, description="ILIKE description"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=200),
    conn: Any = Depends(get_pg_conn),
) -> PipelineIssueListResponse:
    conditions: list[str] = []
    params: dict[str, Any] = {
        "offset": (page - 1) * page_size,
        "page_size": page_size,
    }
    if stage is not None:
        conditions.append("stage = %(stage)s")
        params["stage"] = stage
    if severity is not None:
        conditions.append("severity = %(severity)s")
        params["severity"] = severity
    if resolved is not None:
        conditions.append("resolved = %(resolved)s")
        params["resolved"] = resolved
    if reported_by is not None:
        conditions.append("reported_by = %(reported_by)s")
        params["reported_by"] = reported_by
    if professor_id is not None:
        conditions.append("professor_id = %(professor_id)s")
        params["professor_id"] = professor_id
    if task_id is not None:
        conditions.append("evidence_snapshot->>'task_id' = %(task_id)s")
        params["task_id"] = task_id
    if domain is not None:
        conditions.append("evidence_snapshot->>'domain' = %(domain)s")
        params["domain"] = domain
    if issue_type is not None:
        conditions.append("evidence_snapshot->>'issue_type' = %(issue_type)s")
        params["issue_type"] = issue_type
    if q:
        conditions.append("description ILIKE %(q_like)s")
        params["q_like"] = f"%{q}%"

    where = "WHERE " + " AND ".join(conditions) if conditions else ""
    rows = conn.execute(
        f"""
        SELECT issue_id::text AS issue_id,
               professor_id,
               link_id::text AS link_id,
               institution,
               stage, severity,
               description,
               evidence_snapshot,
               reported_by,
               reported_at,
               resolved,
               resolved_at,
               resolution_notes,
               resolution_round,
               count(*) OVER ()::int AS total_count
          FROM pipeline_issue
          {where}
         ORDER BY reported_at DESC
         OFFSET %(offset)s LIMIT %(page_size)s
        """,
        params,
    ).fetchall()
    total = int(rows[0]["total_count"]) if rows else 0
    items = [_pipeline_issue_from_row(r) for r in rows]
    return PipelineIssueListResponse(
        total=total, page=page, page_size=page_size, items=items
    )


@router.get(
    "/pipeline-issues/{issue_id}/source-rows",
    response_model=PipelineIssueSourceRowsResponse,
)
def get_pipeline_issue_source_rows(
    issue_id: str,
    conn: Any = Depends(get_pg_conn),
) -> PipelineIssueSourceRowsResponse:
    issue = conn.execute(
        """
        SELECT issue_id::text AS issue_id,
               evidence_snapshot
          FROM pipeline_issue
         WHERE issue_id = %(issue_id)s
         LIMIT 1
        """,
        {"issue_id": issue_id},
    ).fetchone()
    if issue is None:
        raise HTTPException(status_code=404, detail="Issue not found")

    evidence = _json_object(_row_value(issue, "evidence_snapshot", 1))
    source_rows = _source_row_numbers(evidence.get("source_rows"))
    task_id = _optional_text(evidence.get("task_id"))
    domain = _optional_text(evidence.get("domain"))
    response_issue_id = str(_row_value(issue, "issue_id", 0))

    if not task_id or not source_rows:
        return PipelineIssueSourceRowsResponse(
            issue_id=response_issue_id,
            task_id=task_id,
            domain=domain,
            upload_path=None,
            sheet_name=None,
            header_row_number=None,
            rows=[],
            warning="missing_task_id_or_source_rows",
        )

    run = conn.execute(
        """
        SELECT run_scope
          FROM pipeline_run
         WHERE run_id = %(task_id)s
         LIMIT 1
        """,
        {"task_id": task_id},
    ).fetchone()
    if run is None:
        return PipelineIssueSourceRowsResponse(
            issue_id=response_issue_id,
            task_id=task_id,
            domain=domain,
            upload_path=None,
            sheet_name=None,
            header_row_number=None,
            rows=[],
            warning="pipeline_run_not_found",
        )

    run_scope = _json_object(_row_value(run, "run_scope", 0))
    upload_path_text = _optional_text(run_scope.get("upload_path"))
    if not upload_path_text:
        return PipelineIssueSourceRowsResponse(
            issue_id=response_issue_id,
            task_id=task_id,
            domain=domain,
            upload_path=None,
            sheet_name=None,
            header_row_number=None,
            rows=[],
            warning="upload_path_not_recorded",
        )

    preview = _read_excel_source_rows(Path(upload_path_text), source_rows)
    return PipelineIssueSourceRowsResponse(
        issue_id=response_issue_id,
        task_id=task_id,
        domain=domain,
        **preview,
    )


@router.patch("/pipeline-issues/{issue_id}", response_model=PipelineIssueRow)
def update_pipeline_issue(
    issue_id: str,
    body: PipelineIssueUpdateRequest,
    conn: Any = Depends(get_pg_conn),
) -> PipelineIssueRow:
    with conn.transaction():
        row = conn.execute(
            """
            UPDATE pipeline_issue
               SET resolved = %(resolved)s,
                   resolved_at = CASE
                       WHEN %(resolved)s THEN COALESCE(resolved_at, now())
                       ELSE NULL
                   END,
                   resolution_notes = %(resolution_notes)s,
                   resolution_round = %(resolution_round)s
             WHERE issue_id = %(issue_id)s
         RETURNING issue_id::text AS issue_id,
                   professor_id,
                   link_id::text AS link_id,
                   institution,
                   stage,
                   severity,
                   description,
                   evidence_snapshot,
                   reported_by,
                   reported_at,
                   resolved,
                   resolved_at,
                   resolution_notes,
                   resolution_round
            """,
            {
                "issue_id": issue_id,
                "resolved": body.resolved,
                "resolution_notes": body.resolution_notes,
                "resolution_round": body.resolution_round,
            },
        ).fetchone()
    if row is None:
        raise HTTPException(status_code=404, detail="Issue not found")
    return _pipeline_issue_from_row(row)


@router.get(
    "/pipeline-issues/guard-runs",
    response_model=list[GuardRunSummary],
)
def list_guard_runs(
    conn: Any = Depends(get_pg_conn),
) -> list[GuardRunSummary]:
    """Per-reported_by summary for the dashboard 质量动态 strip.

    Excludes NULL reported_by (legacy rows without a guard label).
    Sorted by most recent run first.
    """
    rows = conn.execute(
        """
        WITH guard_summary AS (
          SELECT reported_by,
                 max(reported_at) AS last_run_at,
                 count(*) FILTER (
                   WHERE reported_at >= now() - interval '7 days'
                 )::int AS rows_last_7_days,
                 jsonb_object_agg(
                   severity, sev_count
                 ) AS severity_breakdown,
                 max(run_rows) AS rows_last_run
            FROM (
              SELECT reported_by,
                     reported_at,
                     severity,
                     count(*) OVER (
                       PARTITION BY reported_by, severity
                     )::int AS sev_count,
                     count(*) OVER (
                       PARTITION BY reported_by,
                       date_trunc('minute', reported_at)
                     )::int AS run_rows
                FROM pipeline_issue
               WHERE reported_by IS NOT NULL
            ) t
           GROUP BY reported_by
        )
        SELECT reported_by, last_run_at, rows_last_run,
               rows_last_7_days, severity_breakdown
          FROM guard_summary
         ORDER BY last_run_at DESC
        """,
    ).fetchall()
    return [
        GuardRunSummary(
            reported_by=r["reported_by"],
            last_run_at=r["last_run_at"],
            rows_last_run=r["rows_last_run"] or 0,
            rows_last_7_days=r["rows_last_7_days"],
            severity_breakdown=r["severity_breakdown"] or {},
        )
        for r in rows
    ]


def _read_excel_source_rows(
    upload_path: Path,
    source_rows: list[int],
) -> dict[str, Any]:
    if not upload_path.exists():
        return {
            "upload_path": str(upload_path),
            "sheet_name": None,
            "header_row_number": None,
            "rows": [],
            "warning": "upload_path_not_found",
        }
    if upload_path.suffix.lower() != ".xlsx":
        return {
            "upload_path": str(upload_path),
            "sheet_name": None,
            "header_row_number": None,
            "rows": [],
            "warning": "unsupported_source_file_type",
        }

    workbook = load_workbook(upload_path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            return {
                "upload_path": str(upload_path),
                "sheet_name": None,
                "header_row_number": None,
                "rows": [],
                "warning": "workbook_has_no_active_sheet",
            }
        header_row_number = _detect_preview_header_row(sheet, source_rows)
        headers = _read_excel_row_values(sheet, header_row_number)
        previews = [
            SourceRowPreview(
                row_number=row_number,
                cells=_source_row_cells(
                    headers=headers,
                    values=_read_excel_row_values(sheet, row_number),
                ),
            )
            for row_number in source_rows[:50]
        ]
        return {
            "upload_path": str(upload_path),
            "sheet_name": sheet.title,
            "header_row_number": header_row_number,
            "rows": previews,
            "warning": None if len(source_rows) <= 50 else "source_rows_truncated",
        }
    finally:
        workbook.close()


def _detect_preview_header_row(sheet: Any, source_rows: list[int]) -> int | None:
    first_source_row = min(source_rows) if source_rows else 1
    max_scan_row = max(1, min(32, first_source_row - 1))
    best_row_number: int | None = None
    best_score = -1
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=max_scan_row, values_only=True),
        start=1,
    ):
        score = sum(1 for value in row if _cell_text(value))
        if score > best_score:
            best_score = score
            best_row_number = row_number
    return best_row_number


def _read_excel_row_values(sheet: Any, row_number: int | None) -> list[str | None]:
    if row_number is None:
        return []
    rows = list(
        sheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True)
    )
    if not rows:
        return []
    return [_cell_text(value) for value in rows[0]]


def _source_row_cells(
    *,
    headers: list[str | None],
    values: list[str | None],
) -> list[SourceRowCell]:
    cell_count = max(len(headers), len(values))
    cells: list[SourceRowCell] = []
    for index in range(cell_count):
        header = headers[index] if index < len(headers) else None
        value = values[index] if index < len(values) else None
        if header is None and value is None:
            continue
        cells.append(
            SourceRowCell(
                column_index=index + 1,
                column_letter=get_column_letter(index + 1),
                header=header,
                value=value,
            )
        )
    return cells


def _cell_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return text[:500]


def _source_row_numbers(value: Any) -> list[int]:
    if not isinstance(value, list):
        return []
    rows: list[int] = []
    for item in value:
        try:
            row_number = int(item)
        except (TypeError, ValueError):
            continue
        if row_number > 0 and row_number not in rows:
            rows.append(row_number)
    return rows


def _chat_feedback_context(evidence: dict[str, Any]) -> ChatFeedbackIssueContext | None:
    if _optional_text(evidence.get("issue_type")) != "chat_feedback":
        return None
    return ChatFeedbackIssueContext(
        session_id=_optional_text(evidence.get("session_id")),
        query=_optional_text(evidence.get("query")),
        query_type=_optional_text(evidence.get("query_type")),
        answer_text=_optional_text(evidence.get("answer_text")),
        answer_style=_optional_text(evidence.get("answer_style")),
        feedback_type=_optional_text(evidence.get("feedback_type")),
        note=_optional_text(evidence.get("note")),
        citations=_dict_list(evidence.get("citations")),
        citation_map=_json_object(evidence.get("citation_map")),
        structured_payload=_json_object(evidence.get("structured_payload")),
    )


def _dict_list(value: Any) -> list[dict[str, Any]]:
    if not isinstance(value, list):
        return []
    return [item for item in value if isinstance(item, dict)]


def _json_object(value: Any) -> dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _optional_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _row_value(row: Any, key: str, index: int) -> Any:
    return row[key] if isinstance(row, dict) else row[index]
