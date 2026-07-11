#!/usr/bin/env python3
"""Build the Canonical V2 S2 source inventory without mutating evidence."""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path
import sqlite3
from typing import Any


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def inspect_sqlite(path: Path, root: Path) -> dict[str, Any]:
    entry: dict[str, Any] = {
        "kind": "sqlite_snapshot",
        "path": str(path.relative_to(root)),
        "bytes": path.stat().st_size,
        "sha256": sha256_file(path),
        "access_mode": "sqlite_uri_mode_ro_immutable",
    }
    uri = f"file:{path.resolve()}?mode=ro&immutable=1"
    with sqlite3.connect(uri, uri=True) as connection:
        tables = [
            row[0]
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
            )
        ]
        entry["tables"] = tables
        entry["table_counts"] = {}
        if "released_objects" in tables:
            entry["table_counts"]["released_objects"] = connection.execute(
                "SELECT count(*) FROM released_objects"
            ).fetchone()[0]
            entry["object_type_counts"] = {
                str(object_type): count
                for object_type, count in connection.execute(
                    "SELECT object_type, count(*) FROM released_objects "
                    "GROUP BY object_type ORDER BY object_type"
                )
            }
    return entry


def hash_only_file(path: Path, root: Path, *, kind: str) -> dict[str, Any]:
    return {
        "kind": kind,
        "path": str(path.relative_to(root)),
        "bytes": path.stat().st_size,
        "sha256": sha256_file(path),
        "access_mode": "hash_only_never_opened",
    }


def aggregate_files(
    paths: list[Path],
    root: Path,
    *,
    kind: str,
) -> dict[str, Any]:
    manifest_rows = []
    total_bytes = 0
    for path in sorted(paths, key=lambda item: str(item.relative_to(root))):
        relative = str(path.relative_to(root))
        size = path.stat().st_size
        total_bytes += size
        manifest_rows.append(f"{relative}|{size}|{sha256_file(path)}")
    manifest = "\n".join(manifest_rows)
    if manifest_rows:
        manifest += "\n"
    return {
        "kind": kind,
        "root": str(root),
        "files": len(manifest_rows),
        "bytes": total_bytes,
        "manifest_sha256": hashlib.sha256(manifest.encode("utf-8")).hexdigest(),
        "access_mode": "read_hash_only",
    }


def inspect_jsonl(path: Path, root: Path, *, kind: str) -> dict[str, Any]:
    lines = 0
    valid_objects = 0
    invalid_lines = 0
    first_record_keys: list[str] = []
    with path.open("r", encoding="utf-8", errors="replace") as handle:
        for line in handle:
            lines += 1
            if not line.strip():
                continue
            try:
                value = json.loads(line)
            except ValueError:
                invalid_lines += 1
                continue
            if not isinstance(value, dict):
                invalid_lines += 1
                continue
            valid_objects += 1
            if not first_record_keys:
                first_record_keys = sorted(value)
    entry = hash_only_file(path, root, kind=kind)
    entry.update(
        {
            "access_mode": "read_parse_no_write",
            "lines": lines,
            "valid_object_lines": valid_objects,
            "invalid_lines": invalid_lines,
            "first_record_keys": first_record_keys,
        }
    )
    return entry


def inspect_xlsx(path: Path, root: Path, *, kind: str) -> dict[str, Any]:
    entry = hash_only_file(path, root, kind=kind)
    entry["access_mode"] = "openpyxl_read_only_data_only"
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            entry["sheets"] = [
                {
                    "name": sheet.title,
                    "rows": sheet.max_row,
                    "columns": sheet.max_column,
                }
                for sheet in workbook.worksheets
            ]
        finally:
            workbook.close()
    except Exception as exc:
        entry["inspection_error"] = f"{type(exc).__name__}: {exc}"
    return entry


def _annotate(
    entry: dict[str, Any],
    *,
    authority: str,
    domains: list[str],
    limitation: str,
) -> dict[str, Any]:
    entry.update(
        {
            "authority": authority,
            "domains": domains,
            "limitation": limitation,
        }
    )
    return entry


def _existing_files(root: Path, pattern: str) -> list[Path]:
    return sorted(path for path in root.glob(pattern) if path.is_file())


def build_inventory(
    *,
    workspace: Path,
    evidence_root: Path,
    recovery_root: Path,
    recovery_snapshot: dict[str, Any],
    git_commit: str,
    captured_at: str,
) -> dict[str, Any]:
    sources: list[dict[str, Any]] = []

    prds = (
        "docs/Data-Agent-Shared-Spec.md",
        "docs/Agentic-RAG-PRD.md",
        "docs/Company-Data-Agent-PRD.md",
        "docs/Professor-Data-Agent-Requirements-Audit-2026-05-09.md",
        "docs/Paper-Data-Agent-PRD.md",
        "docs/Patent-Data-Agent-PRD.md",
        "docs/Multi-turn-Context-Manager-Design.md",
    )
    for relative in prds:
        path = workspace / relative
        if path.exists():
            sources.append(
                _annotate(
                    hash_only_file(path, workspace, kind="authoritative_prd"),
                    authority="prd_authoritative",
                    domains=["shared"],
                    limitation="Requirements source; not row evidence.",
                )
            )

    workbook = workspace / "docs/测试集答案.xlsx"
    if workbook.exists():
        sources.append(
            _annotate(
                inspect_xlsx(workbook, workspace, kind="seed_scenario_workbook"),
                authority="user_seed_cases",
                domains=["professor", "company", "paper", "patent", "cross_domain"],
                limitation="Seed scenarios only; not the product boundary or sole gold corpus.",
            )
        )

    for path in _existing_files(workspace, "docs/source_backfills/*.jsonl"):
        sources.append(
            _annotate(
                inspect_jsonl(path, workspace, kind="committed_backfill_jsonl"),
                authority="historical_structured_evidence",
                domains=["mixed"],
                limitation="Historical backfill output requires identity/provenance review before canonical use.",
            )
        )
    for path in _existing_files(workspace, "docs/source_backfills/*.xlsx"):
        sources.append(
            _annotate(
                inspect_xlsx(path, workspace, kind="committed_backfill_workbook"),
                authority="historical_structured_evidence",
                domains=["mixed"],
                limitation="Small supplement, not complete domain coverage.",
            )
        )
    for pattern, kind in (
        ("apps/admin-console/tests/fixtures/*.jsonl", "committed_eval_fixture"),
        ("apps/admin-console/tests/fixtures/*.yaml", "committed_eval_fixture"),
        ("apps/admin-console/scripts/*.json", "committed_eval_artifact"),
    ):
        for path in _existing_files(workspace, pattern):
            sources.append(
                _annotate(
                    hash_only_file(path, workspace, kind=kind),
                    authority="legacy_evaluation_evidence",
                    domains=["query_answer"],
                    limitation="Retain original corpus/scorer/time labels; not automatically current baseline.",
                )
            )

    original_milvus = evidence_root / "apps/miroflow-agent/milvus.db"
    if original_milvus.exists():
        sources.append(
            _annotate(
                hash_only_file(
                    original_milvus,
                    evidence_root,
                    kind="milvus_lite_original",
                ),
                authority="forensic_source_hash_only",
                domains=["professor", "company", "paper", "patent"],
                limitation="No verified copy exists; client open and collection inspection are forbidden in S2.",
            )
        )

    individual_evidence = (
        ("docs/企业总表.xlsx", "company_source_workbook", ["company"]),
        ("docs/2025-12-05 专利.xlsx", "patent_source_workbook", ["patent"]),
        (
            "logs/data_agents/professor/enriched_v3_merged.jsonl",
            "professor_merged_jsonl",
            ["professor", "paper"],
        ),
        ("logs/legacy_v2/enriched_v2_2026-04-05.jsonl", "legacy_professor_jsonl", ["professor"]),
    )
    for relative, kind, domains in individual_evidence:
        path = evidence_root / relative
        if not path.exists():
            continue
        inspect = inspect_jsonl if path.suffix == ".jsonl" else inspect_xlsx
        sources.append(
            _annotate(
                inspect(path, evidence_root, kind=kind),
                authority="historical_local_evidence",
                domains=domains,
                limitation="Identity, freshness, and source-assertion mapping must be rebuilt.",
            )
        )

    released_objects = evidence_root / "logs/data_agents/released_objects.db"
    if released_objects.exists():
        sources.append(
            _annotate(
                inspect_sqlite(released_objects, evidence_root),
                authority="historical_published_snapshot",
                domains=["professor", "company", "paper", "patent"],
                limitation="Legacy projection, not canonical truth; inspect only through immutable SQLite mode.",
            )
        )

    legacy_milvus = evidence_root / "logs/legacy_v2/milvus_v2_2026-04-05.db"
    if legacy_milvus.exists():
        sources.append(
            _annotate(
                hash_only_file(legacy_milvus, evidence_root, kind="legacy_milvus_hash_only"),
                authority="historical_index_artifact",
                domains=["professor", "company", "paper", "patent"],
                limitation="Parent release parity is unproven; hash-only in S2.",
            )
        )

    family_specs = (
        ("logs/debug/professor_fetch_cache/*.json", "professor_fetch_cache_family", ["professor"]),
        ("logs/debug/paper_openalex_cache/*.json", "paper_openalex_cache_family", ["paper"]),
        ("logs/debug/paper_orcid_cache/*.json", "paper_orcid_cache_family", ["paper", "professor"]),
        ("logs/debug/*_release_e2e_*/*.jsonl", "legacy_release_jsonl_family", ["mixed"]),
        ("logs/data_agents/**/*.db", "legacy_sqlite_snapshot_family", ["mixed"]),
        ("logs/data_agents/**/*.jsonl", "legacy_data_agent_jsonl_family", ["mixed"]),
        ("data/admin_uploads/**/*.xlsx", "admin_upload_workbook_family", ["company", "patent"]),
        ("data/admin_uploads/**/*.jsonl", "admin_upload_jsonl_family", ["company", "patent"]),
        ("backups/*.csv.gz", "compressed_backup_family", ["mixed"]),
        ("apps/*/logs/raw_pdfs/**/*.pdf", "raw_pdf_family", ["paper"]),
        ("logs/**/*milvus*.db", "historical_milvus_file_family", ["mixed"]),
    )
    for pattern, kind, domains in family_specs:
        paths = _existing_files(evidence_root, pattern)
        if not paths:
            continue
        sources.append(
            _annotate(
                aggregate_files(paths, evidence_root, kind=kind),
                authority="historical_family_manifest",
                domains=domains,
                limitation="Family-level manifest; S4 landing must register individual artifacts and lineage.",
            )
        )

    recovery_files = (
        ("FORENSIC-CHECKPOINT.md", "forensic_checkpoint_document"),
        ("RECOVERY-EXPERIMENT-REPORT.md", "recovery_experiment_document"),
        ("LOGICAL-REBUILD-PLAN.md", "recovery_plan_document"),
        ("host-ext4-journal-20260711T024500Z.bin", "host_ext4_journal_copy"),
        ("lab-01/ext4-dir-inode-27017891.bin", "ext4_directory_inode_copy"),
        ("lab-01/miroflow-real-fpi-salvage.dump", "postgres_salvage_dump"),
        ("lab-01/recovered-paper-ids.txt", "recovered_paper_id_manifest"),
        ("lab-01/recovered-link-ids.txt", "recovered_link_id_manifest"),
    )
    for relative, kind in recovery_files:
        path = recovery_root / relative
        if path.exists():
            sources.append(
                _annotate(
                    hash_only_file(path, recovery_root, kind=kind),
                    authority="forensic_recovery_evidence",
                    domains=["paper", "professor_paper_relation"],
                    limitation="Evidence input only; partial recovery cannot supply missing values or other domains.",
                )
            )

    return {
        "schema_version": "canonical-v2-s2-source-inventory-v1",
        "builder_version": "canonical-v2-s2-source-inventory-builder-v1",
        "captured_at": captured_at,
        "git_commit": git_commit,
        "workspace": str(workspace),
        "evidence_root": str(evidence_root),
        "recovery_root": str(recovery_root),
        "forensic_source_manifest_sha256": (
            "bce14dce8fe2da4d053ac9cd930e1532f4abb436c5d03fff07aa69fd180e9e91"
        ),
        "recovery_database_snapshot": recovery_snapshot,
        "sources": sources,
        "known_hard_limits": [
            "recovery public domain tables are empty; only salvage paper and professor-paper links contain rows",
            "no verified Milvus copy exists, so S2 cannot inspect collections or current index contents",
            "lost TOAST values listed in salvage.field_errors cannot be reconstructed from the dump",
            "historical artifacts have heterogeneous identity, policy, and release semantics",
        ],
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workspace", required=True, type=Path)
    parser.add_argument("--evidence-root", required=True, type=Path)
    parser.add_argument("--recovery-root", required=True, type=Path)
    parser.add_argument("--recovery-snapshot", required=True, type=Path)
    parser.add_argument("--git-commit", required=True)
    parser.add_argument("--captured-at", required=True)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    snapshot = json.loads(args.recovery_snapshot.read_text(encoding="utf-8"))
    result = build_inventory(
        workspace=args.workspace.resolve(),
        evidence_root=args.evidence_root.resolve(),
        recovery_root=args.recovery_root.resolve(),
        recovery_snapshot=snapshot,
        git_commit=args.git_commit,
        captured_at=args.captured_at,
    )
    result["builder_sha256"] = sha256_file(Path(__file__).resolve())
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "output": str(args.output),
                "sources": len(result["sources"]),
            },
            sort_keys=True,
        )
    )


if __name__ == "__main__":
    main()
