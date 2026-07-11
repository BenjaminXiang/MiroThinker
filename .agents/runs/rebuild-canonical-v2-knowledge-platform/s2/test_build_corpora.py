from __future__ import annotations

import importlib.util
import json
import hashlib
from pathlib import Path
import tempfile
import unittest
from typing import Any

from openpyxl import Workbook


MODULE_PATH = Path(__file__).with_name("build_corpora.py")


def _load_builder() -> Any:
    if not MODULE_PATH.exists():
        raise AssertionError("build_corpora.py is not implemented")
    spec = importlib.util.spec_from_file_location("s2_corpus_builder", MODULE_PATH)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


class CorpusBuilderTests(unittest.TestCase):
    def test_workbook_parser_preserves_groups_and_skips_markers(self) -> None:
        builder = _load_builder()
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "cases.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            assert sheet is not None
            sheet.append(["问题1", "答案", "关键点"])
            sheet.append(["介绍甲教授", "甲教授答案", "身份;研究"])
            sheet.append(["他有哪些论文", "论文答案", "多轮"])
            sheet.append(["问题2", None, None])
            sheet.append(["专利 CN123A 的信息", "专利答案", "精确编号"])
            workbook.save(path)

            cases = builder.workbook_seed_cases(path)

            self.assertEqual([case["case_id"] for case in cases], ["wb-r002", "wb-r003", "wb-r005"])
            self.assertEqual([case["group"] for case in cases], ["问题1", "问题1", "问题2"])
            self.assertEqual([case["turn"] for case in cases], [1, 2, 1])
            self.assertEqual(cases[1]["review_status"], "user_confirmed_reference_gold")
            self.assertEqual(cases[0]["reference_answer"], "甲教授答案")
            self.assertEqual(cases[0]["reference_key_points"], "身份;研究")
            self.assertEqual(
                cases[0]["gold"],
                {
                    "authority": "user_confirmed_workbook",
                    "answer_role": "positive_reference",
                    "scope": "case_specific_reference",
                    "provenance": "docs/测试集答案.xlsx#row=2",
                },
            )
            self.assertEqual(cases[0]["domains"], ["professor"])
            self.assertEqual(cases[0]["family"], "exact_lookup")
            self.assertEqual(cases[0]["query_type"], "A")
            self.assertEqual(cases[1]["family"], "multi_turn")
            self.assertEqual(cases[1]["query_type"], "C")
            self.assertEqual(cases[2]["domains"], ["patent"])
            self.assertEqual(cases[2]["protected_slots"], [{"kind": "patent_number", "value": "CN123A"}])

    def test_workbook_refusal_and_known_bad_response_keep_distinct_gold_semantics(self) -> None:
        builder = _load_builder()
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "cases.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            assert sheet is not None
            sheet.append(["问题1", "答案", "关键点"])
            sheet.append(["帮我写一首诗", "不能处理", "不能回答"])
            sheet.append(["问题2", None, None])
            sheet.append(["介绍乙公司", "旧错误响应", "这里的答案是不准确"])
            workbook.save(path)

            refusal, badcase = builder.workbook_seed_cases(path)

            self.assertEqual(refusal["query_type"], "F")
            self.assertFalse(refusal["expected_behavior"]["web_augmentation_required"])
            self.assertEqual(badcase["gold"]["answer_role"], "known_bad_response")
            self.assertIn("reviewed_badcase", badcase["case_roles"])

    def test_build_corpora_writes_deterministic_hash_count_manifest(self) -> None:
        builder = _load_builder()
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            workbook_path = root / "cases.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            assert sheet is not None
            sheet.append(["问题1", "答案", "关键点"])
            sheet.append(["介绍甲教授", "甲教授答案", "身份"])
            workbook.save(workbook_path)
            builder.PRD_CASES = [
                {
                    "case_id": "prd-a",
                    "corpus": "regression-v1",
                    "source": "PRD",
                    "query": "结构化场景",
                    "query_type": "B",
                    "review_status": "pending_user_review",
                }
            ]
            builder.CHALLENGE_CASES = [
                {
                    "case_id": "challenge-a",
                    "corpus": "challenge-v1",
                    "source": "controlled_variation",
                    "query": "别名场景",
                    "query_type": "G",
                    "review_status": "pending_user_review",
                }
            ]
            output = root / "corpora"

            manifest = builder.build_corpora(workbook_path, output)

            regression = output / "regression-v1.jsonl"
            challenge = output / "challenge-v1.jsonl"
            self.assertEqual(manifest["corpora"]["regression-v1"]["cases"], 2)
            self.assertEqual(manifest["corpora"]["challenge-v1"]["cases"], 1)
            self.assertEqual(
                manifest["corpora"]["regression-v1"]["gold_authority_counts"],
                {"none": 1, "user_confirmed_workbook": 1},
            )
            self.assertEqual(
                manifest["corpora"]["regression-v1"]["query_type_counts"],
                {"A": 1, "B": 1},
            )
            self.assertEqual(manifest["approval_state"], "pending_user_acceptance")
            self.assertEqual(
                manifest["corpora"]["regression-v1"]["sha256"],
                hashlib.sha256(regression.read_bytes()).hexdigest(),
            )
            self.assertEqual(
                manifest["corpora"]["challenge-v1"]["sha256"],
                hashlib.sha256(challenge.read_bytes()).hexdigest(),
            )
            self.assertEqual(
                [json.loads(line)["case_id"] for line in regression.read_text().splitlines()],
                ["wb-r002", "prd-a"],
            )

    def test_declared_cases_cover_required_prd_and_challenge_families(self) -> None:
        builder = _load_builder()
        regression_families = {case["family"] for case in builder.PRD_CASES}
        challenge_families = {case["family"] for case in builder.CHALLENGE_CASES}

        self.assertTrue(
            {
                "exact_lookup",
                "semantic_search",
                "structured_filter",
                "relationship_traversal",
                "a_g_interaction",
                "multi_turn",
                "universal_web",
                "provenance_conflict",
                "partial_answer",
                "evidence_based_assessment",
            }.issubset(regression_families)
        )
        self.assertTrue(
            {
                "alias_spelling",
                "protected_constraints",
                "relation_direction",
                "referent_displayed_set",
                "topic_switch",
                "provider_failure",
                "insufficient_evidence",
            }.issubset(challenge_families)
        )
        self.assertEqual(
            {case["query_type"] for case in builder.PRD_CASES if case.get("query_type")},
            set("ABCDEFG"),
        )
        refusal = next(case for case in builder.PRD_CASES if case.get("query_type") == "F")
        self.assertEqual(refusal["expected_behavior"]["interaction"], "refusal")
        self.assertFalse(refusal["expected_behavior"]["web_augmentation_required"])
        self.assertTrue(
            all(
                case["review_status"] == "pending_user_review"
                for case in builder.PRD_CASES
                + [
                    case
                    for case in builder.CHALLENGE_CASES
                    if "reviewed_badcase" not in case.get("case_roles", [])
                ]
            )
        )
        self.assertTrue(
            all(case.get("query_type") in set("ABCDEFG") for case in builder.CHALLENGE_CASES)
        )
        reviewed_badcases = [
            case
            for case in builder.CHALLENGE_CASES
            if "reviewed_badcase" in case.get("case_roles", [])
        ]
        self.assertEqual(len(reviewed_badcases), 1)
        self.assertEqual(reviewed_badcases[0]["source"], "docs/测试集答案.xlsx#row=12")
        self.assertEqual(
            reviewed_badcases[0]["review_status"], "user_confirmed_reference_gold"
        )
        repo_root = MODULE_PATH.parents[4]
        self.assertTrue(
            all(
                (repo_root / case["source"].split("#", 1)[0]).is_file()
                for case in builder.PRD_CASES
            )
        )


if __name__ == "__main__":
    unittest.main()
