#!/usr/bin/env python3
"""Build versioned S2 regression and challenge corpora."""

from __future__ import annotations

import argparse
from collections import Counter
import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


GROUP_MARKER = re.compile(r"问题\d+")
PATENT_NUMBER = re.compile(r"\bCN[A-Z0-9]+\b", re.IGNORECASE)


def _declared_case(
    case_id: str,
    *,
    corpus: str,
    source: str,
    query: str,
    domains: list[str],
    family: str,
    query_type: str | None = None,
    protected_slots: list[dict[str, str]] | None = None,
    expected: dict[str, Any] | None = None,
    review_status: str = "pending_user_review",
    case_roles: list[str] | None = None,
    gold: dict[str, str] | None = None,
) -> dict[str, Any]:
    behavior = {
        "material_claims_require_evidence": True,
        "web_source_nature_disclosed": True,
    }
    if expected:
        behavior.update(expected)
    case = {
        "case_id": case_id,
        "corpus": corpus,
        "source": source,
        "query": query,
        "domains": domains,
        "family": family,
        "query_type": query_type,
        "protected_slots": protected_slots or [],
        "expected_behavior": behavior,
        "review_status": review_status,
        "case_roles": case_roles or [],
    }
    if gold is not None:
        case["gold"] = gold
    return case


PRD_CASES: list[dict[str, Any]] = [
    _declared_case(
        "prd-a-exact-patent",
        corpus="regression-v1",
        source="docs/Agentic-RAG-PRD.md#patent-query",
        query="专利 CN117873146A 的详细信息是什么",
        domains=["patent"],
        family="exact_lookup",
        query_type="A",
        protected_slots=[{"kind": "patent_number", "value": "CN117873146A"}],
        expected={"web_augmentation_required": True, "exact_identifier_preserved": True},
    ),
    _declared_case(
        "prd-b-professor-topic",
        corpus="regression-v1",
        source="docs/Agentic-RAG-PRD.md#professor-query",
        query="深圳有哪些做具身智能的教授",
        domains=["professor"],
        family="semantic_search",
        query_type="B",
        protected_slots=[{"kind": "geography", "value": "深圳"}],
        expected={"web_augmentation_required": True, "top_k_relevance_evaluated": 5},
    ),
    _declared_case(
        "prd-b-company-filter",
        corpus="regression-v1",
        source="docs/Company-Data-Agent-PRD.md#retrieval",
        query="深圳成立于2020年后、做机器人数据采集的企业有哪些",
        domains=["company"],
        family="structured_filter",
        query_type="B",
        protected_slots=[
            {"kind": "geography", "value": "深圳"},
            {"kind": "time_constraint", "value": "2020年后"},
        ],
        expected={"web_augmentation_required": True, "constraints_preserved": True},
    ),
    _declared_case(
        "prd-c-professor-to-paper",
        corpus="regression-v1",
        source="docs/Agentic-RAG-PRD.md#cross-domain",
        query="他的论文有哪些",
        domains=["professor", "paper"],
        family="relationship_traversal",
        query_type="C",
        expected={"web_augmentation_required": True, "requires_prior_anchor": "professor"},
    ),
    _declared_case(
        "prd-d-cross-domain-landscape",
        corpus="regression-v1",
        source="docs/Agentic-RAG-PRD.md#cross-domain-aggregation",
        query="列出深圳具身智能相关教授、企业、代表论文和专利",
        domains=["professor", "company", "paper", "patent"],
        family="a_g_interaction",
        query_type="D",
        protected_slots=[{"kind": "geography", "value": "深圳"}],
        expected={"web_augmentation_required": True, "progressive_not_exhaustive": True},
    ),
    _declared_case(
        "prd-e-knowledge-methods",
        corpus="regression-v1",
        source="docs/Agentic-RAG-PRD.md#knowledge-question",
        query="具身智能合成数据有几种实现方法，各自的适用条件是什么",
        domains=["cross_domain"],
        family="a_g_interaction",
        query_type="E",
        expected={"web_augmentation_required": True, "assessment_dimensions_explicit": True},
    ),
    _declared_case(
        "prd-f-refusal",
        corpus="regression-v1",
        source="docs/Agentic-RAG-PRD.md#query-classification",
        query="帮我写一首诗",
        domains=["none"],
        family="a_g_interaction",
        query_type="F",
        expected={"web_augmentation_required": False, "interaction": "refusal"},
    ),
    _declared_case(
        "prd-g-ambiguous-name",
        corpus="regression-v1",
        source="docs/Agentic-RAG-PRD.md#ambiguity",
        query="介绍王伟",
        domains=["professor"],
        family="a_g_interaction",
        query_type="G",
        protected_slots=[{"kind": "name", "value": "王伟"}],
        expected={"web_augmentation_required": True, "clarification_or_candidates": True},
    ),
    _declared_case(
        "prd-paper-topic-recent",
        corpus="regression-v1",
        source="docs/Agentic-RAG-PRD.md#paper-query",
        query="最近有什么关于人形机器人运动控制的新论文",
        domains=["paper"],
        family="semantic_search",
        query_type="B",
        protected_slots=[{"kind": "time_constraint", "value": "最近"}],
        expected={"web_augmentation_required": True, "top_k_relevance_evaluated": 5},
    ),
    _declared_case(
        "prd-patent-applicant-relation",
        corpus="regression-v1",
        source="docs/Patent-Data-Agent-PRD.md#relations",
        query="这家公司的专利有哪些",
        domains=["company", "patent"],
        family="relationship_traversal",
        query_type="C",
        expected={"web_augmentation_required": True, "relation_direction": "company_to_patent"},
    ),
    _declared_case(
        "prd-multi-turn-progressive",
        corpus="regression-v1",
        source="docs/Multi-turn-Context-Manager-Design.md",
        query="上述企业中，哪些有与机械臂相关的专利",
        domains=["company", "patent"],
        family="multi_turn",
        query_type="C",
        expected={
            "web_augmentation_required": True,
            "uses_displayed_set_only": True,
            "relation_direction": "company_to_patent",
        },
    ),
    _declared_case(
        "prd-universal-web-local-exact",
        corpus="regression-v1",
        source=(
            "openspec/changes/rebuild-canonical-v2-knowledge-platform/specs/"
            "evidence-first-query-orchestration/spec.md"
        ),
        query="介绍深圳智航无界科技",
        domains=["company"],
        family="universal_web",
        query_type="A",
        expected={"web_augmentation_required": True, "local_and_web_fused": True},
    ),
    _declared_case(
        "prd-provenance-conflict",
        corpus="regression-v1",
        source=(
            "openspec/changes/rebuild-canonical-v2-knowledge-platform/specs/"
            "canonical-v2-knowledge/spec.md"
        ),
        query="这家公司的成立时间为什么不同来源说法不一致",
        domains=["company"],
        family="provenance_conflict",
        query_type="A",
        expected={
            "web_augmentation_required": True,
            "conflicting_assertions_retained": True,
            "affected_claim_disclosed": True,
        },
    ),
    _declared_case(
        "prd-partial-answer-evidence-gap",
        corpus="regression-v1",
        source=(
            "openspec/changes/rebuild-canonical-v2-knowledge-platform/specs/"
            "grounded-progressive-answer/spec.md"
        ),
        query="这家公司有哪些机器人专利，目前证据不全的部分也请说明",
        domains=["company", "patent"],
        family="partial_answer",
        query_type="C",
        expected={
            "web_augmentation_required": True,
            "supported_subset_answered": True,
            "unsupported_scope_disclosed": True,
            "missing_evidence_not_filled_from_model_memory": True,
        },
    ),
    _declared_case(
        "prd-assessment-professor",
        corpus="regression-v1",
        source=(
            "openspec/changes/rebuild-canonical-v2-knowledge-platform/specs/"
            "grounded-progressive-answer/spec.md"
        ),
        query="这位教授在机器人领域是否属于高水平专家",
        domains=["professor", "paper", "patent", "company"],
        family="evidence_based_assessment",
        query_type="C",
        expected={
            "web_augmentation_required": True,
            "dimensions_and_uncertainty_required": True,
            "categorical_unsupported_verdict_forbidden": True,
        },
    ),
]


CHALLENGE_CASES: list[dict[str, Any]] = [
    _declared_case(
        "ch-reviewed-badcase-near-name-company",
        corpus="challenge-v1",
        source="docs/测试集答案.xlsx#row=12",
        query="我关注的是深圳智航无界科技，不是深圳智航无人机有限公司",
        domains=["company"],
        family="alias_spelling",
        query_type="A",
        protected_slots=[
            {"kind": "name", "value": "深圳智航无界科技"},
            {"kind": "negation", "value": "不是深圳智航无人机有限公司"},
        ],
        expected={
            "web_augmentation_required": True,
            "excluded_entity_not_returned": "深圳智航无人机有限公司",
            "zero_wrong_identity_substitution": True,
        },
        review_status="user_confirmed_reference_gold",
        case_roles=["reviewed_badcase"],
        gold={
            "authority": "user_confirmed_workbook",
            "scope": "case_specific_behavior",
            "provenance": "docs/测试集答案.xlsx#row=12",
        },
    ),
    _declared_case(
        "ch-alias-company",
        corpus="challenge-v1",
        source="controlled_variation:alias",
        query="无界智航和深圳智航无界科技是不是同一家公司",
        domains=["company"],
        family="alias_spelling",
        query_type="G",
        expected={"web_augmentation_required": True, "identity_evidence_required": True},
    ),
    _declared_case(
        "ch-spelling-professor",
        corpus="challenge-v1",
        source="controlled_variation:spelling",
        query="清华 SIGS 的 Ding Wenbo 有哪些论文",
        domains=["professor", "paper"],
        family="alias_spelling",
        query_type="A",
        protected_slots=[{"kind": "name", "value": "Ding Wenbo"}],
        expected={"web_augmentation_required": True, "alias_resolution_trace_required": True},
    ),
    _declared_case(
        "ch-time-geo-negation",
        corpus="challenge-v1",
        source="controlled_variation:constraints",
        query="只看2023年以后、总部在深圳且不是上市公司的机器人企业",
        domains=["company"],
        family="protected_constraints",
        query_type="B",
        protected_slots=[
            {"kind": "time_constraint", "value": "2023年以后"},
            {"kind": "geography", "value": "深圳"},
            {"kind": "negation", "value": "不是上市公司"},
        ],
        expected={"web_augmentation_required": True, "zero_protected_slot_loss": True},
    ),
    _declared_case(
        "ch-relation-reverse",
        corpus="challenge-v1",
        source="controlled_variation:relation-direction",
        query="专利 CN117873146A 的申请人对应哪家公司",
        domains=["patent", "company"],
        family="relation_direction",
        query_type="A",
        protected_slots=[{"kind": "patent_number", "value": "CN117873146A"}],
        expected={"web_augmentation_required": True, "relation_direction": "patent_to_company"},
    ),
    _declared_case(
        "ch-displayed-set",
        corpus="challenge-v1",
        source="controlled_variation:displayed-set",
        query="第二家公司的创始人是谁",
        domains=["company", "professor"],
        family="referent_displayed_set",
        query_type="C",
        expected={"web_augmentation_required": True, "uses_displayed_set_only": True},
    ),
    _declared_case(
        "ch-undisplayed-referent",
        corpus="challenge-v1",
        source="controlled_variation:displayed-set",
        query="第十家企业的专利有哪些",
        domains=["company", "patent"],
        family="referent_displayed_set",
        query_type="C",
        expected={"web_augmentation_required": True, "must_not_use_undisplayed_member": True},
    ),
    _declared_case(
        "ch-topic-switch",
        corpus="challenge-v1",
        source="controlled_variation:topic-switch",
        query="换个话题，介绍专利 CN117873146A",
        domains=["patent"],
        family="topic_switch",
        query_type="A",
        protected_slots=[{"kind": "patent_number", "value": "CN117873146A"}],
        expected={"web_augmentation_required": True, "prior_anchor_cleared": True},
    ),
    _declared_case(
        "ch-web-timeout",
        corpus="challenge-v1",
        source="controlled_variation:provider-failure",
        query="最近深圳有哪些新成立的具身智能企业",
        domains=["company"],
        family="provider_failure",
        query_type="B",
        expected={
            "web_augmentation_required": True,
            "injected_failure": "web_timeout",
            "supported_partial_answer_or_limitation": True,
        },
    ),
    _declared_case(
        "ch-llm-invalid-schema",
        corpus="challenge-v1",
        source="controlled_variation:provider-failure",
        query="比较这些企业的数据路线",
        domains=["company"],
        family="provider_failure",
        query_type="C",
        expected={
            "web_augmentation_required": True,
            "injected_failure": "llm_invalid_schema",
            "deterministic_degradation_required": True,
        },
    ),
    _declared_case(
        "ch-insufficient-capability",
        corpus="challenge-v1",
        source="controlled_variation:insufficient-evidence",
        query="这家公司能否量产用于酒店电梯的机械臂送餐机器人",
        domains=["company"],
        family="insufficient_evidence",
        query_type="C",
        expected={
            "web_augmentation_required": True,
            "unsupported_capability_inference_forbidden": True,
            "targeted_supplemental_attempt_bounded": True,
        },
    ),
    _declared_case(
        "ch-conflicting-role",
        corpus="challenge-v1",
        source="controlled_variation:conflict",
        query="这位教授是该公司的创始人还是顾问",
        domains=["professor", "company"],
        family="insufficient_evidence",
        query_type="C",
        expected={"web_augmentation_required": True, "conflict_disclosed": True},
    ),
]


def _protected_slots(query: str) -> list[dict[str, str]]:
    return [
        {"kind": "patent_number", "value": match.group(0)}
        for match in PATENT_NUMBER.finditer(query)
    ]


def _infer_domains(query: str) -> list[str]:
    domains: list[str] = []
    if re.search(r"教授|学者|清华|毕业于|企业家|创始人|丁文伯|王学谦", query):
        domains.append("professor")
    if re.search(r"企业|公司|厂商|供应商|PCB|产品|市场|产量", query):
        domains.append("company")
    if re.search(r"论文|paper|pFedGPA", query, re.IGNORECASE):
        domains.append("paper")
    if re.search(r"专利|\bCN[A-Z0-9]+\b", query, re.IGNORECASE):
        domains.append("patent")
    return domains or ["general_web"]


def _infer_workbook_route(query: str, *, turn: int) -> tuple[str, str]:
    if turn > 1:
        return "C", "multi_turn"
    if re.search(r"黄赌毒|违法|不能去|写一首诗|天气|翻译", query):
        return "F", "general_information"
    if PATENT_NUMBER.search(query) or re.search(r"pFedGPA", query, re.IGNORECASE):
        return "A", "exact_lookup"
    if re.search(r"评价|竞争力|是否.*大牛|市场对", query):
        return "D", "evidence_based_assessment"
    if re.search(r"几种|路线|方式|方法|原理|数据需求|有什么不同|具体展开", query):
        return "E", "knowledge_synthesis"
    if re.search(r"专利|论文|创立|创始人|申请人|参与.*企业", query):
        return "A", "relationship_traversal"
    if re.search(r"哪些|有哪些|推荐|找", query):
        return "B", "semantic_search"
    return "A", "exact_lookup"


def _reference_answer_role(key_points: str | None) -> str:
    if key_points and re.search(r"答案.*不准确|不应该出现", key_points):
        return "known_bad_response"
    return "positive_reference"


def workbook_seed_cases(
    path: Path,
    *,
    source_label: str = "docs/测试集答案.xlsx",
) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            raise ValueError(f"workbook has no active sheet: {path}")
        group = "问题1"
        turn = 0
        cases: list[dict[str, Any]] = []
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            query, answer, key_points = (list(row) + [None, None, None])[:3]
            query_text = str(query).strip() if query is not None else ""
            if not query_text:
                continue
            if GROUP_MARKER.fullmatch(query_text):
                group = query_text
                turn = 0
                continue
            turn += 1
            query_type, family = _infer_workbook_route(query_text, turn=turn)
            key_points_text = (
                str(key_points).strip() if key_points is not None else None
            )
            answer_role = _reference_answer_role(key_points_text)
            case_roles = ["workbook_seed"]
            if answer_role == "known_bad_response":
                case_roles.append("reviewed_badcase")
            cases.append(
                {
                    "case_id": f"wb-r{row_number:03d}",
                    "corpus": "regression-v1",
                    "source": source_label,
                    "source_row": row_number,
                    "group": group,
                    "turn": turn,
                    "query": query_text,
                    "domains": _infer_domains(query_text),
                    "family": family,
                    "query_type": query_type,
                    "reference_answer": str(answer).strip()
                    if answer is not None
                    else None,
                    "reference_key_points": key_points_text,
                    "gold": {
                        "authority": "user_confirmed_workbook",
                        "answer_role": answer_role,
                        "scope": "case_specific_reference",
                        "provenance": f"{source_label}#row={row_number}",
                    },
                    "case_roles": case_roles,
                    "protected_slots": _protected_slots(query_text),
                    "expected_behavior": {
                        "interaction": "information_retrieval",
                        "material_claims_require_evidence": True,
                        "web_augmentation_required": query_type != "F",
                        "web_source_nature_disclosed": True,
                    },
                    "review_status": "user_confirmed_reference_gold",
                }
            )
        return cases
    finally:
        workbook.close()


def _write_jsonl(path: Path, cases: list[dict[str, Any]]) -> str:
    payload = "".join(
        json.dumps(case, ensure_ascii=False, sort_keys=True) + "\n" for case in cases
    )
    path.write_text(payload, encoding="utf-8")
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _corpus_summary(path: Path, cases: list[dict[str, Any]], sha256: str) -> dict[str, Any]:
    return {
        "path": path.name,
        "cases": len(cases),
        "sha256": sha256,
        "review_status_counts": dict(
            sorted(Counter(case.get("review_status", "missing") for case in cases).items())
        ),
        "family_counts": dict(
            sorted(Counter(case.get("family", "unspecified") for case in cases).items())
        ),
        "domain_counts": dict(
            sorted(
                Counter(
                    domain
                    for case in cases
                    for domain in case.get("domains", ["unspecified"])
                ).items()
            )
        ),
        "gold_authority_counts": dict(
            sorted(
                Counter(
                    case.get("gold", {}).get("authority", "none") for case in cases
                ).items()
            )
        ),
        "role_counts": dict(
            sorted(
                Counter(role for case in cases for role in case.get("case_roles", []))
                .items()
            )
        ),
        "query_type_counts": dict(
            sorted(Counter(case.get("query_type") or "missing" for case in cases).items())
        ),
    }


def build_corpora(workbook_path: Path, output_dir: Path) -> dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    regression_cases = workbook_seed_cases(workbook_path) + PRD_CASES
    challenge_cases = CHALLENGE_CASES
    regression_path = output_dir / "regression-v1.jsonl"
    challenge_path = output_dir / "challenge-v1.jsonl"
    regression_hash = _write_jsonl(regression_path, regression_cases)
    challenge_hash = _write_jsonl(challenge_path, challenge_cases)
    return {
        "schema_version": "canonical-v2-s2-corpus-manifest-v1",
        "approval_state": "pending_user_acceptance",
        "workbook_sha256": hashlib.sha256(workbook_path.read_bytes()).hexdigest(),
        "corpora": {
            "regression-v1": _corpus_summary(
                regression_path, regression_cases, regression_hash
            ),
            "challenge-v1": _corpus_summary(
                challenge_path, challenge_cases, challenge_hash
            ),
        },
        "gold_policy": (
            "The user-confirmed workbook answers and key points are case-specific reference "
            "ground truth with row provenance. They are not a general answer template or the sole "
            "acceptance source. PRD and controlled variations define observable behavior only and "
            "remain pending user review."
        ),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--manifest", required=True, type=Path)
    args = parser.parse_args()
    manifest = build_corpora(args.workbook, args.output_dir)
    args.manifest.parent.mkdir(parents=True, exist_ok=True)
    args.manifest.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    print(
        json.dumps(
            {
                "manifest": str(args.manifest),
                "regression_cases": manifest["corpora"]["regression-v1"]["cases"],
                "challenge_cases": manifest["corpora"]["challenge-v1"]["cases"],
            },
            sort_keys=True,
        )
    )


if __name__ == "__main__":
    main()
