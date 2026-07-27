"""Focused contract tests for the lean customer-workbook replay."""

from __future__ import annotations

from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
import importlib.util
import json
from pathlib import Path
import sys
from threading import Lock, Thread
from types import ModuleType

from openpyxl import Workbook


TARGET_PATH = Path(__file__).with_name("customer_workbook_replay.py")
RELEASE_ID = "candidate:s12c:test"


def _module() -> ModuleType:
    spec = importlib.util.spec_from_file_location("s12c_customer_workbook_replay", TARGET_PATH)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _write_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("问题1", "答案", "关键点"))
    sheet.append((" 第一问 ", "参考答案一", "纠正点一"))
    sheet.append(("追问", "参考答案二", None))
    sheet.append(("问题2", "答案", "关键点"))
    sheet.append(("第二组", "参考答案三", "关键点三"))
    workbook.save(path)


def test_load_workbook_preserves_groups_turns_and_ground_truth(tmp_path: Path) -> None:
    module = _module()
    workbook_path = tmp_path / "benchmark.xlsx"
    _write_workbook(workbook_path)

    benchmark = module.load_workbook_benchmark(
        workbook_path,
        expected_conversations=2,
        expected_turns=3,
    )

    assert benchmark.workbook_sha256
    assert [item.conversation_id for item in benchmark.conversations] == [
        "question-01",
        "question-02",
    ]
    assert [turn.query for turn in benchmark.conversations[0].turns] == [
        "第一问",
        "追问",
    ]
    assert benchmark.conversations[0].turns[0].workbook_row == 2
    assert benchmark.conversations[0].turns[0].reference_answer == "参考答案一"
    assert benchmark.conversations[0].turns[0].key_points == "纠正点一"
    assert benchmark.conversations[0].turns[1].key_points is None


def test_real_http_replay_keeps_group_sessions_and_renders_failures(
    tmp_path: Path,
) -> None:
    module = _module()
    workbook_path = tmp_path / "benchmark.xlsx"
    _write_workbook(workbook_path)
    benchmark = module.load_workbook_benchmark(
        workbook_path,
        expected_conversations=2,
        expected_turns=3,
    )
    requests: list[tuple[str, str | None]] = []
    counter = 0
    lock = Lock()

    class Handler(BaseHTTPRequestHandler):
        def do_POST(self) -> None:  # noqa: N802 - stdlib handler contract.
            nonlocal counter
            if self.path == "/api/chat/session/reset":
                with lock:
                    counter += 1
                    session_id = f"session-{counter}"
                payload = {"session_id": session_id}
                self._send(
                    200,
                    payload,
                    cookie=f"miroflow_chat_session={session_id}; Path=/; HttpOnly",
                )
                return
            assert self.path == "/api/chat"
            body = json.loads(self.rfile.read(int(self.headers["content-length"])))
            query = body["query"]
            cookie = self.headers.get("cookie")
            with lock:
                requests.append((query, cookie))
            if query == "追问":
                self._send(409, {"detail": "visible_test_failure"})
                return
            self._send(
                200,
                {
                    "query": query,
                    "query_type": "canonical_v2:A:grounded_answer",
                    "answer_text": f"实际回答：{query}",
                    "citations": [
                        {
                            "type": "company",
                            "id": "company:test",
                            "label": "测试企业",
                            "url": "/browse#company/company:test",
                        }
                    ],
                    "evidence": [
                        {
                            "evidence_id": f"evidence:{query}",
                            "object_id": "company:test",
                            "domain": "company",
                            "lane": "exact",
                            "source_nature": "local",
                            "source_locator": "fixture:test",
                            "snippet": "测试证据",
                            "score": 1.0,
                        }
                    ],
                    "clarification": None,
                    "structured_payload": {
                        "canonical_v2": {
                            "release_id": RELEASE_ID,
                            "plan_id": f"plan:{query}",
                            "limitations": [
                                {"code": "partial_coverage", "message": "仅覆盖测试数据"}
                            ],
                            "evidence_ids": [f"evidence:{query}"],
                        }
                    },
                    "answer_style": "template",
                    "citation_map": {"1": "company:test"},
                    "suggested_followups": [],
                },
            )

        def _send(
            self,
            status: int,
            payload: dict[str, object],
            *,
            cookie: str | None = None,
        ) -> None:
            encoded = json.dumps(payload, ensure_ascii=False).encode()
            self.send_response(status)
            self.send_header("content-type", "application/json")
            self.send_header("content-length", str(len(encoded)))
            if cookie is not None:
                self.send_header("set-cookie", cookie)
            self.end_headers()
            self.wfile.write(encoded)

        def log_message(self, format: str, *args: object) -> None:
            del format, args

    server = ThreadingHTTPServer(("127.0.0.1", 0), Handler)
    thread = Thread(target=server.serve_forever, daemon=True)
    thread.start()
    try:
        replay = module.replay_benchmark(
            benchmark,
            base_url=f"http://127.0.0.1:{server.server_port}",
            expected_release_id=RELEASE_ID,
            workers=2,
            timeout_seconds=5.0,
        )
    finally:
        server.shutdown()
        thread.join(timeout=5)
        server.server_close()

    assert replay.summary == {
        "conversation_count": 2,
        "turn_count": 3,
        "ok_count": 2,
        "failure_count": 1,
    }
    assert replay.conversations[0].turns[1].status == "http_error"
    assert replay.conversations[0].turns[1].error == "visible_test_failure"
    cookies_by_query = {query: cookie for query, cookie in requests}
    assert cookies_by_query["第一问"] == cookies_by_query["追问"]
    assert cookies_by_query["第一问"] != cookies_by_query["第二组"]

    json_path = tmp_path / "replay.json"
    markdown_path = tmp_path / "replay.md"
    module.write_replay_outputs(
        replay,
        json_path=json_path,
        markdown_path=markdown_path,
    )
    stored = json.loads(json_path.read_text())
    assert stored["content_sha256"] == replay.content_sha256
    rendered = markdown_path.read_text()
    assert "Ground Truth" in rendered
    assert "实际回答" in rendered
    assert "visible_test_failure" in rendered
    assert "测试企业" in rendered
    assert "仅覆盖测试数据" in rendered
