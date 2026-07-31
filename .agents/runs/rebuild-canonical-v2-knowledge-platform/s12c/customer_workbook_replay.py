"""Replay the customer workbook through one explicit Canonical V2 chat origin."""

from __future__ import annotations

import argparse
from collections.abc import Sequence
from concurrent.futures import ThreadPoolExecutor
from dataclasses import asdict, dataclass
from datetime import UTC, datetime
import hashlib
from http.cookiejar import CookieJar
import json
import os
from pathlib import Path
import re
from time import monotonic
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.parse import urlsplit, urlunsplit
from urllib.request import (
    HTTPRedirectHandler,
    HTTPCookieProcessor,
    ProxyHandler,
    Request,
    build_opener,
)

from openpyxl import load_workbook


_SCHEMA_VERSION = "canonical-v2-customer-workbook-replay-v1"
_QUESTION_MARKER = re.compile(r"^问题\s*(\d+)$")
_MAX_WORKBOOK_BYTES = 10_000_000
_MAX_WORKBOOK_ROWS = 500
_MAX_QUERY_CHARS = 500
_MAX_RESPONSE_BYTES = 8_000_000
_MAX_WORKERS = 96


class WorkbookReplayError(RuntimeError):
    """The workbook or explicit chat transport violated the replay contract."""


@dataclass(frozen=True, slots=True)
class WorkbookTurn:
    turn_number: int
    workbook_row: int
    query: str
    reference_answer: str
    key_points: str | None


@dataclass(frozen=True, slots=True)
class WorkbookConversation:
    conversation_id: str
    question_number: int
    workbook_label: str
    turns: tuple[WorkbookTurn, ...]


@dataclass(frozen=True, slots=True)
class WorkbookBenchmark:
    workbook_path: str
    workbook_sha256: str
    sheet_name: str
    conversations: tuple[WorkbookConversation, ...]


@dataclass(frozen=True, slots=True)
class ReplayTurn:
    turn_number: int
    workbook_row: int
    query: str
    reference_answer: str
    key_points: str | None
    status: str
    http_status: int | None
    elapsed_ms: int
    error: str | None
    response: dict[str, Any] | None
    response_sha256: str | None


@dataclass(frozen=True, slots=True)
class ReplayConversation:
    conversation_id: str
    question_number: int
    workbook_label: str
    session_id: str | None
    turns: tuple[ReplayTurn, ...]


@dataclass(frozen=True, slots=True)
class BenchmarkReplay:
    schema_version: str
    generated_at: str
    base_url: str
    expected_release_id: str
    workbook_path: str
    workbook_sha256: str
    sheet_name: str
    conversations: tuple[ReplayConversation, ...]
    summary: dict[str, int]
    content_sha256: str

    def to_document(self) -> dict[str, Any]:
        return asdict(self)


class _NoRedirect(HTTPRedirectHandler):
    def redirect_request(
        self,
        req: Request,
        fp: Any,
        code: int,
        msg: str,
        headers: Any,
        newurl: str,
    ) -> None:
        del req, fp, code, msg, headers, newurl
        return None


def _sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _canonical_json_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
        allow_nan=False,
    ).encode("utf-8")


def _canonical_sha256(value: Any) -> str:
    return _sha256_bytes(_canonical_json_bytes(value))


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while chunk := stream.read(1024 * 1024):
            digest.update(chunk)
    return digest.hexdigest()


def _required_cell_text(value: object, *, label: str, row_number: int) -> str:
    if not isinstance(value, str) or not value.strip():
        raise WorkbookReplayError(
            f"workbook row {row_number} requires non-empty {label} text"
        )
    return value.strip()


def _optional_cell_text(value: object, *, label: str, row_number: int) -> str | None:
    if value is None:
        return None
    if not isinstance(value, str):
        raise WorkbookReplayError(
            f"workbook row {row_number} {label} must be text when present"
        )
    return value.strip() or None


def load_workbook_benchmark(
    path: Path,
    *,
    expected_conversations: int = 17,
    expected_turns: int = 25,
) -> WorkbookBenchmark:
    """Load the exact ordered benchmark without interpreting answers as runtime data."""

    workbook_path = path.resolve()
    if not workbook_path.is_file():
        raise WorkbookReplayError("workbook path must name one existing file")
    byte_size = workbook_path.stat().st_size
    if not 0 < byte_size <= _MAX_WORKBOOK_BYTES:
        raise WorkbookReplayError("workbook exceeds the bounded replay size")
    if expected_conversations <= 0 or expected_turns <= 0:
        raise WorkbookReplayError("expected workbook counts must be positive")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if len(workbook.sheetnames) != 1:
            raise WorkbookReplayError("customer workbook must contain exactly one sheet")
        sheet = workbook[workbook.sheetnames[0]]
        if sheet.max_row > _MAX_WORKBOOK_ROWS or sheet.max_column != 3:
            raise WorkbookReplayError("customer workbook must remain a bounded three-column sheet")

        conversations: list[WorkbookConversation] = []
        current_number: int | None = None
        current_label: str | None = None
        current_turns: list[WorkbookTurn] = []

        def finish_conversation() -> None:
            nonlocal current_number, current_label, current_turns
            if current_number is None:
                return
            if not current_turns:
                raise WorkbookReplayError(
                    f"question group {current_number} contains no query turns"
                )
            conversations.append(
                WorkbookConversation(
                    conversation_id=f"question-{current_number:02d}",
                    question_number=current_number,
                    workbook_label=current_label or f"问题{current_number}",
                    turns=tuple(current_turns),
                )
            )
            current_number = None
            current_label = None
            current_turns = []

        for row_number, values in enumerate(
            sheet.iter_rows(min_col=1, max_col=3, values_only=True),
            start=1,
        ):
            first, second, third = values
            first_text = first.strip() if isinstance(first, str) else None
            marker = None if first_text is None else _QUESTION_MARKER.fullmatch(first_text)
            if marker is not None:
                if second != "答案" or third != "关键点":
                    raise WorkbookReplayError(
                        f"workbook row {row_number} has an invalid question-group header"
                    )
                finish_conversation()
                question_number = int(marker.group(1))
                expected_number = len(conversations) + 1
                if question_number != expected_number:
                    raise WorkbookReplayError(
                        "question groups must be sequential from 1 without omissions"
                    )
                current_number = question_number
                current_label = first_text
                continue
            if first is None and second is None and third is None:
                continue
            if current_number is None:
                raise WorkbookReplayError(
                    f"workbook row {row_number} appears before a question-group header"
                )
            query = _required_cell_text(first, label="query", row_number=row_number)
            if len(query) > _MAX_QUERY_CHARS:
                raise WorkbookReplayError(
                    f"workbook row {row_number} exceeds the chat query limit"
                )
            current_turns.append(
                WorkbookTurn(
                    turn_number=len(current_turns) + 1,
                    workbook_row=row_number,
                    query=query,
                    reference_answer=_required_cell_text(
                        second,
                        label="reference answer",
                        row_number=row_number,
                    ),
                    key_points=_optional_cell_text(
                        third,
                        label="key points",
                        row_number=row_number,
                    ),
                )
            )
        finish_conversation()
    finally:
        workbook.close()

    turn_count = sum(len(item.turns) for item in conversations)
    if len(conversations) != expected_conversations or turn_count != expected_turns:
        raise WorkbookReplayError(
            "workbook count mismatch: "
            f"expected {expected_conversations} conversations/{expected_turns} turns, "
            f"observed {len(conversations)}/{turn_count}"
        )
    return WorkbookBenchmark(
        workbook_path=str(workbook_path),
        workbook_sha256=_file_sha256(workbook_path),
        sheet_name=workbook.sheetnames[0],
        conversations=tuple(conversations),
    )


def _explicit_origin(base_url: str) -> str:
    try:
        parsed = urlsplit(base_url)
    except ValueError as exc:
        raise WorkbookReplayError("base URL is invalid") from exc
    if (
        parsed.scheme not in {"http", "https"}
        or not parsed.netloc
        or parsed.username is not None
        or parsed.password is not None
        or parsed.query
        or parsed.fragment
        or parsed.path not in {"", "/"}
    ):
        raise WorkbookReplayError("base URL must be one credential-free HTTP origin")
    return urlunsplit((parsed.scheme, parsed.netloc, "", "", ""))


def _read_response(response: Any) -> tuple[bytes, dict[str, Any] | None]:
    raw = response.read(_MAX_RESPONSE_BYTES + 1)
    if len(raw) > _MAX_RESPONSE_BYTES:
        raise WorkbookReplayError("chat response exceeded the replay size bound")
    try:
        payload = json.loads(raw)
    except (UnicodeDecodeError, json.JSONDecodeError):
        return raw, None
    return raw, payload if isinstance(payload, dict) else None


def _payload_error(payload: dict[str, Any] | None, *, fallback: str) -> str:
    if payload is not None:
        detail = payload.get("detail")
        if isinstance(detail, str) and detail.strip():
            return detail.strip()
    return fallback


class _ConversationClient:
    def __init__(self, *, origin: str, timeout_seconds: float) -> None:
        self._origin = origin
        self._timeout_seconds = timeout_seconds
        self._opener = build_opener(
            ProxyHandler({}),
            _NoRedirect(),
            HTTPCookieProcessor(CookieJar()),
        )

    def reset(self) -> tuple[str | None, str | None]:
        endpoint = self._origin + "/api/chat/session/reset"
        request = Request(endpoint, data=b"", method="POST")
        try:
            with self._opener.open(request, timeout=self._timeout_seconds) as response:
                raw, payload = _read_response(response)
                if response.status != 200 or response.geturl() != endpoint:
                    return None, "session reset left the explicit HTTP contract"
        except HTTPError as exc:
            _, payload = _read_response(exc)
            return None, _payload_error(payload, fallback=f"session reset HTTP {exc.code}")
        except (URLError, OSError, WorkbookReplayError) as exc:
            return None, f"session reset failed: {type(exc).__name__}"
        del raw
        session_id = None if payload is None else payload.get("session_id")
        if not isinstance(session_id, str) or not session_id.strip():
            return None, "session reset response lacks a session_id"
        return session_id.strip(), None

    def chat(self, turn: WorkbookTurn, *, expected_release_id: str) -> ReplayTurn:
        endpoint = self._origin + "/api/chat"
        body = _canonical_json_bytes({"query": turn.query, "entity_id_hint": None})
        request = Request(
            endpoint,
            data=body,
            headers={"content-type": "application/json"},
            method="POST",
        )
        started = monotonic()
        try:
            with self._opener.open(request, timeout=self._timeout_seconds) as response:
                raw, payload = _read_response(response)
                status_code = response.status
                response_url = response.geturl()
        except HTTPError as exc:
            raw, payload = _read_response(exc)
            return self._turn_result(
                turn,
                status="http_error",
                http_status=exc.code,
                started=started,
                error=_payload_error(payload, fallback=f"HTTP {exc.code}"),
                payload=payload,
                raw=raw,
            )
        except (URLError, OSError, WorkbookReplayError) as exc:
            return self._turn_result(
                turn,
                status="transport_error",
                http_status=None,
                started=started,
                error=f"chat request failed: {type(exc).__name__}",
                payload=None,
                raw=None,
            )

        contract_errors: list[str] = []
        if status_code != 200:
            contract_errors.append(f"unexpected HTTP {status_code}")
        if response_url != endpoint:
            contract_errors.append("response left the explicit origin")
        if payload is None:
            contract_errors.append("response is not a JSON object")
        else:
            if payload.get("query") != turn.query:
                contract_errors.append("response query does not match the workbook query")
            answer_text = payload.get("answer_text")
            if not isinstance(answer_text, str) or not answer_text.strip():
                contract_errors.append("response lacks a readable answer")
            # The sanitized public envelope intentionally exposes no release ID
            # or internal trace (12.5b public-evidence contract). The Candidate
            # release binding is established by the operator-controlled launch;
            # per turn we only require the public Canonical V2 answer type.
            query_type = payload.get("query_type")
            if not isinstance(query_type, str) or not query_type.startswith(
                "canonical_v2:"
            ):
                contract_errors.append("response lacks a Canonical V2 answer type")
        return self._turn_result(
            turn,
            status="contract_error" if contract_errors else "ok",
            http_status=status_code,
            started=started,
            error="; ".join(contract_errors) or None,
            payload=payload,
            raw=raw,
        )

    @staticmethod
    def _turn_result(
        turn: WorkbookTurn,
        *,
        status: str,
        http_status: int | None,
        started: float,
        error: str | None,
        payload: dict[str, Any] | None,
        raw: bytes | None,
    ) -> ReplayTurn:
        return ReplayTurn(
            turn_number=turn.turn_number,
            workbook_row=turn.workbook_row,
            query=turn.query,
            reference_answer=turn.reference_answer,
            key_points=turn.key_points,
            status=status,
            http_status=http_status,
            elapsed_ms=max(0, round((monotonic() - started) * 1000)),
            error=error,
            response=payload,
            response_sha256=None if raw is None else _sha256_bytes(raw),
        )


def _failed_conversation(
    conversation: WorkbookConversation,
    *,
    error: str,
) -> ReplayConversation:
    return ReplayConversation(
        conversation_id=conversation.conversation_id,
        question_number=conversation.question_number,
        workbook_label=conversation.workbook_label,
        session_id=None,
        turns=tuple(
            ReplayTurn(
                turn_number=turn.turn_number,
                workbook_row=turn.workbook_row,
                query=turn.query,
                reference_answer=turn.reference_answer,
                key_points=turn.key_points,
                status="transport_error",
                http_status=None,
                elapsed_ms=0,
                error=error,
                response=None,
                response_sha256=None,
            )
            for turn in conversation.turns
        ),
    )


def _replay_conversation(
    conversation: WorkbookConversation,
    *,
    origin: str,
    expected_release_id: str,
    timeout_seconds: float,
) -> ReplayConversation:
    client = _ConversationClient(origin=origin, timeout_seconds=timeout_seconds)
    session_id, reset_error = client.reset()
    if reset_error is not None or session_id is None:
        return _failed_conversation(
            conversation,
            error=reset_error or "session reset failed",
        )
    return ReplayConversation(
        conversation_id=conversation.conversation_id,
        question_number=conversation.question_number,
        workbook_label=conversation.workbook_label,
        session_id=session_id,
        turns=tuple(
            client.chat(turn, expected_release_id=expected_release_id)
            for turn in conversation.turns
        ),
    )


def _replay_single_session(
    benchmark: WorkbookBenchmark,
    *,
    origin: str,
    expected_release_id: str,
    timeout_seconds: float,
) -> tuple[ReplayConversation, ...]:
    """Replay every conversation through ONE shared session, in workbook order.

    Cross-topic stress mode: turns from later conversations arrive after
    unrelated earlier topics, so referent binding and topic-switch handling
    are exercised exactly the way a real user mixes topics in one chat.
    """

    client = _ConversationClient(origin=origin, timeout_seconds=timeout_seconds)
    session_id, reset_error = client.reset()
    if reset_error is not None or session_id is None:
        return tuple(
            _failed_conversation(
                conversation,
                error=reset_error or "session reset failed",
            )
            for conversation in benchmark.conversations
        )
    return tuple(
        ReplayConversation(
            conversation_id=conversation.conversation_id,
            question_number=conversation.question_number,
            workbook_label=conversation.workbook_label,
            session_id=session_id,
            turns=tuple(
                client.chat(turn, expected_release_id=expected_release_id)
                for turn in conversation.turns
            ),
        )
        for conversation in benchmark.conversations
    )


def replay_benchmark(
    benchmark: WorkbookBenchmark,
    *,
    base_url: str,
    expected_release_id: str,
    workers: int,
    timeout_seconds: float,
    single_session: bool = False,
) -> BenchmarkReplay:
    """Run conversations concurrently while preserving strict turn order per session."""

    origin = _explicit_origin(base_url)
    if not expected_release_id.strip():
        raise WorkbookReplayError("expected release ID must be explicit")
    if not 1 <= workers <= _MAX_WORKERS:
        raise WorkbookReplayError(f"workers must be between 1 and {_MAX_WORKERS}")
    if not 0 < timeout_seconds <= 3600:
        raise WorkbookReplayError("timeout must be within (0, 3600] seconds")

    if single_session:
        conversations = _replay_single_session(
            benchmark,
            origin=origin,
            expected_release_id=expected_release_id,
            timeout_seconds=timeout_seconds,
        )
    else:
        def execute(conversation: WorkbookConversation) -> ReplayConversation:
            return _replay_conversation(
                conversation,
                origin=origin,
                expected_release_id=expected_release_id,
                timeout_seconds=timeout_seconds,
            )

        with ThreadPoolExecutor(
            max_workers=min(workers, len(benchmark.conversations)),
            thread_name_prefix="customer-workbook",
        ) as executor:
            conversations = tuple(executor.map(execute, benchmark.conversations))
    turns = tuple(turn for item in conversations for turn in item.turns)
    ok_count = sum(turn.status == "ok" for turn in turns)
    summary = {
        "conversation_count": len(conversations),
        "turn_count": len(turns),
        "ok_count": ok_count,
        "failure_count": len(turns) - ok_count,
    }
    generated_at = datetime.now(UTC).isoformat()
    payload: dict[str, Any] = {
        "schema_version": _SCHEMA_VERSION,
        "generated_at": generated_at,
        "base_url": origin,
        "expected_release_id": expected_release_id,
        "workbook_path": benchmark.workbook_path,
        "workbook_sha256": benchmark.workbook_sha256,
        "sheet_name": benchmark.sheet_name,
        "conversations": [asdict(item) for item in conversations],
        "summary": summary,
    }
    return BenchmarkReplay(
        schema_version=_SCHEMA_VERSION,
        generated_at=generated_at,
        base_url=origin,
        expected_release_id=expected_release_id,
        workbook_path=benchmark.workbook_path,
        workbook_sha256=benchmark.workbook_sha256,
        sheet_name=benchmark.sheet_name,
        conversations=conversations,
        summary=summary,
        content_sha256=_canonical_sha256(payload),
    )


def _quoted(value: str | None, *, empty: str = "未提供") -> str:
    text = empty if value is None or not value.strip() else value.strip()
    return "\n".join(f"> {line}" if line else ">" for line in text.splitlines())


def _display_text(value: object, *, fallback: str) -> str:
    return value.strip() if isinstance(value, str) and value.strip() else fallback


def _source_lines(response: dict[str, Any] | None) -> list[str]:
    if response is None:
        return ["- 无（请求未产生可验证响应）"]
    lines: list[str] = []
    citations = response.get("citations")
    if isinstance(citations, list):
        for item in citations[:30]:
            if not isinstance(item, dict):
                continue
            label = _display_text(item.get("label"), fallback="未命名来源")
            domain = _display_text(item.get("type"), fallback="unknown")
            identifier = _display_text(item.get("id"), fallback="unknown")
            lines.append(f"- 引用：{label}（{domain}，`{identifier}`）")
    evidence = response.get("evidence")
    if isinstance(evidence, list):
        for item in evidence[:20]:
            if not isinstance(item, dict):
                continue
            domain = _display_text(item.get("domain"), fallback="unknown")
            lane = _display_text(item.get("lane"), fallback="unknown")
            nature = _display_text(item.get("source_nature"), fallback="unknown")
            locator = _display_text(item.get("source_locator"), fallback="unknown")
            snippet = _display_text(item.get("snippet"), fallback="")
            if len(snippet) > 240:
                snippet = snippet[:237] + "..."
            suffix = f"；{snippet}" if snippet else ""
            lines.append(
                f"- 证据：{domain}/{lane}/{nature}，`{locator}`{suffix}"
            )
    return lines or ["- 响应未提供引用或证据"]


def _limitation_lines(response: dict[str, Any] | None) -> list[str]:
    if response is None:
        return ["- 无结构化限制（请求未产生可验证响应）"]
    structured = response.get("structured_payload")
    trace = structured.get("canonical_v2") if isinstance(structured, dict) else None
    limitations = trace.get("limitations") if isinstance(trace, dict) else None
    lines: list[str] = []
    if isinstance(limitations, list):
        for item in limitations:
            if not isinstance(item, dict):
                continue
            code = _display_text(item.get("code"), fallback="unknown")
            message = _display_text(
                item.get("message") or item.get("public_message") or item.get("rationale"),
                fallback="未提供说明",
            )
            lines.append(f"- `{code}`：{message}")
    return lines or ["- 无结构化限制"]


def render_markdown(replay: BenchmarkReplay) -> str:
    """Render review-first prose; the complete machine trace stays in the JSON companion."""

    summary = replay.summary
    lines = [
        "# 客户工作簿端到端回放报告",
        "",
        f"- Candidate：`{replay.expected_release_id}`",
        f"- 工作簿 SHA-256：`{replay.workbook_sha256}`",
        f"- 运行地址：`{replay.base_url}`",
        f"- 生成时间：`{replay.generated_at}`",
        f"- 会话/轮次：{summary['conversation_count']} / {summary['turn_count']}",
        f"- HTTP/契约成功：{summary['ok_count']}；失败：{summary['failure_count']}",
        "- 语义判断：未自动接受。请对照 Ground Truth、关键点、实际回答与来源进行人工判断。",
        "",
    ]
    for conversation in replay.conversations:
        lines.extend(
            [
                f"## {conversation.workbook_label}",
                "",
                f"会话：`{conversation.conversation_id}`；运行 session："
                f"`{conversation.session_id or '未建立'}`",
                "",
            ]
        )
        for turn in conversation.turns:
            response = turn.response
            answer_text = (
                response.get("answer_text") if isinstance(response, dict) else None
            )
            lines.extend(
                [
                    f"### 第 {turn.turn_number} 轮（Excel 第 {turn.workbook_row} 行）",
                    "",
                    f"- 执行状态：`{turn.status}`",
                    f"- HTTP：`{turn.http_status if turn.http_status is not None else '无'}`",
                    f"- 耗时：{turn.elapsed_ms} ms",
                    f"- 语义复核：{'先修复执行错误' if turn.status != 'ok' else '待用户按语义判断'}",
                    "",
                    "#### 问题",
                    "",
                    _quoted(turn.query),
                    "",
                    "#### Ground Truth（参考答案）",
                    "",
                    _quoted(turn.reference_answer),
                    "",
                    "#### 关键点",
                    "",
                    _quoted(turn.key_points),
                    "",
                    "#### 实际回答",
                    "",
                    _quoted(
                        answer_text if isinstance(answer_text, str) else None,
                        empty="无可读回答",
                    ),
                    "",
                    "#### 来源",
                    "",
                    *_source_lines(response),
                    "",
                    "#### 系统限制",
                    "",
                    *_limitation_lines(response),
                ]
            )
            if turn.error is not None:
                lines.extend(["", "#### 运行错误", "", f"- `{turn.error}`"])
            lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def _atomic_write(path: Path, content: bytes) -> None:
    output = path.resolve()
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_name(f".{output.name}.tmp-{os.getpid()}")
    try:
        with temporary.open("xb") as stream:
            stream.write(content)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary, output)
    finally:
        temporary.unlink(missing_ok=True)


def write_replay_outputs(
    replay: BenchmarkReplay,
    *,
    json_path: Path,
    markdown_path: Path,
) -> None:
    if json_path.resolve() == markdown_path.resolve():
        raise WorkbookReplayError("JSON and Markdown outputs must use distinct paths")
    document = replay.to_document()
    expected_hash = _canonical_sha256(
        {key: value for key, value in document.items() if key != "content_sha256"}
    )
    if replay.content_sha256 != expected_hash:
        raise WorkbookReplayError("replay content hash does not bind the complete result")
    _atomic_write(
        json_path,
        (
            json.dumps(
                document,
                ensure_ascii=False,
                sort_keys=True,
                indent=2,
                allow_nan=False,
            )
            + "\n"
        ).encode("utf-8"),
    )
    _atomic_write(markdown_path, render_markdown(replay).encode("utf-8"))


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Replay the customer workbook through one Canonical V2 Candidate origin."
    )
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--base-url", required=True)
    parser.add_argument("--expected-release-id", required=True)
    parser.add_argument("--json-output", type=Path, required=True)
    parser.add_argument("--markdown-output", type=Path, required=True)
    parser.add_argument(
        "--workers",
        type=int,
        default=min(17, os.cpu_count() or 1),
    )
    parser.add_argument("--timeout-seconds", type=float, default=900.0)
    parser.add_argument(
        "--single-session",
        action="store_true",
        help=(
            "replay every conversation through one shared session in workbook "
            "order (cross-topic stress mode) instead of one session per "
            "conversation"
        ),
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    try:
        benchmark = load_workbook_benchmark(args.workbook)
        replay = replay_benchmark(
            benchmark,
            base_url=args.base_url,
            expected_release_id=args.expected_release_id,
            workers=args.workers,
            timeout_seconds=args.timeout_seconds,
            single_session=args.single_session,
        )
        write_replay_outputs(
            replay,
            json_path=args.json_output,
            markdown_path=args.markdown_output,
        )
    except WorkbookReplayError as exc:
        raise SystemExit(f"customer workbook replay rejected: {exc}") from None
    print(json.dumps(replay.summary, ensure_ascii=False, sort_keys=True))
    return 0 if replay.summary["failure_count"] == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
