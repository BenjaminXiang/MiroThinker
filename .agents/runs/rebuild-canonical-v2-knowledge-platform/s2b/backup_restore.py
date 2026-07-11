#!/usr/bin/env python3
"""Content-addressed backup, restore, and fail-closed S2B admission helpers."""

from __future__ import annotations

import argparse
import gzip
import hashlib
import json
import os
from pathlib import Path
import shutil
import sqlite3
from typing import Any, Iterable
import uuid


FAMILY_PATTERNS = {
    "professor_fetch_cache_family": "logs/debug/professor_fetch_cache/*.json",
    "paper_openalex_cache_family": "logs/debug/paper_openalex_cache/*.json",
    "paper_orcid_cache_family": "logs/debug/paper_orcid_cache/*.json",
    "legacy_release_jsonl_family": "logs/debug/*_release_e2e_*/*.jsonl",
    "legacy_sqlite_snapshot_family": "logs/data_agents/**/*.db",
    "legacy_data_agent_jsonl_family": "logs/data_agents/**/*.jsonl",
    "admin_upload_workbook_family": "data/admin_uploads/**/*.xlsx",
    "admin_upload_jsonl_family": "data/admin_uploads/**/*.jsonl",
    "compressed_backup_family": "backups/*.csv.gz",
    "raw_pdf_family": "apps/*/logs/raw_pdfs/**/*.pdf",
    "historical_milvus_file_family": "logs/**/*milvus*.db",
}
REQUIRED_EXTRA_SOURCE_IDS = {
    "original_postgresql_volume",
    "forensic_recovery_tree",
}
RECOVERY_TREE_EXCLUDES = {"lab-01/cluster-current"}


class BackupGateError(RuntimeError):
    """Raised before a backup, restore, or rebuild write when evidence is unsafe."""


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def serialized_document(document: dict[str, Any]) -> bytes:
    return (
        json.dumps(document, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    ).encode()


def document_sha256(document: dict[str, Any]) -> str:
    return sha256_bytes(serialized_document(document))


def _contains(parent: Path, child: Path) -> bool:
    try:
        child.relative_to(parent)
    except ValueError:
        return False
    return True


def validate_target_roots(
    *,
    source_roots: Iterable[Path],
    backup_root: Path,
    restore_root: Path,
) -> dict[str, str]:
    backup = backup_root.resolve(strict=False)
    restore = restore_root.resolve(strict=False)
    if backup == restore or _contains(backup, restore) or _contains(restore, backup):
        raise BackupGateError("backup and restore roots must be distinct and non-overlapping")

    for raw_source in source_roots:
        source = raw_source.resolve(strict=False)
        if _contains(source, backup) or _contains(backup, source):
            raise BackupGateError(f"backup root is inside source or overlaps it: {source}")
        if _contains(source, restore) or _contains(restore, source):
            raise BackupGateError(f"restore root is inside source or overlaps it: {source}")
    return {"backup_root": str(backup), "restore_root": str(restore)}


def validate_container_mount_policy(
    inspection: dict[str, Any],
    *,
    readonly_destinations: set[str],
    writable_destinations: set[str],
    tmpfs_destinations: set[str],
) -> dict[str, int]:
    expected_persistent = readonly_destinations | writable_destinations
    seen_persistent: set[str] = set()
    for mount in inspection.get("Mounts", []):
        destination = mount.get("Destination")
        if not isinstance(destination, str):
            raise BackupGateError("persistent mount is missing a destination")
        seen_persistent.add(destination)
        expected_rw = destination in writable_destinations
        expected_ro = destination in readonly_destinations
        if not expected_rw and not expected_ro:
            raise BackupGateError(
                f"unexpected persistent mount {destination} ({mount.get('Type')})"
            )
        if bool(mount.get("RW")) != expected_rw:
            expected = "writable" if expected_rw else "read-only"
            raise BackupGateError(f"persistent mount {destination} is not {expected}")
    missing = sorted(expected_persistent - seen_persistent)
    if missing:
        raise BackupGateError(f"missing required persistent mounts: {missing}")

    configured_tmpfs = (inspection.get("HostConfig") or {}).get("Tmpfs") or {}
    seen_tmpfs = set(configured_tmpfs)
    if seen_tmpfs != tmpfs_destinations:
        raise BackupGateError(
            "tmpfs mount mismatch: "
            f"expected={sorted(tmpfs_destinations)}, actual={sorted(seen_tmpfs)}"
        )
    return {
        "persistent_mounts": len(seen_persistent),
        "tmpfs_mounts": len(seen_tmpfs),
    }


def _copy_bytes(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_name(f".{destination.name}.{uuid.uuid4().hex}.part")
    try:
        shutil.copyfile(source, temporary, follow_symlinks=False)
        with temporary.open("rb") as handle:
            os.fsync(handle.fileno())
        os.replace(temporary, destination)
    finally:
        temporary.unlink(missing_ok=True)


def _independent(left: Path, right: Path) -> bool:
    left_stat = left.stat()
    right_stat = right.stat()
    return (left_stat.st_dev, left_stat.st_ino) != (
        right_stat.st_dev,
        right_stat.st_ino,
    )


def copy_file_to_cas(source: Path, backup_root: Path) -> dict[str, Any]:
    source = source.resolve(strict=True)
    if not source.is_file():
        raise BackupGateError(f"source is not a regular file: {source}")
    source_sha256 = sha256_file(source)
    object_path = Path("objects") / "sha256" / source_sha256[:2] / source_sha256
    destination = backup_root.resolve(strict=False) / object_path
    if not destination.exists():
        _copy_bytes(source, destination)
        destination.chmod(0o440)
    backup_sha256 = sha256_file(destination)
    if backup_sha256 != source_sha256 or destination.stat().st_size != source.stat().st_size:
        raise BackupGateError(f"backup hash/size mismatch for {source}")
    independent = _independent(source, destination)
    if not independent:
        raise BackupGateError(f"backup unexpectedly shares source inode: {source}")
    return {
        "source_path": str(source),
        "source_bytes": source.stat().st_size,
        "source_sha256": source_sha256,
        "object_path": str(object_path),
        "backup_bytes": destination.stat().st_size,
        "backup_sha256": backup_sha256,
        "copy_independent": independent,
    }


def _safe_destination(root: Path, relative_path: Path) -> Path:
    if relative_path.is_absolute() or ".." in relative_path.parts:
        raise BackupGateError(f"unsafe restore relative path: {relative_path}")
    resolved_root = root.resolve(strict=False)
    destination = (resolved_root / relative_path).resolve(strict=False)
    if not _contains(resolved_root, destination):
        raise BackupGateError(f"restore path escapes target root: {relative_path}")
    return destination


def materialize_cas_object(
    copy_record: dict[str, Any],
    *,
    backup_root: Path,
    restore_root: Path,
    relative_path: Path,
) -> dict[str, Any]:
    try:
        backup_path = _safe_destination(
            backup_root, Path(copy_record["object_path"])
        )
    except BackupGateError as exc:
        raise BackupGateError(
            f"unsafe backup object path: {copy_record['object_path']}"
        ) from exc
    if not backup_path.is_file():
        raise BackupGateError(f"backup object is missing: {backup_path}")
    destination = _safe_destination(restore_root, relative_path)
    _copy_bytes(backup_path, destination)
    destination.chmod(0o440)
    restore_sha256 = sha256_file(destination)
    if restore_sha256 != copy_record["backup_sha256"]:
        raise BackupGateError(f"restore hash mismatch for {relative_path}")
    independent = _independent(backup_path, destination)
    if not independent:
        raise BackupGateError(f"restore unexpectedly shares backup inode: {relative_path}")
    return {
        "restore_path": str(destination.relative_to(restore_root.resolve(strict=False))),
        "restore_bytes": destination.stat().st_size,
        "restore_sha256": restore_sha256,
        "copy_independent": independent,
    }


def build_member_manifest(
    paths: Iterable[Path], root: Path
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    resolved_root = root.resolve(strict=True)
    rows = []
    for path in sorted(paths, key=lambda item: str(item.relative_to(resolved_root))):
        relative = str(path.relative_to(resolved_root))
        rows.append(
            {
                "relative_path": relative,
                "bytes": path.stat().st_size,
                "sha256": sha256_file(path),
            }
        )
    legacy = "".join(
        f"{row['relative_path']}|{row['bytes']}|{row['sha256']}\n" for row in rows
    )
    return (
        {
            "files": len(rows),
            "bytes": sum(row["bytes"] for row in rows),
            "legacy_manifest_sha256": sha256_bytes(legacy.encode()),
        },
        rows,
    )


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.part")
    try:
        with temporary.open("w", encoding="utf-8") as handle:
            for row in rows:
                handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows = []
    with path.open("r", encoding="utf-8") as handle:
        for line_number, line in enumerate(handle, start=1):
            if not line.strip():
                continue
            value = json.loads(line)
            if not isinstance(value, dict):
                raise BackupGateError(f"non-object JSONL row at {path}:{line_number}")
            rows.append(value)
    return rows


def _inventory_source_members(
    source: dict[str, Any],
    *,
    evidence_root: Path,
    recovery_root: Path,
) -> tuple[Path, str, list[Path]]:
    kind = source["kind"]
    if "root" in source:
        pattern = FAMILY_PATTERNS.get(kind)
        if pattern is None:
            raise BackupGateError(f"no frozen family pattern for {kind}")
        root = Path(source["root"]).resolve(strict=True)
        if root != evidence_root.resolve(strict=True):
            raise BackupGateError(f"family root changed for {kind}: {root}")
        members = sorted(path for path in root.glob(pattern) if path.is_file())
        return root, "workspace", members

    if source.get("authority") == "forensic_recovery_evidence":
        root = recovery_root.resolve(strict=True)
        namespace = "recovery"
    else:
        root = evidence_root.resolve(strict=True)
        namespace = "workspace"
    member = _safe_destination(root, Path(source["path"]))
    if not member.is_file():
        raise BackupGateError(f"inventoried source is missing: {member}")
    return root, namespace, [member]


def _verify_inventory_record(
    source: dict[str, Any],
    *,
    root: Path,
    members: list[Path],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    summary, rows = build_member_manifest(members, root)
    if "root" in source:
        expected = (
            source["files"],
            source["bytes"],
            source["manifest_sha256"],
        )
        actual = (
            summary["files"],
            summary["bytes"],
            summary["legacy_manifest_sha256"],
        )
        if actual != expected:
            raise BackupGateError(
                f"frozen family inventory mismatch for {source['kind']}: "
                f"expected={expected}, actual={actual}"
            )
    else:
        if len(rows) != 1:
            raise BackupGateError(f"individual inventory record expanded: {source['kind']}")
        row = rows[0]
        if row["bytes"] != source["bytes"] or row["sha256"] != source["sha256"]:
            raise BackupGateError(f"frozen source hash/size mismatch for {source['kind']}")
    return summary, rows


def backup_inventory_sources(
    *,
    inventory: dict[str, Any],
    evidence_root: Path,
    recovery_root: Path,
    backup_root: Path,
    restore_root: Path,
    run_id: str,
    copied_at: str,
) -> dict[str, Any]:
    target_roots = validate_target_roots(
        source_roots=[evidence_root, recovery_root],
        backup_root=backup_root,
        restore_root=restore_root,
    )
    backup_root = backup_root.resolve(strict=False)
    backup_root.mkdir(parents=True, mode=0o700, exist_ok=False)
    backup_root.chmod(0o700)
    entries = []
    for index, source in enumerate(inventory["sources"]):
        source_id = source_record_id(source)
        root, namespace, members = _inventory_source_members(
            source,
            evidence_root=evidence_root,
            recovery_root=recovery_root,
        )
        summary, expected_rows = _verify_inventory_record(
            source, root=root, members=members
        )
        backup_rows = []
        for path, expected in zip(members, expected_rows, strict=True):
            copy = copy_file_to_cas(path, backup_root)
            if (
                copy["source_bytes"] != expected["bytes"]
                or copy["source_sha256"] != expected["sha256"]
            ):
                raise BackupGateError(f"source changed during copy: {path}")
            backup_rows.append(
                {
                    "namespace": namespace,
                    "relative_path": expected["relative_path"],
                    **copy,
                }
            )
        member_manifest = (
            Path("manifests")
            / "inventory"
            / f"{index:03d}-{source_id.removeprefix('inventory:')}.jsonl"
        )
        member_manifest_path = backup_root / member_manifest
        _write_jsonl(member_manifest_path, backup_rows)
        entries.append(
            {
                "source_id": source_id,
                "kind": source["kind"],
                "authority": source.get("authority"),
                "source_root": str(root),
                "member_count": len(backup_rows),
                "source_bytes": summary["bytes"],
                "source_member_manifest_sha256": summary[
                    "legacy_manifest_sha256"
                ],
                "backup_member_manifest_path": str(member_manifest),
                "backup_member_manifest_sha256": sha256_file(member_manifest_path),
                "copy_independent": all(
                    row["copy_independent"] for row in backup_rows
                ),
                "hash_verified": all(
                    row["source_sha256"] == row["backup_sha256"]
                    for row in backup_rows
                ),
            }
        )
    return {
        "schema_version": "canonical-v2-s2b-inventory-backup-v1",
        "run_id": run_id,
        "copied_at": copied_at,
        "backup_root": str(backup_root),
        "restore_root": target_roots["restore_root"],
        "sources": entries,
    }


def _sample_rows(rows: list[dict[str, Any]], maximum: int = 5) -> list[dict[str, Any]]:
    if len(rows) <= maximum:
        return rows
    return rows[:3] + rows[-2:]


def _materialize_or_verify(
    row: dict[str, Any], *, backup_root: Path, restore_root: Path
) -> dict[str, Any]:
    relative = Path(row["namespace"]) / row["relative_path"]
    destination = _safe_destination(restore_root, relative)
    backup_path = backup_root / row["object_path"]
    if destination.exists():
        restore_sha256 = sha256_file(destination)
        if restore_sha256 != row["backup_sha256"]:
            raise BackupGateError(f"existing restore hash mismatch for {relative}")
        return {
            "restore_path": str(relative),
            "restore_bytes": destination.stat().st_size,
            "restore_sha256": restore_sha256,
            "copy_independent": _independent(backup_path, destination),
        }
    return materialize_cas_object(
        row,
        backup_root=backup_root,
        restore_root=restore_root,
        relative_path=relative,
    )


def restore_inventory_sources(
    *,
    backup_manifest: dict[str, Any],
    backup_root: Path,
    restore_root: Path,
    verified_at: str,
) -> dict[str, Any]:
    backup_root = backup_root.resolve(strict=True)
    restore_root = restore_root.resolve(strict=False)
    validate_target_roots(
        source_roots=[],
        backup_root=backup_root,
        restore_root=restore_root,
    )
    restore_root.mkdir(parents=True, mode=0o700, exist_ok=False)
    restore_root.chmod(0o700)
    entries = []
    for source in backup_manifest["sources"]:
        member_manifest = backup_root / source["backup_member_manifest_path"]
        if sha256_file(member_manifest) != source["backup_member_manifest_sha256"]:
            raise BackupGateError(
                f"backup member manifest hash mismatch for {source['source_id']}"
            )
        rows = _read_jsonl(member_manifest)
        restored_rows = []
        for row in rows:
            restored_rows.append(
                {
                    **row,
                    **_materialize_or_verify(
                        row,
                        backup_root=backup_root,
                        restore_root=restore_root,
                    ),
                }
            )
        hash_verified = all(
            row["restore_sha256"] == row["backup_sha256"] for row in restored_rows
        )
        copy_independent = all(
            row["copy_independent"] for row in restored_rows
        )
        probes = []
        for row in _sample_rows(restored_rows):
            path = restore_root / row["restore_path"]
            probes.append(
                {
                    "restore_path": row["restore_path"],
                    **probe_materialized_file(path),
                }
            )
        probes_passed = all(probe["status"] == "passed" for probe in probes)
        entries.append(
            {
                "source_id": source["source_id"],
                "kind": source["kind"],
                "member_count": len(restored_rows),
                "restore_bytes": sum(row["restore_bytes"] for row in restored_rows),
                "hash_verified": hash_verified,
                "copy_independent": copy_independent,
                "probes": probes,
                "status": (
                    "passed"
                    if hash_verified and copy_independent and probes_passed
                    else "failed"
                ),
            }
        )
    return {
        "schema_version": "canonical-v2-s2b-inventory-restore-v1",
        "run_id": backup_manifest["run_id"],
        "verified_at": verified_at,
        "backup_root": str(backup_root),
        "restore_root": str(restore_root),
        "sources": entries,
    }


def register_archive_backup(
    *,
    source_id: str,
    kind: str,
    source_identity: str,
    source_root: str,
    archive: Path,
    source_tree_manifest: Path,
    backup_root: Path,
    details: dict[str, Any],
) -> dict[str, Any]:
    archive_copy = copy_file_to_cas(archive, backup_root)
    manifest_copy = copy_file_to_cas(source_tree_manifest, backup_root)
    entries = sum(
        1
        for line in source_tree_manifest.read_text(
            encoding="utf-8", errors="strict"
        ).splitlines()
        if line
    )
    return {
        **details,
        "source_id": source_id,
        "kind": kind,
        "source_identity": source_identity,
        "source_root": source_root,
        "source_tree_entries": entries,
        "source_tree_manifest_sha256": manifest_copy["source_sha256"],
        "backup_tree_manifest_object_path": manifest_copy["object_path"],
        "backup_archive_object_path": archive_copy["object_path"],
        "backup_archive_bytes": archive_copy["backup_bytes"],
        "backup_archive_sha256": archive_copy["backup_sha256"],
        "copy_independent": (
            archive_copy["copy_independent"]
            and manifest_copy["copy_independent"]
        ),
        "hash_verified": (
            archive_copy["source_sha256"] == archive_copy["backup_sha256"]
            and manifest_copy["source_sha256"] == manifest_copy["backup_sha256"]
        ),
    }


def _json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    return str(value)


def probe_milvus_copy(
    path: Path,
    *,
    forbidden_paths: set[Path],
    expected_sha256: str,
    client_factory: Any | None = None,
) -> dict[str, Any]:
    resolved = path.resolve(strict=True)
    resolved_stat = resolved.stat()
    for forbidden_path in forbidden_paths:
        forbidden = forbidden_path.resolve(strict=True)
        forbidden_stat = forbidden.stat()
        if resolved == forbidden or (
            resolved_stat.st_dev,
            resolved_stat.st_ino,
        ) == (forbidden_stat.st_dev, forbidden_stat.st_ino):
            raise BackupGateError(f"refusing forbidden Milvus source path: {forbidden}")
    before_sha256 = sha256_file(resolved)
    if before_sha256 != expected_sha256:
        raise BackupGateError(
            f"Milvus probe copy hash mismatch: expected {expected_sha256}, got {before_sha256}"
        )
    if client_factory is None:
        from pymilvus import MilvusClient

        client_factory = MilvusClient
    client = client_factory(str(resolved))
    collections = []
    try:
        for name in sorted(client.list_collections()):
            description = client.describe_collection(name)
            stats = client.get_collection_stats(name)
            row_count = int(stats.get("row_count", 0))
            collections.append(
                {
                    "name": name,
                    "row_count": row_count,
                    "description": _json_safe(description),
                }
            )
    finally:
        client.close()
    after_sha256 = sha256_file(resolved)
    return {
        "path": str(resolved),
        "before_sha256": before_sha256,
        "after_sha256": after_sha256,
        "probe_mutated_copy": before_sha256 != after_sha256,
        "collection_count": len(collections),
        "total_rows": sum(item["row_count"] for item in collections),
        "collections": collections,
    }


def source_record_id(source: dict[str, Any]) -> str:
    compact = json.dumps(
        source, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode()
    return f"inventory:{sha256_bytes(compact)}"


def expected_source_ids(inventory: dict[str, Any]) -> set[str]:
    return {source_record_id(source) for source in inventory["sources"]}


def _indexed_sources(document: dict[str, Any], *, label: str) -> dict[str, dict[str, Any]]:
    records: dict[str, dict[str, Any]] = {}
    for item in document.get("sources", []):
        source_id = item.get("source_id")
        if not isinstance(source_id, str) or not source_id:
            raise BackupGateError(f"{label} contains a source without source_id")
        if source_id in records:
            raise BackupGateError(f"{label} contains duplicate source_id {source_id}")
        records[source_id] = item
    return records


def require_accepted_backup_gate(
    *,
    inventory: dict[str, Any],
    backup_manifest: dict[str, Any],
    restore_verification: dict[str, Any],
    acceptance_record: dict[str, Any],
) -> dict[str, Any]:
    validate_target_roots(
        source_roots=[],
        backup_root=Path(backup_manifest["backup_root"]),
        restore_root=Path(backup_manifest["restore_root"]),
    )
    expected = expected_source_ids(inventory) | REQUIRED_EXTRA_SOURCE_IDS
    backups = _indexed_sources(backup_manifest, label="backup manifest")
    missing_backups = sorted(expected - set(backups))
    extra_backups = sorted(set(backups) - expected)
    if missing_backups:
        raise BackupGateError(f"missing backup source records: {missing_backups}")
    if extra_backups:
        raise BackupGateError(f"unexpected backup source records: {extra_backups}")
    if backup_manifest.get("backup_root") == backup_manifest.get("restore_root"):
        raise BackupGateError("backup and restore roots must be distinct")
    for source_id, record in backups.items():
        if record.get("copy_independent") is not True:
            raise BackupGateError(f"copy independence failed for {source_id}")
        if record.get("hash_verified") is not True:
            raise BackupGateError(f"backup hash verification failed for {source_id}")

    restores = _indexed_sources(restore_verification, label="restore verification")
    missing_restores = sorted(expected - set(restores))
    extra_restores = sorted(set(restores) - expected)
    if missing_restores or extra_restores:
        raise BackupGateError(
            "restore verification coverage mismatch: "
            f"missing={missing_restores}, extra={extra_restores}"
        )
    for source_id, record in restores.items():
        if record.get("status") != "passed":
            raise BackupGateError(f"restore verification failed for {source_id}")

    if restore_verification.get("backup_manifest_sha256") != document_sha256(
        backup_manifest
    ):
        raise BackupGateError(
            "restore verification does not reference the exact backup manifest"
        )

    required_probe_ids = {"forensic", "milvus", "postgresql"}
    probes = restore_verification.get("required_probes") or {}
    missing_probes = sorted(required_probe_ids - set(probes))
    failed_probes = sorted(
        probe_id
        for probe_id in required_probe_ids & set(probes)
        if probes[probe_id].get("status") != "passed"
    )
    if missing_probes or failed_probes:
        raise BackupGateError(
            "required restore probes are incomplete: "
            f"missing={missing_probes}, failed={failed_probes}"
        )

    if acceptance_record.get("state") != "accepted":
        raise BackupGateError("backup/restore acceptance record is not accepted")
    if acceptance_record.get("backup_manifest_sha256") != document_sha256(
        backup_manifest
    ):
        raise BackupGateError("acceptance does not match the exact backup manifest")
    if acceptance_record.get("restore_verification_sha256") != document_sha256(
        restore_verification
    ):
        raise BackupGateError("acceptance does not match the exact restore verification")
    return {
        "state": "accepted",
        "source_count": len(expected),
        "backup_manifest_sha256": document_sha256(backup_manifest),
        "restore_verification_sha256": document_sha256(restore_verification),
    }


def probe_materialized_file(path: Path) -> dict[str, Any]:
    suffixes = path.suffixes
    try:
        if path.suffix == ".jsonl":
            records = 0
            with path.open("r", encoding="utf-8") as handle:
                for line in handle:
                    if line.strip():
                        json.loads(line)
                        records += 1
            return {"status": "passed", "probe": "jsonl_parse", "records": records}
        if path.suffix == ".json":
            json.loads(path.read_text(encoding="utf-8"))
            return {"status": "passed", "probe": "json_parse"}
        if path.suffix in {".db", ".sqlite", ".sqlite3"}:
            uri = f"file:{path.resolve()}?mode=ro&immutable=1"
            with sqlite3.connect(uri, uri=True) as connection:
                result = connection.execute("PRAGMA quick_check").fetchone()[0]
            if result != "ok":
                raise ValueError(f"SQLite quick_check returned {result}")
            return {"status": "passed", "probe": "sqlite_quick_check"}
        if path.suffix == ".xlsx":
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                sheets = len(workbook.worksheets)
            finally:
                workbook.close()
            return {"status": "passed", "probe": "xlsx_read_only", "sheets": sheets}
        if path.suffix == ".pdf":
            if not path.read_bytes()[:5].startswith(b"%PDF-"):
                raise ValueError("missing PDF header")
            return {"status": "passed", "probe": "pdf_header"}
        if suffixes[-2:] == [".csv", ".gz"] or path.suffix == ".gz":
            with gzip.open(path, "rb") as handle:
                while handle.read(1024 * 1024):
                    pass
            return {"status": "passed", "probe": "gzip_stream"}
        with path.open("rb") as handle:
            handle.read(4096)
        return {"status": "passed", "probe": "bounded_read"}
    except Exception as exc:
        return {
            "status": "failed",
            "probe": "format_readability",
            "error": f"{type(exc).__name__}: {exc}",
        }


def _read_document(path: Path) -> dict[str, Any]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(value, dict):
        raise BackupGateError(f"expected JSON object: {path}")
    return value


def _write_document(path: Path, document: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(f".{path.name}.{uuid.uuid4().hex}.part")
    try:
        with temporary.open("wb") as handle:
            handle.write(serialized_document(document))
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)

    backup = subparsers.add_parser("backup-inventory")
    backup.add_argument("--inventory", required=True, type=Path)
    backup.add_argument("--evidence-root", required=True, type=Path)
    backup.add_argument("--recovery-root", required=True, type=Path)
    backup.add_argument("--backup-root", required=True, type=Path)
    backup.add_argument("--restore-root", required=True, type=Path)
    backup.add_argument("--run-id", required=True)
    backup.add_argument("--copied-at", required=True)
    backup.add_argument("--output", required=True, type=Path)

    restore = subparsers.add_parser("restore-inventory")
    restore.add_argument("--backup-manifest", required=True, type=Path)
    restore.add_argument("--backup-root", required=True, type=Path)
    restore.add_argument("--restore-root", required=True, type=Path)
    restore.add_argument("--verified-at", required=True)
    restore.add_argument("--output", required=True, type=Path)

    archive = subparsers.add_parser("register-archive")
    archive.add_argument("--source-id", required=True)
    archive.add_argument("--kind", required=True)
    archive.add_argument("--source-identity", required=True)
    archive.add_argument("--source-root", required=True)
    archive.add_argument("--archive", required=True, type=Path)
    archive.add_argument("--source-tree-manifest", required=True, type=Path)
    archive.add_argument("--backup-root", required=True, type=Path)
    archive.add_argument("--details-json", type=Path)
    archive.add_argument("--output", required=True, type=Path)

    milvus = subparsers.add_parser("probe-milvus")
    milvus.add_argument("--copy", required=True, type=Path)
    milvus.add_argument("--forbidden", action="append", required=True, type=Path)
    milvus.add_argument("--expected-sha256", required=True)
    milvus.add_argument("--output", required=True, type=Path)

    assemble = subparsers.add_parser("assemble")
    assemble.add_argument("--base", required=True, type=Path)
    assemble.add_argument("--extra", action="append", default=[], type=Path)
    assemble.add_argument("--metadata", required=True, type=Path)
    assemble.add_argument("--output", required=True, type=Path)

    acceptance = subparsers.add_parser("create-acceptance")
    acceptance.add_argument("--backup-manifest", required=True, type=Path)
    acceptance.add_argument("--restore-verification", required=True, type=Path)
    acceptance.add_argument("--accepted-on", required=True)
    acceptance.add_argument("--authority", required=True)
    acceptance.add_argument("--statement", required=True)
    acceptance.add_argument("--output", required=True, type=Path)

    gate = subparsers.add_parser("verify-gate")
    gate.add_argument("--inventory", required=True, type=Path)
    gate.add_argument("--backup-manifest", required=True, type=Path)
    gate.add_argument("--restore-verification", required=True, type=Path)
    gate.add_argument("--acceptance-record", required=True, type=Path)
    return parser


def main() -> None:
    parser = _build_parser()
    args = parser.parse_args()
    if args.command == "backup-inventory":
        document = backup_inventory_sources(
            inventory=_read_document(args.inventory),
            evidence_root=args.evidence_root,
            recovery_root=args.recovery_root,
            backup_root=args.backup_root,
            restore_root=args.restore_root,
            run_id=args.run_id,
            copied_at=args.copied_at,
        )
        _write_document(args.output, document)
        print(json.dumps({"sources": len(document["sources"]), "output": str(args.output)}))
        return
    if args.command == "restore-inventory":
        document = restore_inventory_sources(
            backup_manifest=_read_document(args.backup_manifest),
            backup_root=args.backup_root,
            restore_root=args.restore_root,
            verified_at=args.verified_at,
        )
        _write_document(args.output, document)
        print(json.dumps({"sources": len(document["sources"]), "output": str(args.output)}))
        return
    if args.command == "register-archive":
        details = _read_document(args.details_json) if args.details_json else {}
        document = register_archive_backup(
            source_id=args.source_id,
            kind=args.kind,
            source_identity=args.source_identity,
            source_root=args.source_root,
            archive=args.archive,
            source_tree_manifest=args.source_tree_manifest,
            backup_root=args.backup_root,
            details=details,
        )
        _write_document(args.output, document)
        print(json.dumps({"source_id": document["source_id"], "output": str(args.output)}))
        return
    if args.command == "probe-milvus":
        document = probe_milvus_copy(
            args.copy,
            forbidden_paths=set(args.forbidden),
            expected_sha256=args.expected_sha256,
        )
        _write_document(args.output, document)
        print(
            json.dumps(
                {
                    "collections": document["collection_count"],
                    "rows": document["total_rows"],
                    "output": str(args.output),
                }
            )
        )
        return
    if args.command == "assemble":
        base = _read_document(args.base)
        metadata = _read_document(args.metadata)
        extras = [_read_document(path) for path in args.extra]
        document = {**base, **metadata, "sources": [*base["sources"], *extras]}
        _write_document(args.output, document)
        print(json.dumps({"sources": len(document["sources"]), "output": str(args.output)}))
        return
    if args.command == "create-acceptance":
        backup_manifest = _read_document(args.backup_manifest)
        restore_verification = _read_document(args.restore_verification)
        document = {
            "schema_version": "canonical-v2-s2b-acceptance-v1",
            "state": "accepted",
            "accepted_on": args.accepted_on,
            "authority": args.authority,
            "statement": args.statement,
            "backup_manifest_sha256": document_sha256(backup_manifest),
            "restore_verification_sha256": document_sha256(restore_verification),
        }
        _write_document(args.output, document)
        print(json.dumps({"state": document["state"], "output": str(args.output)}))
        return
    if args.command == "verify-gate":
        result = require_accepted_backup_gate(
            inventory=_read_document(args.inventory),
            backup_manifest=_read_document(args.backup_manifest),
            restore_verification=_read_document(args.restore_verification),
            acceptance_record=_read_document(args.acceptance_record),
        )
        print(json.dumps(result, sort_keys=True))
        return
    parser.error(f"unsupported command: {args.command}")


if __name__ == "__main__":
    main()
